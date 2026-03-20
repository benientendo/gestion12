# views_commercant.py
# Vues pour l'interface commerçant multi-boutiques

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum, Count, F, Avg, Max, Min, Prefetch
from django.db import transaction
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.core.cache import cache
from decimal import Decimal
from datetime import datetime, timedelta
import json
import csv
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from .models import Commercant, Boutique, Article, Vente, LigneVente, MouvementStock, Client, RapportCaisse, ArticleNegocie, RetourArticle, VenteRejetee, TransfertStock, VarianteArticle, Fournisseur, FactureApprovisionnement, LigneApprovisionnement, Categorie, Inventaire, LigneInventaire, AlerteStock
from .forms import BoutiqueForm, ArticleForm, VarianteArticleForm
import json
import qrcode
from PIL import Image
import io

# ... (rest of the code remains the same)
import uuid
from django.core.files.base import ContentFile

# ===== DÉCORATEURS ET UTILITAIRES =====

def generer_qr_code_article(article):
    """Génère un code QR pour un article et l'enregistre"""
    try:
        # Créer les données du QR code avec toutes les informations de l'article
        qr_data = {
            'id': article.id,
            'code': article.code,
            'nom': article.nom,
            'prix_vente': float(article.prix_vente),
            'boutique_id': article.boutique.id if article.boutique else None,
            'boutique_nom': article.boutique.nom if article.boutique else None
        }
        
        # Convertir en JSON pour le QR code
        qr_content = json.dumps(qr_data, ensure_ascii=False)
        
        # Créer le QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_content)
        qr.make(fit=True)
        
        # Créer l'image
        img = qr.make_image(fill_color="black", back_color="white")
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        # Créer le nom du fichier
        filename = f"qr_{article.code}_{article.id}.png"
        
        # Sauvegarder dans le champ qr_code de l'article
        article.qr_code.save(
            filename,
            ContentFile(buffer.read()),
            save=True
        )
        
        return True
        
    except Exception as e:
        print(f"Erreur lors de la génération du QR code pour l'article {article.id}: {str(e)}")
        return False

def commercant_required(view_func):
    """Décorateur pour vérifier que l'utilisateur est un commerçant ou un collaborateur actif"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('inventory:login_commercant')
        
        try:
            # Essayer d'abord comme commerçant
            commercant = request.user.profil_commercant
            if not commercant.est_actif:
                messages.error(request, "Votre compte commerçant est désactivé.")
                return redirect('inventory:login_commercant')
            return view_func(request, *args, **kwargs)
        except Commercant.DoesNotExist:
            # Si pas commerçant, vérifier si c'est un collaborateur
            try:
                from .models import Collaborateur
                collaborateur = Collaborateur.objects.get(user=request.user, est_actif=True)
                # Les collaborateurs ont accès
                return view_func(request, *args, **kwargs)
            except Collaborateur.DoesNotExist:
                messages.error(request, "Vous n'avez pas de profil commerçant ou collaborateur.")
                return redirect('inventory:login_commercant')
    return wrapper

def boutique_access_required(view_func):
    """Décorateur pour vérifier l'accès à une boutique spécifique (commerçant ou collaborateur)"""
    def wrapper(request, boutique_id, *args, **kwargs):
        try:
            # Essayer d'abord comme commerçant
            commercant = request.user.profil_commercant
            boutique = get_object_or_404(Boutique, id=boutique_id, commercant=commercant)
            request.boutique = boutique
            return view_func(request, boutique_id, *args, **kwargs)
        except Commercant.DoesNotExist:
            # Si pas commerçant, essayer comme collaborateur
            try:
                from .models import Collaborateur
                collaborateur = Collaborateur.objects.get(user=request.user, est_actif=True)
                boutique = get_object_or_404(Boutique, id=boutique_id, commercant=collaborateur.commercant)
                request.boutique = boutique
                request.collaborateur = collaborateur  # Ajouter le collaborateur à la requête
                return view_func(request, boutique_id, *args, **kwargs)
            except Collaborateur.DoesNotExist:
                return HttpResponseForbidden("Accès non autorisé")
    return wrapper

# ===== AUTHENTIFICATION =====

def login_commercant(request):
    """Page de connexion pour les commerçants"""
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            try:
                commercant = user.profil_commercant
                if commercant.est_actif:
                    login(request, user)
                    return redirect('inventory:commercant_dashboard')
                else:
                    messages.error(request, "Votre compte commerçant est désactivé. Veuillez contacter l'administrateur pour réactiver votre compte.")
            except Commercant.DoesNotExist:
                messages.error(request, "Vous n'avez pas de profil commerçant. Veuillez contacter l'administrateur.")
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect.")
    
    return render(request, 'inventory/commercant/login.html')

@login_required
def logout_commercant(request):
    """Déconnexion du commerçant"""
    logout(request)
    messages.success(request, "Vous avez été déconnecté avec succès.")
    return redirect('inventory:login_commercant')

# ===== TABLEAU DE BORD COMMERÇANT =====

@login_required
@commercant_required
def dashboard_commercant(request):
    """Tableau de bord principal du commerçant"""
    commercant = request.user.profil_commercant
    
    # Statistiques générales - Requêtes optimisées
    boutiques_toutes = commercant.boutiques.select_related('commercant').prefetch_related('clients', 'articles')
    
    # Séparer les dépôts des boutiques normales (une seule requête)
    depots = boutiques_toutes.filter(est_depot=True)
    boutiques_normales = boutiques_toutes.filter(est_depot=False)
    
    # Utiliser les boutiques normales (sans dépôts) pour l'affichage principal
    boutiques_list = list(boutiques_normales)
    total_boutiques = len(boutiques_list)
    
    # Ajouter les statistiques pour chaque dépôt
    depots_list = list(depots)
    for depot in depots_list:
        articles_depot = depot.articles.filter(est_actif=True)
        depot.nb_articles = articles_depot.count()
        # Valeur stock en CDF (articles avec devise CDF)
        depot.valeur_stock_cdf = articles_depot.filter(devise='CDF').aggregate(
            total=Sum(F('quantite_stock') * F('prix_achat'))
        )['total'] or 0
        # Valeur stock en USD (articles avec devise USD)
        depot.valeur_stock_usd = articles_depot.filter(devise='USD').aggregate(
            total=Sum(F('quantite_stock') * F('prix_achat'))
        )['total'] or 0
        # Valeur totale pour compatibilité
        depot.valeur_stock = depot.valeur_stock_cdf
        depot.nb_transferts_mois = TransfertStock.objects.filter(
            depot_source=depot,
            date_transfert__gte=timezone.now().replace(day=1)
        ).count()
    
    # Ajouter le compteur des ventes refusées du jour pour chaque boutique
    aujourd_hui = timezone.now().date()
    for boutique in boutiques_list:
        boutique.nb_ventes_refusees_jour = VenteRejetee.objects.filter(
            boutique=boutique,
            date_tentative__date=aujourd_hui
        ).count()
    
    boutiques = boutiques_list
    
    # Statistiques des 30 derniers jours
    date_debut = timezone.now() - timedelta(days=30)
    aujourd_hui = timezone.now().date()
    # Début du mois en cours (pour les dépenses mensuelles)
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Calculs par boutique
    stats_boutiques = []
    total_ventes = 0
    total_ca = 0
    
    for boutique in boutiques:
        # Récupérer les ventes via les clients MAUI de la boutique (exclure les ventes annulées)
        try:
            ventes_boutique = Vente.objects.filter(
                client_maui__boutique=boutique,
                date_vente__gte=date_debut,
                paye=True,
                est_annulee=False
            )
            nb_ventes = ventes_boutique.count()
            ca_boutique = ventes_boutique.aggregate(total=Sum('montant_total'))['total'] or 0
        except (ValueError, TypeError):
            # Relations pas encore mises à jour après migration
            nb_ventes = 0
            ca_boutique = 0
        
        stats_boutiques.append({
            'boutique': boutique,
            'nb_ventes': nb_ventes,
            'chiffre_affaires': ca_boutique,
            'nb_articles': boutique.articles.count(),
            'nb_terminaux': boutique.clients.count()  # Utiliser clients au lieu de terminaux
        })
        
        total_ventes += nb_ventes
        total_ca += ca_boutique
    
    # Recette du jour - Séparation CDF et USD
    ventes_jour = Vente.objects.filter(
        client_maui__boutique__in=boutiques,
        date_vente__date=aujourd_hui,
        paye=True,
        est_annulee=False
    )
    
    # Recette CDF du jour (ventes en CDF uniquement)
    ventes_jour_cdf = ventes_jour.filter(devise='CDF')
    ca_jour_cdf = ventes_jour_cdf.aggregate(total=Sum('montant_total'))['total'] or 0
    
    # Recette USD du jour (ventes en USD uniquement)
    ventes_jour_usd = ventes_jour.filter(devise='USD')
    ca_jour_usd = ventes_jour_usd.aggregate(total=Sum('montant_total_usd'))['total'] or 0
    
    # Total jour (pour compatibilité, on garde le CDF)
    ca_jour = ca_jour_cdf
    
    # Recette 30 jours - Séparation CDF et USD
    ventes_30j = Vente.objects.filter(
        client_maui__boutique__in=boutiques,
        date_vente__gte=date_debut,
        paye=True,
        est_annulee=False
    )
    
    # Recette CDF 30 jours
    ventes_30j_cdf = ventes_30j.filter(devise='CDF')
    ca_30j_cdf = ventes_30j_cdf.aggregate(total=Sum('montant_total'))['total'] or 0
    
    # Recette USD 30 jours
    ventes_30j_usd = ventes_30j.filter(devise='USD')
    ca_30j_usd = ventes_30j_usd.aggregate(total=Sum('montant_total_usd'))['total'] or 0

    # Valeur totale de la marchandise (stock) en CDF
    articles_commercant = Article.objects.filter(
        boutique__commercant=commercant,
        est_actif=True
    )
    valeur_marchandise = articles_commercant.aggregate(
        total=Sum(F('quantite_stock') * F('prix_achat'))
    )['total'] or 0

    # Dépenses totales de toutes les boutiques (rapports de caisse en CDF) sur le mois en cours
    depenses_qs = RapportCaisse.objects.filter(
        boutique__in=boutiques,
        devise='CDF',
        date_rapport__gte=debut_mois
    )
    depenses_totales = depenses_qs.aggregate(total=Sum('depense'))['total'] or 0

    # Articles en stock bas (tous les boutiques)
    articles_stock_bas = Article.objects.filter(
        boutique__commercant=commercant
    ).filter(
        quantite_stock__lte=5  # Seuil par défaut
    ).select_related('boutique')[:10]
    
    # 💰 NÉGOCIATIONS - Statistiques des prix négociés ce mois
    from .models import LigneVente
    lignes_negociees_mois = LigneVente.objects.filter(
        vente__boutique__in=boutiques,
        vente__date_vente__gte=debut_mois,
        est_negocie=True
    ).aggregate(
        nombre=Count('id'),
        total_reduction=Sum(F('prix_original') - F('prix_unitaire'))
    )
    negociations_mois = lignes_negociees_mois['nombre'] or 0
    montant_negocie_mois = lignes_negociees_mois['total_reduction'] or 0
    
    context = {
        'commercant': commercant,
        'boutiques': boutiques,  # Ajouter la liste des boutiques
        'depots': depots_list,  # Ajouter la liste des dépôts
        'total_boutiques': total_boutiques,
        'total_ventes': total_ventes,
        'total_ca': total_ca,
        'chiffre_affaires_30j': ca_30j_cdf,  # CDF 30 jours
        'chiffre_affaires_30j_usd': ca_30j_usd,  # USD 30 jours
        'recette_jour': ca_jour_cdf,  # CDF du jour
        'recette_jour_usd': ca_jour_usd,  # USD du jour
        'nb_ventes_jour_cdf': ventes_jour_cdf.count(),
        'nb_ventes_jour_usd': ventes_jour_usd.count(),
        'valeur_marchandise': valeur_marchandise,
        'depenses_totales': depenses_totales,
        'boutiques_avec_clients': boutiques_toutes.filter(clients__isnull=False).distinct().count(),
        'stats_boutiques': stats_boutiques,
        'articles_stock_bas': articles_stock_bas,
        'peut_ajouter_boutique': commercant.peut_creer_boutique(),
        # 💰 Négociations
        'negociations_mois': negociations_mois,
        'montant_negocie_mois': montant_negocie_mois,
    }
    
    return render(request, 'inventory/commercant/dashboard.html', context)

@login_required
@commercant_required
def modifier_taux_dollar(request):
    """Modifier le taux de change USD/CDF du commerçant"""
    if request.method == 'POST':
        commercant = request.user.profil_commercant
        taux = request.POST.get('taux_dollar')
        try:
            taux_decimal = Decimal(taux)
            if taux_decimal >= 1:
                commercant.taux_dollar = taux_decimal
                commercant.save(update_fields=['taux_dollar'])
                messages.success(request, f"Taux de change mis à jour: 1$ = {taux_decimal:,.0f} FC")
            else:
                messages.error(request, "Le taux doit être supérieur ou égal à 1")
        except (ValueError, TypeError, InvalidOperation):
            messages.error(request, "Veuillez entrer un taux valide")
    return redirect('inventory:commercant_dashboard')

# ===== GESTION DES BOUTIQUES =====

@login_required
@commercant_required
def liste_boutiques(request):
    """Liste des boutiques du commerçant"""
    commercant = request.user.profil_commercant
    boutiques = commercant.boutiques.all().annotate(
        nb_articles=Count('articles'),
        nb_terminaux=Count('clients')
    )
    
    context = {
        'commercant': commercant,
        'boutiques': boutiques,
        'peut_ajouter_boutique': commercant.peut_creer_boutique()
    }
    
    return render(request, 'inventory/commercant/liste_boutiques.html', context)

@login_required
@commercant_required
def creer_boutique(request):
    """Créer une nouvelle boutique"""
    commercant = request.user.profil_commercant
    
    if not commercant.peut_creer_boutique():
        messages.error(request, f"Vous avez atteint la limite de {commercant.max_boutiques} boutique(s).")
        return redirect('inventory:commercant_boutiques')
    
    if request.method == 'POST':
        form = BoutiqueForm(request.POST)
        if form.is_valid():
            boutique = form.save(commit=False)
            boutique.commercant = commercant
            # Forcer le flag dépôt si le type sélectionné est "DEPOT"
            boutique.est_depot = boutique.type_commerce == 'DEPOT'
            boutique.save()
            messages.success(request, f"Boutique '{boutique.nom}' créée avec succès!")
            return redirect('inventory:commercant_detail_boutique', boutique_id=boutique.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        form = BoutiqueForm()
    
    context = {
        'commercant': commercant,
        'form': form
    }
    
    return render(request, 'inventory/commercant/ajouter_boutique.html', context)

@login_required
@commercant_required
@boutique_access_required
def detail_boutique(request, boutique_id):
    """Détail d'une boutique avec statistiques"""
    boutique = request.boutique
    
    # Statistiques des 30 derniers jours
    date_debut = timezone.now() - timedelta(days=30)
    
    # Récupérer les ventes via les clients MAUI de la boutique (exclure les ventes annulées)
    try:
        ventes_recentes = Vente.objects.filter(
            client_maui__boutique=boutique,
            date_vente__gte=date_debut,
            paye=True,
            est_annulee=False
        )
    except (ValueError, TypeError):
        # Relations pas encore mises à jour après migration
        ventes_recentes = Vente.objects.none()
    
    nb_ventes = ventes_recentes.count()
    ca_total = ventes_recentes.aggregate(total=Sum('montant_total'))['total'] or 0
    
    # Articles les plus vendus
    from django.db.models import Sum as DbSum
    articles_populaires = Article.objects.filter(
        boutique=boutique,
        lignevente__vente__date_vente__gte=date_debut,
        lignevente__vente__paye=True
    ).annotate(
        total_vendu=DbSum('lignevente__quantite')
    ).order_by('-total_vendu')[:5]
    
    # Terminaux actifs
    terminaux = boutique.clients.filter(est_actif=True)
    
    # Articles en stock bas
    articles_stock_bas = boutique.articles.filter(
        est_actif=True,
        quantite_stock__lte=boutique.alerte_stock_bas
    )[:10]
    
    # Statistiques supplémentaires pour le template
    total_articles = boutique.articles.count()
    total_categories = boutique.categories.count()
    total_ventes = nb_ventes
    chiffre_affaires = ca_total
    
    # Ventes récentes limitées pour affichage
    ventes_recentes_display = ventes_recentes[:10]
    
    # Compteur des ventes refusées du jour + total potentiel
    aujourd_hui = timezone.now().date()
    ventes_refusees_jour = VenteRejetee.objects.filter(
        boutique=boutique,
        date_tentative__date=aujourd_hui
    )
    nb_ventes_refusees_jour = ventes_refusees_jour.count()
    
    # Calculer le total potentiel des ventes refusées
    total_potentiel_refusees = Decimal('0')
    for vente in ventes_refusees_jour:
        try:
            donnees = vente.donnees_vente
            if isinstance(donnees, dict):
                montant = donnees.get('montant_total', 0)
                total_potentiel_refusees += Decimal(str(montant))
        except (TypeError, ValueError, KeyError):
            pass
    
    context = {
        'boutique': boutique,
        'nb_ventes': nb_ventes,
        'ca_total': ca_total,
        'articles_populaires': articles_populaires,
        'terminaux': terminaux,
        'articles_stock_bas': articles_stock_bas,
        'nb_articles_total': total_articles,
        'total_articles': total_articles,
        'total_categories': total_categories,
        'total_ventes': total_ventes,
        'chiffre_affaires': chiffre_affaires,
        'ventes_recentes': ventes_recentes_display,
        'nb_ventes_refusees_jour': nb_ventes_refusees_jour,
        'total_potentiel_refusees': total_potentiel_refusees
    }
    
    return render(request, 'inventory/commercant/details_boutique.html', context)

# ===== GESTION DES ARTICLES =====

@login_required
@commercant_required
@boutique_access_required
def articles_boutique(request, boutique_id):
    """Liste des articles d'une boutique"""
    boutique = request.boutique
    
    # Récupérer les filtres
    search = request.GET.get('search', '')
    categorie_id = request.GET.get('categorie', '')
    stock_filter = request.GET.get('stock', '')
    populaires_filter = request.GET.get('populaires', '')
    
    # Requête optimisée avec select_related et prefetch_related pour les variantes
    articles = boutique.articles.filter(est_actif=True).select_related('categorie').prefetch_related('variantes')
    
    if search:
        # ⭐ Chercher dans articles ET dans les codes-barres des variantes
        articles = articles.filter(
            Q(nom__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search) |
            Q(variantes__code_barre__icontains=search, variantes__est_actif=True)
        ).distinct()
    
    if categorie_id:
        articles = articles.filter(categorie_id=categorie_id)
    
    if stock_filter == 'bas':
        articles = articles.filter(quantite_stock__lte=boutique.alerte_stock_bas)
    elif stock_filter == 'zero':
        articles = articles.filter(quantite_stock=0)
    elif stock_filter == 'normal':
        articles = articles.filter(quantite_stock__gt=boutique.alerte_stock_bas)
    
    # Filtre pour les articles populaires (ayant des ventes)
    if populaires_filter:
        from django.db.models import Count
        articles = articles.annotate(
            nb_ventes=Count('lignevente')
        ).filter(nb_ventes__gt=0).order_by('-nb_ventes')
    else:
        articles = articles.order_by('nom')
    
    # Comptage total avant pagination
    total_articles = articles.count()
    
    # Pagination - 30 articles par page
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(articles, 30)
    page = request.GET.get('page', 1)
    
    try:
        articles_page = paginator.page(page)
    except PageNotAnInteger:
        articles_page = paginator.page(1)
    except EmptyPage:
        articles_page = paginator.page(paginator.num_pages)
    
    # Catégories pour le filtre (mise en cache possible)
    categories = boutique.categories.all()
    
    # Calculer le nombre d'articles en stock bas (optimisé avec only)
    articles_stock_bas = boutique.articles.filter(
        quantite_stock__lte=boutique.alerte_stock_bas,
        est_actif=True
    ).only('id').count()
    
    # Articles en attente de validation MAUI
    articles_en_attente = boutique.articles.filter(
        est_actif=True,
        est_valide_client=False
    ).only('id').count()
    
    context = {
        'boutique': boutique,
        'articles': articles_page,
        'categories': categories,
        'articles_stock_bas': articles_stock_bas,
        'articles_en_attente': articles_en_attente,
        'search': search,
        'categorie_id': int(categorie_id) if categorie_id else None,
        'stock_filter': stock_filter,
        'populaires_filter': populaires_filter,
        'total_articles': total_articles,
        'paginator': paginator,
    }
    
    return render(request, 'inventory/commercant/articles_boutique.html', context)

@login_required
@commercant_required
@boutique_access_required
def articles_search_ajax(request, boutique_id):
    """Recherche AJAX d'articles - cherche dans TOUS les articles de la boutique ET leurs variantes"""
    from django.http import JsonResponse
    from inventory.models import VarianteArticle
    
    boutique = request.boutique
    search = request.GET.get('q', '').strip()
    categorie_id = request.GET.get('categorie_id', '').strip()
    
    if len(search) < 2:
        return JsonResponse({'articles': [], 'count': 0})
    
    # 1. Chercher dans les articles directement
    qs = boutique.articles.filter(est_actif=True)
    if categorie_id:
        qs = qs.filter(categorie_id=categorie_id)
    articles = qs.filter(
        Q(nom__icontains=search) |
        Q(code__icontains=search) |
        Q(description__icontains=search)
    ).select_related('categorie').order_by('nom')[:50]
    
    # 2. ⭐ Chercher aussi dans les codes-barres des variantes
    # Récupérer les articles parents des variantes correspondantes
    variantes_qs = boutique.articles.filter(
        est_actif=True,
        variantes__code_barre__icontains=search,
        variantes__est_actif=True
    )
    if categorie_id:
        variantes_qs = variantes_qs.filter(categorie_id=categorie_id)
    articles_from_variantes = variantes_qs.select_related('categorie').distinct()
    
    # Combiner et dédupliquer
    all_article_ids = set(art.id for art in articles) | set(art.id for art in articles_from_variantes)
    articles = boutique.articles.filter(id__in=all_article_ids).select_related('categorie').order_by('nom')[:50]
    
    articles_data = []
    for art in articles:
        articles_data.append({
            'id': art.id,
            'nom': art.nom,
            'code': art.code or '',
            'categorie': art.categorie.nom if art.categorie else '',
            'prix_vente': str(art.prix_vente),
            'devise': art.devise,
            'quantite_stock': art.stock_total,  # Utiliser stock_total pour inclure les variantes
            'a_variantes': art.a_variantes,
            'nb_variantes': art.nb_variantes if art.a_variantes else 0,
            'image_url': art.image.url if art.image else None,
            'description': art.description[:100] if art.description else '',
            'est_valide_client': art.est_valide_client,
            'quantite_envoyee': art.quantite_envoyee,
            'url': f'/commercant/boutiques/{boutique.id}/articles/{art.id}/'
        })
    
    return JsonResponse({
        'articles': articles_data,
        'count': len(articles_data),
        'search': search
    })

@login_required
@commercant_required
@boutique_access_required
def article_variantes_ajax(request, boutique_id, article_id):
    """Récupérer les variantes d'un article via AJAX"""
    from django.http import JsonResponse
    from inventory.models import Article, VarianteArticle
    
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Article non trouvé'})
    
    variantes = article.variantes.filter(est_actif=True).order_by('nom_variante')
    
    variantes_data = []
    for v in variantes:
        variantes_data.append({
            'id': v.id,
            'code_barre': v.code_barre,
            'nom_variante': v.nom_variante,
            'type_attribut': v.get_type_attribut_display(),
        })
    
    return JsonResponse({
        'success': True,
        'article_id': article.id,
        'article_nom': article.nom,
        'stock_parent': article.quantite_stock,
        'variantes': variantes_data,
        'count': len(variantes_data)
    })

@login_required
@commercant_required
@boutique_access_required
def categories_boutique(request, boutique_id):
    """Liste des catégories d'une boutique"""
    boutique = request.boutique
    
    # Récupérer les catégories de la boutique
    categories = boutique.categories.all().order_by('nom')
    
    # Statistiques par catégorie
    categories_stats = []
    for categorie in categories:
        nb_articles = categorie.articles.filter(est_actif=True).count()
        categories_stats.append({
            'categorie': categorie,
            'nb_articles': nb_articles
        })
    
    context = {
        'boutique': boutique,
        'categories': categories,
        'categories_stats': categories_stats,
        'total_categories': categories.count()
    }
    
    return render(request, 'inventory/commercant/categories_boutique.html', context)

# ===== GESTION DES TERMINAUX =====

@login_required
@commercant_required
@boutique_access_required
def terminaux_boutique(request, boutique_id):
    """Gestion des terminaux MAUI d'une boutique"""
    boutique = request.boutique
    terminaux = boutique.clients.all().order_by('nom_terminal')
    
    context = {
        'boutique': boutique,
        'terminaux': terminaux
    }
    
    return render(request, 'inventory/commercant/terminaux_boutique.html', context)

@login_required
@commercant_required
@boutique_access_required
def creer_terminal(request, boutique_id):
    """Créer un nouveau terminal MAUI"""
    boutique = request.boutique
    
    if request.method == 'POST':
        nom_terminal = request.POST.get('nom_terminal')
        numero_serie = request.POST.get('numero_serie')
        nom_utilisateur = request.POST.get('nom_utilisateur', '')
        
        if nom_terminal and numero_serie:
            # Vérifier l'unicité du numéro de série
            if TerminalMaui.objects.filter(numero_serie=numero_serie).exists():
                messages.error(request, "Ce numéro de série existe déjà.")
            else:
                terminal = TerminalMaui.objects.create(
                    nom_terminal=nom_terminal,
                    boutique=boutique,
                    numero_serie=numero_serie,
                    nom_utilisateur=nom_utilisateur
                )
                
                messages.success(request, f"Terminal '{nom_terminal}' créé avec succès!")
                messages.info(request, f"Clé API générée: {terminal.cle_api}")
                return redirect('inventory:commercant_terminaux_boutique', boutique_id=boutique.id)
        else:
            messages.error(request, "Nom du terminal et numéro de série sont requis.")
    
    context = {
        'boutique': boutique
    }
    
    return render(request, 'inventory/commercant/ajouter_client_maui.html', context)

# ===== API AJAX =====

@login_required
@commercant_required
def api_stats_boutique(request, boutique_id):
    """API pour récupérer les statistiques d'une boutique en temps réel"""
    try:
        commercant = request.user.profil_commercant
        boutique = get_object_or_404(Boutique, id=boutique_id, commercant=commercant)
        
        # Statistiques d'aujourd'hui
        aujourd_hui = timezone.now().date()
        try:
            ventes_aujourd_hui = Vente.objects.filter(
                boutique=boutique,
                date_vente__date=aujourd_hui,
                paye=True
            )
        except (ValueError, TypeError):
            # Relations pas encore mises à jour après migration
            ventes_aujourd_hui = Vente.objects.none()
        
        nb_ventes = ventes_aujourd_hui.count()
        ca_aujourd_hui_brut = ventes_aujourd_hui.aggregate(total=Sum('montant_total'))['total'] or 0

        depenses_appliquees_aujourd_hui = RapportCaisse.objects.filter(
            boutique=boutique,
            date_rapport__date=aujourd_hui,
            depense_appliquee=True
        ).aggregate(total=Sum('depense'))['total'] or 0

        ca_aujourd_hui_net = ca_aujourd_hui_brut - depenses_appliquees_aujourd_hui
        
        # Terminaux connectés
        terminaux_actifs = boutique.clients.filter(
            est_actif=True,
            sessions__est_active=True
        ).distinct().count()
        
        return JsonResponse({
            'success': True,
            'stats': {
                'ventes_aujourd_hui': nb_ventes,
                'ca_aujourd_hui': float(ca_aujourd_hui_net),
                'ca_aujourd_hui_brut': float(ca_aujourd_hui_brut),
                'depenses_appliquees_aujourd_hui': float(depenses_appliquees_aujourd_hui),
                'terminaux_connectes': terminaux_actifs
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@commercant_required
@boutique_access_required
def api_ca_jour_boutique(request, boutique_id):
    """
    Endpoint léger : CA du jour en temps réel pour un point de vente.
    - Sépare CDF et USD
    - Renvoie le timestamp de la dernière vente pour détection de changement côté client
    - Conçu pour le polling intelligent (Visibility API + backoff exponentiel)
    """
    boutique = request.boutique
    aujourd_hui = timezone.now().date()

    ventes_qs = Vente.objects.filter(
        client_maui__boutique=boutique,
        date_vente__date=aujourd_hui,
        paye=True,
        est_annulee=False,
    )

    # Agrégats CDF
    agg_cdf = ventes_qs.filter(devise='CDF').aggregate(
        total=Sum('montant_total'),
        nb=Count('id'),
    )
    ca_cdf = float(agg_cdf['total'] or 0)
    nb_cdf = int(agg_cdf['nb'] or 0)

    # Agrégats USD
    agg_usd = ventes_qs.filter(devise='USD').aggregate(
        total=Sum('montant_total_usd'),
        nb=Count('id'),
    )
    ca_usd = float(agg_usd['total'] or 0)
    nb_usd = int(agg_usd['nb'] or 0)

    # Dépenses du jour appliquées (rapports de caisse CDF)
    depenses_cdf = float(
        RapportCaisse.objects.filter(
            boutique=boutique,
            date_rapport__date=aujourd_hui,
            depense_appliquee=True,
            devise='CDF',
        ).aggregate(total=Sum('depense'))['total'] or 0
    )

    # Timestamp de la dernière vente → permet au client de savoir si les données ont changé
    derniere_vente_ts = (
        ventes_qs.order_by('-date_vente')
        .values_list('date_vente', flat=True)
        .first()
    )

    return JsonResponse({
        'ca_cdf':        ca_cdf,
        'ca_cdf_net':    round(ca_cdf - depenses_cdf, 2),
        'ca_usd':        ca_usd,
        'nb_ventes_cdf': nb_cdf,
        'nb_ventes_usd': nb_usd,
        'nb_ventes':     nb_cdf + nb_usd,
        'depenses_cdf':  depenses_cdf,
        'derniere_vente_ts': derniere_vente_ts.isoformat() if derniere_vente_ts else None,
        'heure_serveur': timezone.now().strftime('%H:%M:%S'),
    })


# ===== GESTION AVANCÉE DES BOUTIQUES =====

@login_required
@commercant_required
@boutique_access_required
def modifier_boutique(request, boutique_id):
    """Modifier une boutique"""
    boutique = request.boutique
    
    if request.method == 'POST':
        form = BoutiqueForm(request.POST, instance=boutique)
        if form.is_valid():
            form.save()
            messages.success(request, f"Boutique '{boutique.nom}' modifiée avec succès!")
            return redirect('inventory:commercant_detail_boutique', boutique_id=boutique.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        form = BoutiqueForm(instance=boutique)
    
    context = {
        'boutique': boutique,
        'form': form
    }
    
    return render(request, 'inventory/commercant/modifier_boutique.html', context)

@login_required
@commercant_required
@boutique_access_required
def supprimer_boutique(request, boutique_id):
    """Supprimer une boutique"""
    boutique = request.boutique
    
    if request.method == 'POST':
        nom_boutique = boutique.nom
        boutique.delete()
        messages.success(request, f"Boutique '{nom_boutique}' supprimée avec succès!")
        return redirect('inventory:commercant_boutiques')
    
    context = {
        'boutique': boutique
    }
    
    return render(request, 'inventory/commercant/supprimer_boutique.html', context)

@login_required
@commercant_required
@boutique_access_required
def toggle_boutique_pos(request, boutique_id):
    boutique = request.boutique
    boutique.pos_autorise = not boutique.pos_autorise
    boutique.save(update_fields=['pos_autorise'])
    if boutique.pos_autorise:
        messages.success(request, f"POS MAUI autorisé pour la boutique '{boutique.nom}'.")
    else:
        messages.warning(request, f"POS MAUI désactivé pour la boutique '{boutique.nom}'. Aucun terminal ne pourra enregistrer de ventes.")
    return redirect('inventory:commercant_dashboard')

@login_required
@commercant_required
@boutique_access_required
def entrer_boutique(request, boutique_id):
    """Entrer dans une boutique (dashboard boutique)"""
    boutique = request.boutique
    
    # Statistiques de base
    nb_articles = boutique.articles.count()
    nb_terminaux = boutique.clients.count()
    
    # Ventes d'aujourd'hui (exclure les ventes annulées)
    date_aujourd_hui = timezone.now().date()
    try:
        ventes_aujourd_hui = Vente.objects.filter(
            boutique=boutique,
            date_vente__date=date_aujourd_hui,
            paye=True,
            est_annulee=False
        )
        nb_ventes_aujourd_hui = ventes_aujourd_hui.count()
        ca_aujourd_hui_brut = ventes_aujourd_hui.aggregate(total=Sum('montant_total'))['total'] or 0
    except (ValueError, TypeError):
        nb_ventes_aujourd_hui = 0
        ca_aujourd_hui_brut = 0

    depenses_appliquees_ca_jour = RapportCaisse.objects.filter(
        boutique=boutique,
        date_rapport__date=date_aujourd_hui,
        depense_appliquee=True
    ).aggregate(total=Sum('depense'))['total'] or 0

    ca_aujourd_hui = ca_aujourd_hui_brut - depenses_appliquees_ca_jour
    
    # Ventes du mois en cours (exclure les ventes annulées)
    try:
        premier_jour_mois = timezone.now().date().replace(day=1)
        ventes_mois = Vente.objects.filter(
            boutique=boutique,
            date_vente__date__gte=premier_jour_mois,
            paye=True,
            est_annulee=False
        )
        nb_ventes_mois = ventes_mois.count()
        ca_mois_brut = ventes_mois.aggregate(total=Sum('montant_total'))['total'] or 0
    except (ValueError, TypeError):
        nb_ventes_mois = 0
        ca_mois_brut = 0

    depenses_appliquees_ca_mois = RapportCaisse.objects.filter(
        boutique=boutique,
        date_rapport__date__gte=premier_jour_mois,
        date_rapport__date__lte=date_aujourd_hui,
        depense_appliquee=True
    ).aggregate(total=Sum('depense'))['total'] or 0

    ca_mois = ca_mois_brut - depenses_appliquees_ca_mois
    
    # Variables supplémentaires pour le template dashboard.html
    total_articles = nb_articles
    total_categories = boutique.categories.count()
    ca_jour = ca_aujourd_hui
    
    # Chiffre d'affaires en USD
    try:
        ventes_usd_jour = Vente.objects.filter(
            boutique=boutique,
            date_vente__date=date_aujourd_hui,
            paye=True,
            est_annulee=False,
            devise='USD'
        )
        ca_aujourd_hui_usd = ventes_usd_jour.aggregate(total=Sum('montant_total_usd'))['total'] or 0
        
        ventes_usd_mois = Vente.objects.filter(
            boutique=boutique,
            date_vente__date__gte=premier_jour_mois,
            paye=True,
            est_annulee=False,
            devise='USD'
        )
        ca_mois_usd = ventes_usd_mois.aggregate(total=Sum('montant_total_usd'))['total'] or 0
    except:
        ca_aujourd_hui_usd = 0
        ca_mois_usd = 0
    
    # Valeur totale du stock disponible (prix de vente x quantité) - séparé par devise
    try:
        articles_actifs = boutique.articles.filter(est_actif=True, quantite_stock__gt=0)
        
        # Stock en CDF
        valeur_stock_cdf = articles_actifs.filter(devise='CDF').aggregate(
            total=Sum(F('prix_vente') * F('quantite_stock'))
        )['total'] or 0
        
        # Stock en USD
        valeur_stock_usd = articles_actifs.filter(devise='USD').aggregate(
            total=Sum(F('prix_vente') * F('quantite_stock'))
        )['total'] or 0
        
        # Valeur totale pour compatibilité (en CDF)
        valeur_stock_disponible = valeur_stock_cdf
    except Exception:
        valeur_stock_disponible = 0
        valeur_stock_cdf = 0
        valeur_stock_usd = 0
    
    # Articles en stock faible
    try:
        articles_stock_faible = boutique.articles.filter(
            est_actif=True,
            quantite_stock__lte=boutique.alerte_stock_bas
        )
    except:
        articles_stock_faible = boutique.articles.none()
    
    # Mouvements de stock des 7 derniers jours
    try:
        date_7_jours = timezone.now() - timezone.timedelta(days=7)
        mouvements_recents = MouvementStock.objects.filter(
            article__boutique=boutique,
            date_mouvement__gte=date_7_jours
        ).count()
    except:
        mouvements_recents = 0
    
    # Ventes récentes (exclure les ventes annulées)
    try:
        ventes_recentes = Vente.objects.filter(
            client_maui__boutique=boutique,
            est_annulee=False
        ).order_by('-date_vente')[:10]
    except:
        ventes_recentes = Vente.objects.none()
    
    # Articles populaires (top 5 des articles les plus vendus sur 30 jours)
    try:
        date_30_jours = timezone.now() - timezone.timedelta(days=30)
        articles_populaires = Article.objects.filter(
            boutique=boutique,
            est_actif=True,
            lignevente__vente__est_annulee=False,
            lignevente__vente__date_vente__gte=date_30_jours
        ).annotate(
            total_vendus=Sum('lignevente__quantite'),
            total_revenus=Sum(F('lignevente__quantite') * F('lignevente__prix_unitaire'))
        ).filter(total_vendus__gt=0).order_by('-total_vendus')[:5]
    except:
        articles_populaires = []
    
    # Données pour graphiques (simplifiées)
    labels_jours = []
    ca_quotidien = []
    
    # 💰 NÉGOCIATIONS - Statistiques des prix négociés pour cette boutique
    debut_mois = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    aujourd_hui = timezone.now().date()
    
    # Négociations du jour
    lignes_negociees_jour = LigneVente.objects.filter(
        vente__boutique=boutique,
        vente__date_vente__date=aujourd_hui,
        est_negocie=True
    ).aggregate(
        nombre=Count('id'),
        total_reduction=Sum(F('prix_original') - F('prix_unitaire'))
    )
    negociations_jour = lignes_negociees_jour['nombre'] or 0
    montant_negocie_jour = lignes_negociees_jour['total_reduction'] or 0
    
    # Négociations du mois
    lignes_negociees_mois = LigneVente.objects.filter(
        vente__boutique=boutique,
        vente__date_vente__gte=debut_mois,
        est_negocie=True
    ).aggregate(
        nombre=Count('id'),
        total_reduction=Sum(F('prix_original') - F('prix_unitaire'))
    )
    negociations_mois = lignes_negociees_mois['nombre'] or 0
    montant_negocie_mois = lignes_negociees_mois['total_reduction'] or 0
    
    # Liste des articles négociés du jour (pour modal)
    articles_negocies_jour = LigneVente.objects.filter(
        vente__boutique=boutique,
        vente__date_vente__date=aujourd_hui,
        est_negocie=True
    ).select_related('article', 'vente').order_by('-vente__date_vente')
    
    # Liste des articles négociés du mois (pour modal)
    articles_negocies_mois = LigneVente.objects.filter(
        vente__boutique=boutique,
        vente__date_vente__gte=debut_mois,
        est_negocie=True
    ).select_related('article', 'vente').order_by('-vente__date_vente')
    
    context = {
        'boutique': boutique,
        'nb_articles': nb_articles,
        'nb_terminaux': nb_terminaux,
        'nb_ventes_aujourd_hui': nb_ventes_aujourd_hui,
        'ca_aujourd_hui': ca_aujourd_hui,
        'ca_aujourd_hui_brut': ca_aujourd_hui_brut,
        'ca_aujourd_hui_usd': ca_aujourd_hui_usd,
        'depenses_appliquees_ca_jour': depenses_appliquees_ca_jour,
        'total_articles': total_articles,
        'total_categories': total_categories,
        'ca_mois': ca_mois,
        'ca_mois_brut': ca_mois_brut,
        'ca_mois_usd': ca_mois_usd,
        'depenses_appliquees_ca_mois': depenses_appliquees_ca_mois,
        'ca_jour': ca_jour,
        'valeur_stock_disponible': valeur_stock_disponible,
        'valeur_stock_cdf': valeur_stock_cdf,
        'valeur_stock_usd': valeur_stock_usd,
        'articles_stock_faible': articles_stock_faible,
        'mouvements_recents': mouvements_recents,
        'ventes_recentes': ventes_recentes,
        'articles_populaires': articles_populaires,
        'labels_jours': labels_jours,
        'ca_quotidien': ca_quotidien,
        # 💰 Négociations
        'negociations_jour': negociations_jour,
        'montant_negocie_jour': montant_negocie_jour,
        'negociations_mois': negociations_mois,
        'montant_negocie_mois': montant_negocie_mois,
        'articles_negocies_jour': articles_negocies_jour,
        'articles_negocies_mois': articles_negocies_mois,
        # 📦 Dépôt du commerçant (pour lien rapide)
        'depot': Boutique.objects.filter(commercant=request.user.profil_commercant, est_depot=True).first(),
    }
    
    return render(request, 'inventory/boutique/dashboard.html', context)

@login_required
@commercant_required
@boutique_access_required
def ajouter_client_maui_boutique(request, boutique_id):
    """Ajouter un client MAUI à une boutique spécifique"""
    boutique = request.boutique
    
    if request.method == 'POST':
        nom_terminal = request.POST.get('nom_terminal')
        numero_serie = request.POST.get('numero_serie')
        description = request.POST.get('description', '')
        
        if nom_terminal and numero_serie:
            # Vérifier l'unicité du numéro de série
            if Client.objects.filter(numero_serie=numero_serie).exists():
                messages.error(request, "Ce numéro de série existe déjà.")
            else:
                client = Client.objects.create(
                    nom_terminal=nom_terminal,
                    boutique=boutique,
                    compte_proprietaire=boutique.commercant.user,
                    numero_serie=numero_serie,
                    description=description
                )
                
                messages.success(request, f"Terminal '{nom_terminal}' créé avec succès!")
                messages.info(request, f"Clé API générée: {client.cle_api}")
                return redirect('inventory:commercant_detail_boutique', boutique_id=boutique.id)
        else:
            messages.error(request, "Nom du terminal et numéro de série sont requis.")
    
    context = {
        'boutique': boutique
    }
    
    return render(request, 'inventory/commercant/ajouter_client_maui.html', context)

@login_required
@commercant_required
@boutique_access_required
def rapport_ca_quotidien(request, boutique_id):
    """Afficher le rapport du chiffre d'affaires pour une journée donnée (par défaut aujourd'hui)"""
    boutique = request.boutique

    # Récupérer la date ciblée (par défaut aujourd'hui)
    date_str = request.GET.get('date')
    if date_str:
        try:
            date_cible = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            date_cible = timezone.now().date()
    else:
        date_cible = timezone.now().date()

    ventes = Vente.objects.filter(
        boutique=boutique,
        date_vente__date=date_cible,
        paye=True,
        est_annulee=False
    ).select_related('client_maui').prefetch_related('lignes__article')

    total_ventes = ventes.count()
    
    # CA en CDF (ventes en CDF)
    ventes_cdf = ventes.filter(devise='CDF')
    total_ca_cdf = ventes_cdf.aggregate(total=Sum('montant_total'))['total'] or 0
    
    # CA en USD (ventes en USD)
    ventes_usd = ventes.filter(devise='USD')
    total_ca_usd = ventes_usd.aggregate(total=Sum('montant_total_usd'))['total'] or 0
    
    # Total CA (pour compatibilité - somme CDF)
    total_ca = total_ca_cdf

    depenses_appliquees_qs = RapportCaisse.objects.filter(
        boutique=boutique,
        date_rapport__date=date_cible,
        depense_appliquee=True
    )
    total_depenses_appliquees = depenses_appliquees_qs.aggregate(total=Sum('depense'))['total'] or 0

    total_ca_brut = total_ca_cdf
    total_ca_brut_usd = total_ca_usd
    total_ca_net = total_ca_brut - total_depenses_appliquees

    ventes_par_mode = ventes.values('mode_paiement').annotate(
        count=Count('id'),
        total=Sum('montant_total')
    )
    
    # Ventes par mode en USD
    ventes_par_mode_usd = ventes_usd.values('mode_paiement').annotate(
        count=Count('id'),
        total=Sum('montant_total_usd')
    )
    
    # ===== CALCUL DES BÉNÉFICES =====
    # Calculer le coût d'achat total et le bénéfice brut
    total_cout_achat = Decimal('0')
    total_benefice_brut = Decimal('0')
    articles_sans_prix_achat = []  # Articles avec prix_achat = 0
    
    for vente in ventes:
        for ligne in vente.lignes.all():
            if ligne.article:
                prix_achat = ligne.article.prix_achat or Decimal('0')
                cout_achat_ligne = prix_achat * ligne.quantite
                prix_vente_ligne = ligne.prix_unitaire * ligne.quantite
                total_cout_achat += cout_achat_ligne
                total_benefice_brut += (prix_vente_ligne - cout_achat_ligne)
                
                # Détecter les articles sans prix d'achat
                if prix_achat <= 0 and ligne.article.nom not in articles_sans_prix_achat:
                    articles_sans_prix_achat.append(ligne.article.nom)
    
    # Bénéfice net = Bénéfice brut - Dépenses appliquées
    total_benefice_net = total_benefice_brut - total_depenses_appliquees
    
    # Marge bénéficiaire en pourcentage
    marge_beneficiaire = 0
    if total_ca_brut > 0:
        marge_beneficiaire = (total_benefice_brut / total_ca_brut) * 100
    
    # Avertissement si calcul incomplet
    benefice_incomplet = len(articles_sans_prix_achat) > 0

    context = {
        'boutique': boutique,
        'ventes': ventes,
        'date_cible': date_cible,
        'total_ventes': total_ventes,
        'total_ca': total_ca_net,
        'total_ca_brut': total_ca_brut,
        'total_ca_usd': total_ca_usd,
        'total_ca_brut_usd': total_ca_brut_usd,
        'total_depenses_appliquees': total_depenses_appliquees,
        'ventes_par_mode': ventes_par_mode,
        'ventes_par_mode_usd': ventes_par_mode_usd,
        # Données de bénéfices
        'total_cout_achat': total_cout_achat,
        'total_benefice_brut': total_benefice_brut,
        'total_benefice_net': total_benefice_net,
        'marge_beneficiaire': marge_beneficiaire,
        'benefice_incomplet': benefice_incomplet,
        'articles_sans_prix_achat': articles_sans_prix_achat,
    }

    return render(request, 'inventory/commercant/rapport_ca_quotidien.html', context)

@login_required
@commercant_required
@boutique_access_required
def rapport_ca_mensuel(request, boutique_id):
    """Afficher le rapport du chiffre d'affaires mensuel"""
    boutique = request.boutique
    
    # Récupérer le mois et l'année depuis les paramètres GET (optionnel)
    try:
        annee = int(request.GET.get('annee', timezone.now().year))
        mois = int(request.GET.get('mois', timezone.now().month))
    except (ValueError, TypeError):
        annee = timezone.now().year
        mois = timezone.now().month
    
    # Noms des mois en français
    mois_noms = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
    }
    nom_mois = mois_noms.get(mois, 'Mois')
    
    # Calculer le premier et dernier jour du mois
    from calendar import monthrange
    premier_jour = datetime(annee, mois, 1).date()
    dernier_jour_num = monthrange(annee, mois)[1]
    dernier_jour = datetime(annee, mois, dernier_jour_num).date()
    
    rapports_jours = []
    current_date = premier_jour
    total_ca = 0
    total_ventes = 0
    total_ca_brut = 0
    total_depenses_appliquees = 0
    
    # Parcourir tous les jours du mois
    while current_date <= dernier_jour:
        ventes_jour = Vente.objects.filter(
            boutique=boutique,
            date_vente__date=current_date,
            paye=True,
            est_annulee=False
        )
        nb_ventes = ventes_jour.count()
        ca_jour_brut = ventes_jour.aggregate(total=Sum('montant_total'))['total'] or 0

        depenses_jour = RapportCaisse.objects.filter(
            boutique=boutique,
            date_rapport__date=current_date,
            depense_appliquee=True
        ).aggregate(total=Sum('depense'))['total'] or 0

        ca_jour = ca_jour_brut - depenses_jour
        
        rapports_jours.append({
            'date': current_date,
            'nb_ventes': nb_ventes,
            'ca': ca_jour,
            'ca_brut': ca_jour_brut,
            'depenses_appliquees': depenses_jour
        })
        
        total_ventes += nb_ventes
        total_ca += ca_jour
        total_ca_brut += ca_jour_brut
        total_depenses_appliquees += depenses_jour
        current_date += timedelta(days=1)
    
    # Inverser pour avoir les dates les plus récentes en premier
    rapports_jours.reverse()
    
    # Calculer mois précédent et suivant pour navigation
    if mois == 1:
        mois_precedent = 12
        annee_precedente = annee - 1
    else:
        mois_precedent = mois - 1
        annee_precedente = annee
    
    if mois == 12:
        mois_suivant = 1
        annee_suivante = annee + 1
    else:
        mois_suivant = mois + 1
        annee_suivante = annee
    
    context = {
        'boutique': boutique,
        'rapports_jours': rapports_jours,
        'total_ventes': total_ventes,
        'total_ca': total_ca,
        'total_ca_brut': total_ca_brut,
        'total_depenses_appliquees': total_depenses_appliquees,
        'mois': mois,
        'annee': annee,
        'nom_mois': nom_mois,
        'mois_precedent': mois_precedent,
        'annee_precedente': annee_precedente,
        'mois_suivant': mois_suivant,
        'annee_suivante': annee_suivante,
    }
    
    return render(request, 'inventory/commercant/rapport_ca_mensuel.html', context)

@login_required
@commercant_required
@boutique_access_required
def rapports_caisse_boutique(request, boutique_id):
    """Liste des rapports de caisse envoyés depuis les terminaux MAUI pour une boutique."""
    boutique = request.boutique

    rapports = RapportCaisse.objects.filter(
        boutique=boutique
    ).select_related('terminal').order_by('-date_rapport', '-created_at')

    # Filtres optionnels
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    terminal_id = request.GET.get('terminal_id')

    if date_debut:
        try:
            d = datetime.strptime(date_debut, '%Y-%m-%d').date()
            rapports = rapports.filter(date_rapport__date__gte=d)
        except ValueError:
            pass

    if date_fin:
        try:
            d = datetime.strptime(date_fin, '%Y-%m-%d').date()
            rapports = rapports.filter(date_rapport__date__lte=d)
        except ValueError:
            pass

    terminal_selected = 'TOUS'
    if terminal_id:
        if terminal_id != 'TOUS':
            try:
                terminal_id_int = int(terminal_id)
                rapports = rapports.filter(terminal_id=terminal_id_int)
                terminal_selected = terminal_id_int
            except (TypeError, ValueError):
                terminal_selected = 'TOUS'

    terminaux = boutique.clients.all().order_by('nom_terminal')

    # Statistiques par devise
    total_depense_cdf = rapports.filter(devise='CDF').aggregate(total=Sum('depense'))['total'] or 0
    total_depense_usd = rapports.filter(devise='USD').aggregate(total=Sum('depense'))['total'] or 0
    total_depense = total_depense_cdf  # Pour compatibilité

    # Notifications de rapports non lus (style "Facebook")
    now = timezone.now()
    boutique.derniere_lecture_rapports_caisse = now
    boutique.save(update_fields=['derniere_lecture_rapports_caisse'])

    # Sur cette page, tous les rapports de caisse sont considérés comme lus
    unread_rapports_count = 0

    # Articles négociés non lus
    if boutique.derniere_lecture_articles_negocies:
        unread_articles_negocies_count = ArticleNegocie.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_articles_negocies
        ).count()
    else:
        unread_articles_negocies_count = ArticleNegocie.objects.filter(
            boutique=boutique
        ).count()

    # Retours d'articles non lus
    if boutique.derniere_lecture_retours_articles:
        unread_retours_count = RetourArticle.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_retours_articles
        ).count()
    else:
        unread_retours_count = RetourArticle.objects.filter(
            boutique=boutique
        ).count()

    context = {
        'boutique': boutique,
        'rapports': rapports,
        'terminaux': terminaux,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'terminal_selected': terminal_selected,
        'total_depense': total_depense,
        'total_depense_cdf': total_depense_cdf,
        'total_depense_usd': total_depense_usd,
        'unread_rapports_count': unread_rapports_count,
        'unread_articles_negocies_count': unread_articles_negocies_count,
        'unread_retours_count': unread_retours_count,
    }

    return render(request, 'inventory/commercant/rapports_caisse_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def articles_negocies_boutique(request, boutique_id):
    boutique = request.boutique

    articles = ArticleNegocie.objects.filter(
        boutique=boutique
    ).select_related('terminal', 'article').order_by('-date_operation', '-created_at')

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    terminal_id = request.GET.get('terminal_id')

    if date_debut:
        try:
            d = datetime.strptime(date_debut, '%Y-%m-%d').date()
            articles = articles.filter(date_operation__date__gte=d)
        except ValueError:
            pass

    if date_fin:
        try:
            d = datetime.strptime(date_fin, '%Y-%m-%d').date()
            articles = articles.filter(date_operation__date__lte=d)
        except ValueError:
            pass

    terminal_selected = 'TOUS'
    if terminal_id:
        if terminal_id != 'TOUS':
            try:
                terminal_id_int = int(terminal_id)
                articles = articles.filter(terminal_id=terminal_id_int)
                terminal_selected = terminal_id_int
            except (TypeError, ValueError):
                terminal_selected = 'TOUS'

    terminaux = boutique.clients.all().order_by('nom_terminal')

    stats = articles.aggregate(total_montant=Sum('montant_negocie'))
    total_montant = stats['total_montant'] or 0

    # Notifications de rapports non lus (style "Facebook")
    now = timezone.now()
    boutique.derniere_lecture_articles_negocies = now
    boutique.save(update_fields=['derniere_lecture_articles_negocies'])

    # Sur cette page, tous les articles négociés sont considérés comme lus
    unread_articles_negocies_count = 0

    # Rapports de caisse non lus
    if boutique.derniere_lecture_rapports_caisse:
        unread_rapports_count = RapportCaisse.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_rapports_caisse
        ).count()
    else:
        unread_rapports_count = RapportCaisse.objects.filter(
            boutique=boutique
        ).count()

    # Retours d'articles non lus
    if boutique.derniere_lecture_retours_articles:
        unread_retours_count = RetourArticle.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_retours_articles
        ).count()
    else:
        unread_retours_count = RetourArticle.objects.filter(
            boutique=boutique
        ).count()

    # Statut d'application des négociations (vente déjà créée ou non)
    article_ids = list(articles.values_list('id', flat=True))
    applied_negociations_ids = []
    if article_ids:
        numero_map = {f"NEG-{boutique.id}-{nid}": nid for nid in article_ids}
        existing_num = Vente.objects.filter(
            boutique=boutique,
            numero_facture__in=numero_map.keys()
        ).values_list('numero_facture', flat=True)
        applied_negociations_ids = [
            numero_map[nf]
            for nf in existing_num
            if nf in numero_map
        ]

    context = {
        'boutique': boutique,
        'articles_negocies': articles,
        'terminaux': terminaux,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'terminal_selected': terminal_selected,
        'total_montant': total_montant,
        'unread_rapports_count': unread_rapports_count,
        'unread_articles_negocies_count': unread_articles_negocies_count,
        'unread_retours_count': unread_retours_count,
        'applied_negociations_ids': applied_negociations_ids,
    }

    return render(request, 'inventory/commercant/articles_negocies_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def retours_articles_boutique(request, boutique_id):
    boutique = request.boutique

    retours = RetourArticle.objects.filter(
        boutique=boutique
    ).select_related('terminal', 'article').order_by('-date_operation', '-created_at')

    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    terminal_id = request.GET.get('terminal_id')

    if date_debut:
        try:
            d = datetime.strptime(date_debut, '%Y-%m-%d').date()
            retours = retours.filter(date_operation__date__gte=d)
        except ValueError:
            pass

    if date_fin:
        try:
            d = datetime.strptime(date_fin, '%Y-%m-%d').date()
            retours = retours.filter(date_operation__date__lte=d)
        except ValueError:
            pass

    terminal_selected = 'TOUS'
    if terminal_id:
        if terminal_id != 'TOUS':
            try:
                terminal_id_int = int(terminal_id)
                retours = retours.filter(terminal_id=terminal_id_int)
                terminal_selected = terminal_id_int
            except (TypeError, ValueError):
                terminal_selected = 'TOUS'

    terminaux = boutique.clients.all().order_by('nom_terminal')

    stats = retours.aggregate(total_montant=Sum('montant_retourne'))
    total_montant = stats['total_montant'] or 0

    # Notifications de rapports non lus (style "Facebook")
    now = timezone.now()
    boutique.derniere_lecture_retours_articles = now
    boutique.save(update_fields=['derniere_lecture_retours_articles'])

    # Sur cette page, tous les retours d'articles sont considérés comme lus
    unread_retours_count = 0

    # Rapports de caisse non lus
    if boutique.derniere_lecture_rapports_caisse:
        unread_rapports_count = RapportCaisse.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_rapports_caisse
        ).count()
    else:
        unread_rapports_count = RapportCaisse.objects.filter(
            boutique=boutique
        ).count()

    # Articles négociés non lus
    if boutique.derniere_lecture_articles_negocies:
        unread_articles_negocies_count = ArticleNegocie.objects.filter(
            boutique=boutique,
            created_at__gt=boutique.derniere_lecture_articles_negocies
        ).count()
    else:
        unread_articles_negocies_count = ArticleNegocie.objects.filter(
            boutique=boutique
        ).count()

    context = {
        'boutique': boutique,
        'retours_articles': retours,
        'terminaux': terminaux,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'terminal_selected': terminal_selected,
        'total_montant': total_montant,
        'unread_rapports_count': unread_rapports_count,
        'unread_articles_negocies_count': unread_articles_negocies_count,
        'unread_retours_count': unread_retours_count,
    }

    return render(request, 'inventory/commercant/retours_articles_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def appliquer_article_negocie(request, boutique_id, negociation_id):
    boutique = request.boutique
    negociation = get_object_or_404(ArticleNegocie, id=negociation_id, boutique=boutique)

    if request.method != 'POST':
        return HttpResponseForbidden("Méthode non autorisée")

    # Idempotence simple : numéro de facture déterministe basé sur la boutique et l'ID de la négociation
    numero_facture = f"NEG-{boutique.id}-{negociation.id}"
    if Vente.objects.filter(numero_facture=numero_facture).exists():
        messages.warning(request, "Cette négociation a déjà été appliquée à la recette.")
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    if not negociation.article:
        messages.error(request, "Impossible d'appliquer cette négociation car l'article lié est introuvable.")
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    article = negociation.article
    if article.boutique_id != boutique.id:
        messages.error(request, "Cet article n'appartient plus à cette boutique.")
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    quantite = negociation.quantite or 0
    if quantite <= 0:
        messages.error(request, "Quantité invalide pour cette négociation.")
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    montant_unitaire = negociation.montant_negocie or Decimal('0')
    montant_total = montant_unitaire * quantite
    if montant_total <= 0:
        messages.error(request, "Montant négocié invalide pour cette négociation.")
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    # Vérifier le stock disponible avant de lancer la transaction
    if article.quantite_stock < quantite:
        messages.error(
            request,
            f"Stock insuffisant pour l'article {article.code}. Stock disponible : {article.quantite_stock}, quantité demandée : {quantite}."
        )
        return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)

    terminal = negociation.terminal

    try:
        with transaction.atomic():
            # Créer la vente comme si elle venait de MAUI
            vente = Vente.objects.create(
                numero_facture=numero_facture,
                date_vente=timezone.now(),
                montant_total=montant_total,
                mode_paiement='CASH',
                paye=True,
                boutique=boutique,
                client_maui=terminal,
                adresse_ip_client=request.META.get('REMOTE_ADDR'),
                version_app_maui=terminal.version_app_maui if terminal else ''
            )

            # Créer la ligne de vente
            LigneVente.objects.create(
                vente=vente,
                article=article,
                quantite=quantite,
                prix_unitaire=montant_unitaire,
            )

            # Mettre à jour le stock de l'article et journaliser le mouvement
            stock_avant = article.quantite_stock
            article.quantite_stock = stock_avant - quantite
            article.save(update_fields=['quantite_stock'])

            MouvementStock.objects.create(
                article=article,
                type_mouvement='VENTE',
                quantite=-quantite,
                stock_avant=stock_avant,
                stock_apres=article.quantite_stock,
                reference_document=vente.numero_facture,
                utilisateur=request.user.username,
                commentaire=(
                    f"Négociation appliquée depuis l'interface commerçant "
                    f"(Article négocié ID {negociation.id}, motif: {negociation.motif})"
                )
            )

        messages.success(
            request,
            f"La négociation sur l'article {article.code} a été appliquée avec succès "
            f"({quantite} × {montant_unitaire} {negociation.devise} ajoutés au CA du jour)."
        )
    except Exception as e:
        messages.error(request, f"Erreur lors de l'application de la négociation : {e}")

    return redirect('inventory:commercant_articles_negocies_boutique', boutique_id=boutique.id)


@login_required
@commercant_required
@boutique_access_required
def appliquer_depense_rapport_caisse(request, boutique_id, rapport_id):
    boutique = request.boutique
    rapport = get_object_or_404(RapportCaisse, id=rapport_id, boutique=boutique)

    if request.method != 'POST':
        return HttpResponseForbidden("Méthode non autorisée")

    if not rapport.depense_appliquee:
        rapport.depense_appliquee = True
        rapport.date_application_depense = timezone.now()
        rapport.save(update_fields=['depense_appliquee', 'date_application_depense'])
        messages.success(
            request,
            f"La dépense de {rapport.depense} {rapport.devise} du {rapport.date_rapport.date()} a été appliquée à la recette du jour."
        )
    else:
        messages.info(request, "Cette dépense est déjà appliquée à la recette du jour.")

    return redirect('inventory:commercant_rapports_caisse_boutique', boutique_id=boutique.id)

@login_required
@commercant_required
@boutique_access_required
def supprimer_rapport_caisse(request, boutique_id, rapport_id):
    """Supprimer un rapport de caisse."""
    boutique = request.boutique
    rapport = get_object_or_404(RapportCaisse, id=rapport_id, boutique=boutique)

    if request.method != 'POST':
        return HttpResponseForbidden("Méthode non autorisée")

    date_rapport = rapport.date_rapport.strftime('%d/%m/%Y %H:%i') if rapport.date_rapport else ''
    depense = rapport.depense
    devise = rapport.devise
    
    rapport.delete()
    
    messages.success(
        request,
        f"Le rapport de caisse du {date_rapport} (dépense: {depense} {devise}) a été supprimé."
    )

    return redirect('inventory:commercant_rapports_caisse_boutique', boutique_id=boutique.id)

@login_required
@commercant_required
@boutique_access_required
def exporter_ca_quotidien_pdf(request, boutique_id):
    """Exporter le chiffre d'affaires quotidien en PDF"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from datetime import datetime, timedelta
    import io
    
    boutique = request.boutique
    
    # Créer le buffer pour le PDF
    buffer = io.BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Titre
    title = Paragraph(f"Rapport CA Quotidien - {boutique.nom}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Informations boutique
    info_text = f"""
    <b>Boutique:</b> {boutique.nom}<br/>
    <b>Type:</b> {boutique.get_type_commerce_display()}<br/>
    <b>Adresse:</b> {boutique.adresse}, {boutique.ville}<br/>
    <b>Date d'export:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}<br/>
    """
    info = Paragraph(info_text, styles['Normal'])
    story.append(info)
    story.append(Spacer(1, 20))
    
    # Données des 30 derniers jours
    date_fin = timezone.now().date()
    date_debut = date_fin - timedelta(days=30)
    
    # Créer les données du tableau
    data = [['Date', 'Nb Ventes', 'Chiffre d\'Affaires (CDF)']]
    
    current_date = date_debut
    total_ca = 0
    total_ventes = 0
    
    while current_date <= date_fin:
        try:
            ventes_jour = Vente.objects.filter(
                boutique=boutique,
                date_vente__date=current_date,
                paye=True
            )
            nb_ventes = ventes_jour.count()
            ca_jour = ventes_jour.aggregate(total=Sum('montant_total'))['total'] or 0
        except (ValueError, TypeError):
            nb_ventes = 0
            ca_jour = 0
        
        data.append([
            current_date.strftime('%d/%m/%Y'),
            str(nb_ventes),
            f"{ca_jour:,.0f}"
        ])
        
        total_ventes += nb_ventes
        total_ca += ca_jour
        current_date += timedelta(days=1)
    
    # Ajouter ligne de total
    data.append(['TOTAL', str(total_ventes), f"{total_ca:,.0f}"])
    
    # Créer le tableau
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construire le PDF
    doc.build(story)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="CA_quotidien_{boutique.nom}_{date_fin.strftime("%Y%m%d")}.pdf"'
    
    return response

@login_required
@commercant_required
@boutique_access_required
def exporter_ca_mensuel_pdf(request, boutique_id):
    """Exporter le chiffre d'affaires mensuel en PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from datetime import datetime
    from calendar import monthrange
    import io
    import locale
    
    boutique = request.boutique
    
    # Récupérer le mois et l'année depuis les paramètres GET (optionnel)
    try:
        annee = int(request.GET.get('annee', timezone.now().year))
        mois = int(request.GET.get('mois', timezone.now().month))
    except (ValueError, TypeError):
        annee = timezone.now().year
        mois = timezone.now().month
    
    # Créer le buffer pour le PDF
    buffer = io.BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Noms des mois en français
    mois_noms = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
    }
    nom_mois = mois_noms.get(mois, 'Mois')
    
    # Titre avec le nom du mois
    title = Paragraph(
        f"Rapport CA Mensuel - {nom_mois} {annee}<br/>{boutique.nom}", 
        styles['Title']
    )
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Informations boutique
    info_text = f"""
    <b>Boutique:</b> {boutique.nom}<br/>
    <b>Type:</b> {boutique.get_type_commerce_display()}<br/>
    <b>Adresse:</b> {boutique.adresse}, {boutique.ville}<br/>
    <b>Période:</b> {nom_mois} {annee}<br/>
    <b>Date d'export:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}<br/>
    """
    info = Paragraph(info_text, styles['Normal'])
    story.append(info)
    story.append(Spacer(1, 20))
    
    # Calculer le premier et dernier jour du mois
    premier_jour = datetime(annee, mois, 1).date()
    dernier_jour_num = monthrange(annee, mois)[1]
    dernier_jour = datetime(annee, mois, dernier_jour_num).date()
    
    # Créer les données du tableau
    data = [['Date', 'Nb Ventes', 'Chiffre d\'Affaires (CDF)']]
    
    current_date = premier_jour
    total_ca = 0
    total_ventes = 0
    
    # Parcourir tous les jours du mois
    while current_date <= dernier_jour:
        try:
            ventes_jour = Vente.objects.filter(
                boutique=boutique,
                date_vente__date=current_date,
                paye=True
            )
            nb_ventes = ventes_jour.count()
            ca_jour = ventes_jour.aggregate(total=Sum('montant_total'))['total'] or 0
        except (ValueError, TypeError):
            nb_ventes = 0
            ca_jour = 0
        
        data.append([
            current_date.strftime('%d/%m/%Y'),
            str(nb_ventes),
            f"{ca_jour:,.0f}"
        ])
        
        total_ventes += nb_ventes
        total_ca += ca_jour
        current_date += timedelta(days=1)
    
    # Ajouter ligne de total
    data.append(['TOTAL', str(total_ventes), f"{total_ca:,.0f}"])
    
    # Créer le tableau
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Construire le PDF
    doc.build(story)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="CA_Mensuel_{nom_mois}_{annee}_{boutique.nom}.pdf"'
    
    return response

@login_required
@commercant_required
@boutique_access_required
def verifier_code_barre(request, boutique_id):
    """Vérifier si un code-barres existe déjà dans la boutique (article ou variante)"""
    from django.http import JsonResponse
    from inventory.models import VarianteArticle
    import logging
    logger = logging.getLogger(__name__)
    
    boutique = request.boutique
    code = request.GET.get('code', '').strip()
    
    logger.info(f"🔍 verifier_code_barre: code='{code}', boutique={boutique.id} ({boutique.nom})")
    
    if not code:
        return JsonResponse({'existe': False})
    
    # 1. Chercher d'abord dans les articles
    article = Article.objects.filter(code=code, boutique=boutique).first()
    logger.info(f"   → Article trouvé par code: {article}")
    
    if article:
        # Vérifier si l'article a des variantes
        variantes_actives = article.variantes.filter(est_actif=True)
        variantes = list(variantes_actives.values(
            'id', 'nom_variante', 'code_barre', 'type_attribut'
        ))
        
        # Stock TOUJOURS sur le parent (variants = identifiants uniquement)
        stock_effectif = article.quantite_stock
        
        return JsonResponse({
            'existe': True,
            'type': 'article',
            'article': {
                'id': article.id,
                'nom': article.nom,
                'code': article.code,
                'stock': stock_effectif,  # ⭐ Stock effectif (somme variantes ou stock article)
                'prix': float(article.prix_vente),
                'devise': article.devise,
                'a_variantes': len(variantes) > 0,
                'variantes': variantes
            }
        })
    
    # 2. Sinon, chercher dans les variantes → retourner le PARENT avec SOMME des quantités variantes
    logger.info(f"   → Recherche dans variantes: code_barre='{code}', boutique={boutique.id}")
    variante = VarianteArticle.objects.filter(
        code_barre=code, 
        article_parent__boutique=boutique,
        est_actif=True
    ).select_related('article_parent').first()
    logger.info(f"   → Variante trouvée: {variante}")
    
    if variante:
        parent = variante.article_parent
        # Stock toujours sur le parent
        stock_parent = parent.quantite_stock
        
        return JsonResponse({
            'existe': True,
            'type': 'variante',
            'article': {
                'id': parent.id,
                'nom': parent.nom,
                'code': parent.code,
                'stock': stock_parent,
                'prix': float(parent.prix_vente or 0),
                'devise': parent.devise,
                'a_variantes': True
            },
            'variante': {
                'id': variante.id,
                'nom': variante.nom_variante,
                'nom_complet': variante.nom_complet,
                'code_barre': variante.code_barre,
                'stock': stock_parent,
                'type_attribut': variante.type_attribut
            }
        })
    
    logger.info(f"   → RIEN TROUVÉ pour code='{code}'")
    return JsonResponse({'existe': False})


@login_required
@commercant_required
@boutique_access_required
def modifier_article_existant(request, boutique_id):
    """
    Modifier un article ou une variante: ajouter du stock et/ou modifier le prix.
    Appelé depuis le modal de recherche par code-barres.
    """
    from django.http import JsonResponse
    from inventory.models import MouvementStock, VarianteArticle
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    boutique = request.boutique
    article_id = request.POST.get('article_id')
    variante_id = request.POST.get('variante_id')  # Nouveau paramètre
    quantite = request.POST.get('quantite', 0)
    prix_vente = request.POST.get('prix_vente', '')
    
    try:
        quantite = int(quantite)
    except (ValueError, TypeError):
        quantite = 0
    
    try:
        prix_vente = Decimal(str(prix_vente)) if prix_vente else None
    except (ValueError, TypeError, InvalidOperation):
        prix_vente = None
    
    modifications = []
    nom_element = ""
    
    # === CAS 1: Modification d'une VARIANTE → Stock toujours sur le PARENT ===
    if variante_id:
        try:
            variante = VarianteArticle.objects.select_related('article_parent').get(
                id=variante_id, 
                article_parent__boutique=boutique
            )
        except VarianteArticle.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Variante non trouvée'})
        
        article = variante.article_parent
        nom_element = variante.nom_complet
        
        # Le prix se modifie sur l'article parent (hérité par toutes les variantes)
        if prix_vente is not None and prix_vente > 0:
            ancien_prix = article.prix_vente or Decimal('0')
            if abs(ancien_prix - prix_vente) > Decimal('0.01'):
                article.prix_vente = prix_vente
                article.save()
                modifications.append(f"Prix: {ancien_prix} → {prix_vente}")
        
        # Stock toujours sur le PARENT — la variante est un identifiant de vente
        if quantite > 0:
            ancien_stock = article.quantite_stock
            article.quantite_stock += quantite
            article.save()
            modifications.append(f"Stock parent: {ancien_stock} → {article.quantite_stock} (+{quantite})")
            
            MouvementStock.objects.create(
                article=article,
                type_mouvement='ENTREE',
                quantite=quantite,
                stock_avant=ancien_stock,
                stock_apres=article.quantite_stock,
                reference_document=f"VAR-{variante.id}-{boutique.id}",
                commentaire=f"Ajout stock via variante '{variante.nom_variante}'"
            )
        
        if modifications:
            return JsonResponse({
                'success': True,
                'message': f"Modifié ({variante.nom_variante}): {', '.join(modifications)}",
                'article': {
                    'id': article.id,
                    'nom': article.nom,
                    'stock': article.quantite_stock,
                    'prix': float(article.prix_vente or 0)
                }
            })
        else:
            return JsonResponse({'success': True, 'message': 'Aucune modification effectuée'})
    
    # === CAS 2: Modification d'un ARTICLE (sans variante) ===
    if not article_id:
        return JsonResponse({'success': False, 'message': 'Article non spécifié'})
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Article non trouvé'})
    
    nom_element = article.nom
    
    # Modifier le prix si fourni et différent
    if prix_vente is not None and prix_vente > 0:
        ancien_prix = article.prix_vente or Decimal('0')
        if abs(ancien_prix - prix_vente) > Decimal('0.01'):
            article.prix_vente = prix_vente
            modifications.append(f"Prix: {ancien_prix} → {prix_vente}")
    
    # Ajouter du stock si quantité > 0 → envoyer en validation MAUI
    if quantite > 0:
        article.est_valide_client = False
        article.quantite_envoyee = quantite
        modifications.append(f"Stock: +{quantite} envoyé pour validation client")
    
    if modifications:
        article.save()
        return JsonResponse({
            'success': True,
            'message': f"Article modifié: {', '.join(modifications)}",
            'article': {
                'id': article.id,
                'nom': article.nom,
                'stock': article.quantite_stock,
                'prix': float(article.prix_vente)
            }
        })
    else:
        return JsonResponse({
            'success': True,
            'message': 'Aucune modification effectuée'
        })


@login_required
@commercant_required
@boutique_access_required
def ajouter_article_boutique(request, boutique_id):
    """Ajouter un article à une boutique spécifique (interface commerçant)"""
    from django.http import JsonResponse
    from inventory.models import Article, Categorie
    
    boutique = request.boutique
    
    if request.method == 'POST':
        # Vérifier si c'est une requête AJAX (ajout rapide)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            try:
                # Créer l'article avec les données minimales
                nom = request.POST.get('nom', '').strip()
                prix_achat = request.POST.get('prix_achat', '')
                prix_vente = request.POST.get('prix_vente', '')
                prix_vente_usd = request.POST.get('prix_vente_usd', '').strip()
                date_expiration = request.POST.get('date_expiration', '').strip()
                quantite_stock = request.POST.get('quantite_stock', 0)
                categorie_id = request.POST.get('categorie', '')
                devise = request.POST.get('devise', 'CDF')
                
                # Validations
                errors = {}
                if not nom:
                    errors['nom'] = ['Le nom est requis']
                    
                # Si un code est fourni, vérifier s'il existe déjà
                code = request.POST.get('code', '').strip()
                if code:
                    article_existant = Article.objects.filter(code=code, boutique=boutique).first()
                    if article_existant:
                        # ⭐ Article existe déjà : proposer uniquement l'ajout de stock
                        quantite_ajout = int(quantite_stock) if quantite_stock else 0
                        if quantite_ajout > 0:
                            # Ne PAS ajouter le stock maintenant — sera appliqué après validation MAUI
                            article_existant.est_valide_client = False
                            article_existant.quantite_envoyee = quantite_ajout
                            article_existant.save()
                            
                            return JsonResponse({
                                'success': True,
                                'article_exists': True,
                                'message': f'Ajout de +{quantite_ajout} unités à "{article_existant.nom}" envoyé pour validation client',
                                'article_id': article_existant.id,
                                'article_nom': article_existant.nom,
                                'stock_actuel': article_existant.quantite_stock,
                                'quantite_en_attente': quantite_ajout
                            })
                        else:
                            return JsonResponse({
                                'success': False,
                                'article_exists': True,
                                'message': f'L\'article "{article_existant.nom}" existe déjà. Veuillez indiquer une quantité à ajouter.',
                                'article_id': article_existant.id,
                                'article_nom': article_existant.nom,
                                'stock_actuel': article_existant.quantite_stock
                            })
                
                if errors:
                    return JsonResponse({'success': False, 'errors': errors})
                
                # Générer un code-barres automatique si aucun n'est fourni
                if not code:
                    base_code = f"{boutique.id}-{uuid.uuid4().hex[:8].upper()}"
                    # S'assurer que le code est unique
                    while Article.objects.filter(code=base_code).exists():
                        base_code = f"{boutique.id}-{uuid.uuid4().hex[:8].upper()}"
                    code = base_code
                
                # Créer l'article
                stock_initial = int(quantite_stock) if quantite_stock else 0
                
                # Parser la date d'expiration si fournie
                date_exp_parsed = None
                if date_expiration:
                    try:
                        from datetime import datetime
                        date_exp_parsed = datetime.strptime(date_expiration, '%Y-%m-%d').date()
                    except ValueError:
                        pass
                
                article = Article.objects.create(
                    boutique=boutique,
                    nom=nom,
                    code=code,
                    devise=devise,
                    prix_achat=float(prix_achat) if prix_achat else 0,
                    prix_vente=float(prix_vente) if prix_vente else 0,
                    prix_vente_usd=float(prix_vente_usd) if prix_vente_usd else None,
                    date_expiration=date_exp_parsed,
                    quantite_stock=0,  # Stock à 0 — sera appliqué après validation MAUI
                    est_actif=True,
                    est_valide_client=False,
                    quantite_envoyee=stock_initial
                )
                
                # Ajouter la catégorie si fournie
                if categorie_id:
                    try:
                        categorie = Categorie.objects.get(id=categorie_id, boutique=boutique)
                        article.categorie = categorie
                        article.save()
                    except Categorie.DoesNotExist:
                        pass
                
                # ⭐ Créer les variantes si présentes (AJAX)
                variantes_creees = 0
                variante_keys = [k for k in request.POST.keys() if 'variante' in k]
                print(f"🔍 DEBUG AJAX - POST keys pour variantes: {variante_keys}")
                for key in request.POST:
                    if key.startswith('variante_code_'):
                        idx = key.replace('variante_code_', '')
                        code_barre_v = request.POST.get(f'variante_code_{idx}', '').strip()
                        nom_variante = request.POST.get(f'variante_nom_{idx}', '').strip()
                        type_attribut = request.POST.get(f'variante_type_{idx}', 'AUTRE')
                        stock_variante = request.POST.get(f'variante_stock_{idx}', '0')
                        
                        print(f"🏷️ Variante {idx}: code='{code_barre_v}', nom='{nom_variante}', type='{type_attribut}', stock='{stock_variante}'")
                        
                        if code_barre_v and nom_variante:
                            try:
                                stock_v = int(stock_variante) if stock_variante else 0
                                # Vérifier unicité du code-barres dans la MÊME BOUTIQUE seulement
                                if not VarianteArticle.objects.filter(
                                    code_barre=code_barre_v,
                                    article_parent__boutique=boutique
                                ).exists():
                                    VarianteArticle.objects.create(
                                        article_parent=article,
                                        code_barre=code_barre_v,
                                        nom_variante=nom_variante,
                                        type_attribut=type_attribut,
                                        quantite_stock=max(0, stock_v),
                                        est_actif=True
                                    )
                                    variantes_creees += 1
                            except Exception as e:
                                print(f"Erreur création variante: {e}")
                
                # Générer le code QR automatiquement
                try:
                    generer_qr_code_article(article)
                except Exception as e:
                    print(f"Erreur génération QR: {e}")
                
                msg = f'Article "{article.nom}" ajouté avec succès'
                if variantes_creees > 0:
                    msg += f' avec {variantes_creees} variante(s)'
                
                return JsonResponse({
                    'success': True, 
                    'message': msg,
                    'article_id': article.id,
                    'article_nom': article.nom,
                    'variantes_creees': variantes_creees
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False, 
                    'message': f'Erreur lors de l\'ajout: {str(e)}'
                })
        
        # Si ce n'est pas AJAX, utiliser le formulaire normal
        form = ArticleForm(request.POST, request.FILES)
        if form.is_valid():
            # ⭐ Vérifier si l'article existe déjà par code-barres
            code = form.cleaned_data.get('code')
            if code:
                article_existant = Article.objects.filter(code=code, boutique=boutique).first()
                if article_existant:
                    # Article existe : envoyer pour validation MAUI (stock pas appliqué maintenant)
                    quantite_ajout = form.cleaned_data.get('quantite_stock', 0)
                    if quantite_ajout > 0:
                        article_existant.est_valide_client = False
                        article_existant.quantite_envoyee = quantite_ajout
                        article_existant.save()
                        
                        messages.success(request, f'✅ Ajout de +{quantite_ajout} unités à "{article_existant.nom}" envoyé pour validation client')
                    else:
                        messages.warning(request, f'⚠️ L\'article "{article_existant.nom}" existe déjà (Stock actuel: {article_existant.quantite_stock}). Veuillez indiquer une quantité à ajouter.')
                    
                    return redirect('inventory:entrer_boutique', boutique_id=boutique.id)
            
            # Nouvel article : créer normalement
            article = form.save(commit=False)
            article.boutique = boutique
            stock_initial = article.quantite_stock
            article.quantite_stock = 0  # Stock à 0 — sera appliqué après validation MAUI
            article.est_valide_client = False
            article.quantite_envoyee = stock_initial
            article.save()
            
            # ⭐ Créer les variantes si présentes dans le formulaire
            # Stock géré au niveau de chaque variante (enfant)
            variantes_creees = 0
            variante_keys = [k for k in request.POST.keys() if 'variante' in k]
            print(f"🔍 DEBUG FORM POST - variante keys: {variante_keys}")
            for key in request.POST:
                if key.startswith('variante_code_'):
                    idx = key.replace('variante_code_', '')
                    code_barre = request.POST.get(f'variante_code_{idx}', '').strip()
                    nom_variante = request.POST.get(f'variante_nom_{idx}', '').strip()
                    type_attribut = request.POST.get(f'variante_type_{idx}', 'AUTRE')
                    stock_variante = request.POST.get(f'variante_stock_{idx}', '0')
                    
                    # Valider et créer la variante avec son stock
                    if code_barre and nom_variante:
                        try:
                            stock_v = int(stock_variante) if stock_variante else 0
                            # Vérifier unicité du code-barres dans la MÊME BOUTIQUE seulement
                            if not VarianteArticle.objects.filter(
                                code_barre=code_barre,
                                article_parent__boutique=boutique
                            ).exists():
                                VarianteArticle.objects.create(
                                    article_parent=article,
                                    code_barre=code_barre,
                                    nom_variante=nom_variante,
                                    type_attribut=type_attribut,
                                    quantite_stock=max(0, stock_v),
                                    est_actif=True
                                )
                                variantes_creees += 1
                        except Exception as e:
                            print(f"Erreur création variante: {e}")
            
            # Générer le code QR automatiquement
            try:
                generer_qr_code_article(article)
                msg = f'Article "{article.nom}" ajouté avec succès à {boutique.nom}.'
                if variantes_creees > 0:
                    msg += f' {variantes_creees} variante(s) créée(s).'
                messages.success(request, msg)
            except Exception as e:
                messages.warning(request, f'Article "{article.nom}" ajouté, mais erreur lors de la génération du code QR: {str(e)}')
            
            return redirect('inventory:entrer_boutique', boutique_id=boutique.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        form = ArticleForm()
    
    # Récupérer les catégories de la boutique
    categories = boutique.categories.all()
    
    context = {
        'form': form,
        'boutique': boutique,
        'categories': categories
    }
    
    return render(request, 'inventory/commercant/ajouter_article.html', context)

@login_required
@commercant_required
@boutique_access_required
def modifier_article_boutique(request, boutique_id, article_id):
    """Modifier un article d'une boutique spécifique"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
    except Article.DoesNotExist:
        messages.error(request, "Article introuvable.")
        return redirect('inventory:commercant_articles_boutique', boutique_id=boutique.id)
    
    if request.method == 'POST':
        # Conserver le code-barres existant (on ne souhaite plus le modifier via ce formulaire)
        ancien_code = article.code
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            article = form.save(commit=False)
            article.boutique = boutique
            # Réappliquer le code d'origine pour éviter de le vider ou de le dupliquer
            article.code = ancien_code
            article.save()
            
            # Régénérer le code QR si le code a changé
            try:
                generer_qr_code_article(article)
                messages.success(request, f'Article "{article.nom}" modifié avec succès.')
            except Exception as e:
                messages.warning(request, f'Article modifié, mais erreur lors de la génération du code QR: {str(e)}')
            
            return redirect('inventory:commercant_articles_boutique', boutique_id=boutique.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erreur dans {field}: {error}")
    else:
        form = ArticleForm(instance=article)
    
    # Récupérer les catégories de la boutique
    categories = boutique.categories.all()
    
    # Récupérer les variantes de cet article
    variantes = article.variantes.all().order_by('nom_variante')
    variante_form = VarianteArticleForm()
    
    # Stock toujours sur le parent
    stock_total_variantes = article.quantite_stock
    
    context = {
        'form': form,
        'boutique': boutique,
        'categories': categories,
        'article': article,
        'mode_edition': True,
        # Variantes
        'variantes': variantes,
        'variante_form': variante_form,
        'stock_total_variantes': stock_total_variantes,
        'has_variantes': variantes.exists(),
    }
    
    return render(request, 'inventory/commercant/modifier_article.html', context)

@login_required
@commercant_required
@boutique_access_required
def bulk_delete_articles(request, boutique_id):
    """Suppression multiple d'articles (AJAX POST JSON)"""
    from django.http import JsonResponse
    from django.views.decorators.csrf import csrf_exempt
    import json
    
    boutique = request.boutique
    
    if request.method != 'POST' or request.headers.get('X-Requested-With') != 'XMLHttpRequest':
        return JsonResponse({'success': False, 'message': 'Requête invalide.'})
    
    try:
        data = json.loads(request.body)
        article_ids = data.get('article_ids', [])
        if not isinstance(article_ids, list):
            raise ValueError('article_ids doit être une liste')
        # Convertir en entiers
        article_ids = [int(i) for i in article_ids]
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Données invalides : {str(e)}'})
    
    if not article_ids:
        return JsonResponse({'success': False, 'message': 'Aucun article à supprimer.'})
    
    # Récupérer les articles appartenant à cette boutique
    articles = Article.objects.filter(id__in=article_ids, boutique=boutique)
    if not articles.exists():
        return JsonResponse({'success': False, 'message': 'Aucun article trouvé pour cette boutique.'})
    
    deleted_count = 0
    failed = []
    for article in articles:
        try:
            article.delete()
            deleted_count += 1
        except Exception as e:
            failed.append(f"{article.nom}: {str(e)}")
    
    if failed:
        message = f'{deleted_count} article(s) supprimé(s). Erreurs : {"; ".join(failed)}'
        return JsonResponse({'success': False, 'message': message})
    else:
        return JsonResponse({'success': True, 'message': f'{deleted_count} article(s) supprimé(s) avec succès.'})


# ===== IMPORT ARTICLES ENTRE BOUTIQUES =====

@login_required
@commercant_required
@boutique_access_required
def importer_articles_entre_boutiques(request, boutique_id):
    """Importer des articles depuis un autre point de vente (même commerçant), sans quantité ni prix"""
    commercant = request.user.profil_commercant
    boutique = request.boutique
    
    # Autres boutiques du même commerçant (exclure la boutique actuelle)
    autres_boutiques = commercant.boutiques.filter(
        est_active=True
    ).exclude(id=boutique.id).order_by('nom')
    
    # Boutique source sélectionnée
    source_id = request.GET.get('source', '')
    boutique_source = None
    articles_source = []
    
    if source_id:
        boutique_source = autres_boutiques.filter(id=source_id).first()
        if boutique_source:
            # Codes déjà présents dans la boutique destination
            codes_existants = set(Article.objects.filter(
                boutique=boutique, est_actif=True
            ).values_list('code', flat=True))
            
            articles_source = Article.objects.filter(
                boutique=boutique_source, est_actif=True
            ).select_related('categorie').order_by('nom')
            
            # Marquer ceux déjà présents
            for art in articles_source:
                art.deja_present = art.code in codes_existants
    
    if request.method == 'POST':
        source_post_id = request.POST.get('boutique_source')
        if not source_post_id:
            messages.error(request, "Veuillez sélectionner un point de vente source")
            return redirect('inventory:importer_articles_entre_boutiques', boutique_id=boutique.id)
        
        boutique_src = autres_boutiques.filter(id=source_post_id).first()
        if not boutique_src:
            messages.error(request, "Point de vente source introuvable")
            return redirect('inventory:importer_articles_entre_boutiques', boutique_id=boutique.id)
        
        articles_selectionnes = request.POST.getlist('articles_selectionnes')
        if not articles_selectionnes:
            messages.error(request, "Veuillez sélectionner au moins un article à importer")
            return redirect(f"{request.path}?source={source_post_id}")
        
        codes_existants = set(Article.objects.filter(
            boutique=boutique, est_actif=True
        ).values_list('code', flat=True))
        
        importes = 0
        erreurs = []
        
        try:
            with transaction.atomic():
                for article_id in articles_selectionnes:
                    try:
                        art_src = Article.objects.get(id=article_id, boutique=boutique_src)
                        
                        if art_src.code in codes_existants:
                            erreurs.append(f"{art_src.nom} (code: {art_src.code}): existe déjà")
                            continue
                        
                        # Copier ou récupérer la catégorie dans la boutique destination
                        categorie_dest = None
                        if art_src.categorie:
                            categorie_dest, _ = Categorie.objects.get_or_create(
                                nom__iexact=art_src.categorie.nom,
                                boutique=boutique,
                                defaults={'nom': art_src.categorie.nom}
                            )
                        
                        nouvel_article = Article.objects.create(
                            code=art_src.code,
                            nom=art_src.nom,
                            description=art_src.description,
                            devise=art_src.devise,
                            prix_achat=0,
                            prix_vente=0,
                            categorie=categorie_dest,
                            boutique=boutique,
                            quantite_stock=0,
                            est_actif=True
                        )
                        
                        # Copier les variantes actives (sans stock)
                        variantes_src = art_src.variantes.filter(est_actif=True)
                        for var_src in variantes_src:
                            # Vérifier que le code-barres n'existe pas déjà dans cette boutique
                            if not VarianteArticle.objects.filter(
                                code_barre=var_src.code_barre,
                                article_parent__boutique=boutique
                            ).exists():
                                VarianteArticle.objects.create(
                                    article_parent=nouvel_article,
                                    code_barre=var_src.code_barre,
                                    nom_variante=var_src.nom_variante,
                                    type_attribut=var_src.type_attribut,
                                    quantite_stock=0,
                                    est_actif=True
                                )
                        
                        codes_existants.add(art_src.code)
                        importes += 1
                        
                    except Article.DoesNotExist:
                        erreurs.append(f"Article ID {article_id} introuvable")
                    except Exception as e:
                        erreurs.append(f"Erreur: {str(e)}")
                
                if importes == 0 and erreurs:
                    raise Exception("Aucun article importé")
        
        except Exception:
            for err in erreurs:
                messages.error(request, err)
            return redirect(f"{request.path}?source={source_post_id}")
        
        if importes > 0:
            messages.success(request, f"{importes} article(s) importé(s) avec succès (sans quantité ni prix)")
        for err in erreurs:
            messages.warning(request, err)
        
        return redirect('inventory:commercant_articles_boutique', boutique_id=boutique.id)
    
    context = {
        'boutique': boutique,
        'autres_boutiques': autres_boutiques,
        'boutique_source': boutique_source,
        'articles_source': articles_source,
        'commercant': commercant,
    }
    return render(request, 'inventory/commercant/importer_articles_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def sync_variantes_entre_boutiques(request, boutique_id):
    """Synchroniser les variantes manquantes depuis un autre point de vente"""
    commercant = request.user.profil_commercant
    boutique = request.boutique
    
    if request.method != 'POST':
        return redirect('inventory:importer_articles_entre_boutiques', boutique_id=boutique.id)
    
    source_id = request.POST.get('boutique_source')
    if not source_id:
        messages.error(request, "Veuillez sélectionner un point de vente source")
        return redirect('inventory:importer_articles_entre_boutiques', boutique_id=boutique.id)
    
    boutique_src = get_object_or_404(Boutique, id=source_id, commercant=commercant)
    
    # Trouver les articles communs (même code) entre source et destination
    articles_dest = Article.objects.filter(boutique=boutique, est_actif=True)
    codes_dest = {a.code: a for a in articles_dest}
    
    articles_src = Article.objects.filter(
        boutique=boutique_src, est_actif=True
    ).prefetch_related('variantes')
    
    variantes_ajoutees = 0
    erreurs = []
    
    try:
        with transaction.atomic():
            for art_src in articles_src:
                art_dest = codes_dest.get(art_src.code)
                if not art_dest:
                    continue  # Article pas présent dans la destination
                
                # Codes-barres déjà existants dans la destination (par boutique)
                for var_src in art_src.variantes.filter(est_actif=True):
                    if VarianteArticle.objects.filter(
                        code_barre=var_src.code_barre,
                        article_parent__boutique=boutique_dest
                    ).exists():
                        continue  # Variante déjà existante dans cette boutique
                    
                    try:
                        VarianteArticle.objects.create(
                            article_parent=art_dest,
                            code_barre=var_src.code_barre,
                            nom_variante=var_src.nom_variante,
                            type_attribut=var_src.type_attribut,
                            quantite_stock=0,
                            est_actif=True
                        )
                        variantes_ajoutees += 1
                    except Exception as e:
                        erreurs.append(f"Variante {var_src.nom_variante} ({var_src.code_barre}): {str(e)}")
    
    except Exception:
        for err in erreurs:
            messages.error(request, err)
        return redirect(f"/commercant/boutiques/{boutique.id}/articles/importer/?source={source_id}")
    
    if variantes_ajoutees > 0:
        messages.success(request, f"{variantes_ajoutees} variante(s) synchronisée(s) avec succès (stock = 0)")
    else:
        messages.info(request, "Aucune nouvelle variante à synchroniser — tout est déjà à jour")
    
    for err in erreurs:
        messages.warning(request, err)
    
    return redirect(f"/commercant/boutiques/{boutique.id}/articles/importer/?source={source_id}")


# ===== GESTION DES VARIANTES D'ARTICLES =====

@login_required
@commercant_required
@boutique_access_required
def ajouter_variante(request, boutique_id, article_id):
    """Ajouter une variante à un article (AJAX POST)"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
    except Article.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Article introuvable.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})
    
    # Récupérer les données du formulaire
    code_barre = request.POST.get('code_barre', '').strip()
    nom_variante = request.POST.get('nom_variante', '').strip()
    type_attribut = request.POST.get('type_attribut', 'AUTRE')
    quantite_stock = request.POST.get('quantite_stock', 0)
    
    # Validation
    if not code_barre:
        return JsonResponse({'success': False, 'message': 'Le code-barres est obligatoire.'})
    if not nom_variante:
        return JsonResponse({'success': False, 'message': 'Le nom de la variante est obligatoire.'})
    
    # Vérifier que le code-barres n'existe pas déjà dans cette boutique
    if VarianteArticle.objects.filter(code_barre=code_barre, article_parent__boutique=boutique).exists():
        return JsonResponse({'success': False, 'message': f'Le code-barres "{code_barre}" existe déjà dans cette boutique.'})
    
    # Vérifier que le code-barres n'est pas utilisé par un article
    if Article.objects.filter(code=code_barre, boutique=boutique).exists():
        return JsonResponse({'success': False, 'message': f'Ce code-barres est déjà utilisé par un article.'})
    
    try:
        quantite_stock = int(quantite_stock)
        if quantite_stock < 0:
            quantite_stock = 0
    except (ValueError, TypeError):
        quantite_stock = 0
    
    # Créer la variante
    try:
        variante = VarianteArticle.objects.create(
            article_parent=article,
            code_barre=code_barre,
            nom_variante=nom_variante,
            type_attribut=type_attribut,
            quantite_stock=quantite_stock,
            est_actif=True
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Variante "{nom_variante}" ajoutée avec succès.',
            'variante': {
                'id': variante.id,
                'code_barre': variante.code_barre,
                'nom_variante': variante.nom_variante,
                'type_attribut': variante.get_type_attribut_display(),
                'quantite_stock': variante.quantite_stock,
                'prix_vente': float(variante.prix_vente),
                'est_actif': variante.est_actif
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


@login_required
@commercant_required
@boutique_access_required
def modifier_variante(request, boutique_id, article_id, variante_id):
    """Modifier une variante d'article (AJAX POST)"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
        variante = VarianteArticle.objects.get(id=variante_id, article_parent=article)
    except (Article.DoesNotExist, VarianteArticle.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Variante introuvable.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})
    
    # Récupérer les données
    nom_variante = request.POST.get('nom_variante', '').strip()
    type_attribut = request.POST.get('type_attribut', variante.type_attribut)
    est_actif = request.POST.get('est_actif') == 'true'
    
    if not nom_variante:
        return JsonResponse({'success': False, 'message': 'Le nom de la variante est obligatoire.'})
    
    try:
        variante.nom_variante = nom_variante
        variante.type_attribut = type_attribut
        variante.est_actif = est_actif
        variante.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Variante "{nom_variante}" modifiée avec succès.',
            'variante': {
                'id': variante.id,
                'nom_variante': variante.nom_variante,
                'type_attribut': variante.get_type_attribut_display(),
                'est_actif': variante.est_actif
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


@login_required
@commercant_required
@boutique_access_required
def supprimer_variante(request, boutique_id, article_id, variante_id):
    """Supprimer une variante d'article (AJAX POST)"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
        variante = VarianteArticle.objects.get(id=variante_id, article_parent=article)
    except (Article.DoesNotExist, VarianteArticle.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Variante introuvable.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})
    
    nom_variante = variante.nom_variante
    try:
        variante.delete()
        return JsonResponse({
            'success': True,
            'message': f'Variante "{nom_variante}" supprimée avec succès.'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


@login_required
@commercant_required
@boutique_access_required
def ajuster_stock_variante(request, boutique_id, article_id, variante_id):
    """Ajuster le stock d'une variante (AJAX POST)"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
        variante = VarianteArticle.objects.get(id=variante_id, article_parent=article)
    except (Article.DoesNotExist, VarianteArticle.DoesNotExist):
        return JsonResponse({'success': False, 'message': 'Variante introuvable.'})
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})
    
    try:
        quantite = int(request.POST.get('quantite', 0))
        operation = request.POST.get('operation', 'ajouter')  # 'ajouter' ou 'retirer'
        
        stock_avant = variante.quantite_stock
        
        if operation == 'retirer':
            quantite = -abs(quantite)
        else:
            quantite = abs(quantite)
        
        nouveau_stock = stock_avant + quantite
        
        if nouveau_stock < 0:
            return JsonResponse({
                'success': False,
                'message': f'Stock insuffisant. Stock actuel: {stock_avant}'
            })
        
        variante.quantite_stock = nouveau_stock
        variante.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Stock ajusté: {stock_avant} → {nouveau_stock}',
            'stock_avant': stock_avant,
            'stock_apres': nouveau_stock
        })
        
    except (ValueError, TypeError) as e:
        return JsonResponse({'success': False, 'message': 'Quantité invalide.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


@login_required
@commercant_required
@boutique_access_required
def supprimer_article_boutique(request, boutique_id, article_id):
    """Supprimer un article d'une boutique spécifique (interface commerçant)"""
    boutique = request.boutique
    
    try:
        article = Article.objects.get(id=article_id, boutique=boutique)
    except Article.DoesNotExist:
        messages.error(request, "Article introuvable.")
        return redirect('inventory:commercant_articles_boutique', boutique_id=boutique.id)
    
    if request.method == 'POST':
        nom_article = article.nom
        try:
            article.delete()
            messages.success(request, f"Article '{nom_article}' supprimé avec succès!")
        except Exception as e:
            messages.error(request, f"Impossible de supprimer l'article : {str(e)}")
    
    return redirect('inventory:commercant_articles_boutique', boutique_id=boutique.id)

@login_required
@commercant_required
@boutique_access_required
def ajuster_stock_article(request, boutique_id, article_id):
    """Ajuster rapidement le stock d'un article avec traçabilité complète"""
    from django.http import JsonResponse
    
    boutique = request.boutique
    
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            article = Article.objects.get(id=article_id, boutique=boutique)
            
            type_ajustement = request.POST.get('type_ajustement')
            quantite = int(request.POST.get('quantite', 0))
            commentaire = request.POST.get('commentaire', '')
            
            # Capturer le stock avant modification
            stock_avant = article.quantite_stock
            
            # Calculer la différence et le type de mouvement
            if type_ajustement == 'ajouter':
                # Envoyer en validation MAUI au lieu d'appliquer directement
                article.est_valide_client = False
                article.quantite_envoyee = quantite
                article.save(update_fields=['est_valide_client', 'quantite_envoyee'])
                return JsonResponse({
                    'success': True,
                    'message': f'+{quantite} unités envoyées pour validation client (stock actuel: {stock_avant})',
                    'stock_avant': stock_avant,
                    'stock_apres': stock_avant,
                    'quantite_en_attente': quantite
                })
            elif type_ajustement == 'retirer':
                article.quantite_stock = max(0, article.quantite_stock - quantite)
                type_mouvement = 'SORTIE'
                quantite_mouvement = -(stock_avant - article.quantite_stock)
            elif type_ajustement == 'definir':
                article.quantite_stock = quantite
                difference = quantite - stock_avant
                type_mouvement = 'AJUSTEMENT'
                quantite_mouvement = difference
            else:
                return JsonResponse({'success': False, 'message': 'Type d\'ajustement invalide'})
            
            article.save(update_fields=['quantite_stock'])
            
            # ⭐ Créer le mouvement de stock pour traçabilité
            MouvementStock.objects.create(
                article=article,
                type_mouvement=type_mouvement,
                quantite=quantite_mouvement,
                stock_avant=stock_avant,
                stock_apres=article.quantite_stock,
                reference_document=f"AJUST-{boutique.code_boutique}-{article.id}",
                utilisateur=request.user.username,
                commentaire=commentaire or f"Ajustement {type_ajustement}: {stock_avant} → {article.quantite_stock}"
            )
            
            return JsonResponse({
                'success': True,
                'message': f'Stock ajusté: {stock_avant} → {article.quantite_stock}',
                'ancien_stock': stock_avant,
                'nouveau_stock': article.quantite_stock
            })
            
        except Article.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Article introuvable'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
@commercant_required
@boutique_access_required
def modifier_prix_article(request, boutique_id, article_id):
    """Modifier rapidement le prix d'un article"""
    from django.http import JsonResponse
    
    boutique = request.boutique
    
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            article = Article.objects.get(id=article_id, boutique=boutique)
            
            nouveau_prix = Decimal(str(request.POST.get('prix_vente', 0) or 0))
            commentaire = request.POST.get('commentaire', '')
            
            if nouveau_prix < 0:
                return JsonResponse({'success': False, 'message': 'Le prix ne peut pas être négatif'})
            
            ancien_prix = article.prix_vente or Decimal('0')
            article.prix_vente = nouveau_prix
            article.save()
            
            # Régénérer le QR code avec le nouveau prix
            try:
                generer_qr_code_article(article)
            except Exception:
                pass  # Ne pas bloquer si le QR code échoue
            
            return JsonResponse({
                'success': True,
                'message': f'Prix modifié: {ancien_prix} → {nouveau_prix} CDF',
                'ancien_prix': float(ancien_prix),
                'nouveau_prix': float(nouveau_prix)
            })
            
        except Article.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Article introuvable'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@login_required
@commercant_required
@boutique_access_required
def generer_pdf_qr_codes(request, boutique_id):
    """Générer un PDF avec tous les codes QR des articles d'une boutique"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    import os
    
    boutique = request.boutique
    
    # Récupérer tous les articles actifs de la boutique avec QR code
    articles = boutique.articles.filter(est_actif=True, qr_code__isnull=False).exclude(qr_code='')
    
    if not articles.exists():
        messages.warning(request, "Aucun article avec code QR trouvé dans cette boutique.")
        return redirect('inventory:entrer_boutique', boutique_id=boutique.id)
    
    # Créer le buffer pour le PDF
    buffer = io.BytesIO()
    
    # Créer le document PDF
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []
    
    # Titre
    title = Paragraph(f"Codes QR - {boutique.nom}", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Informations boutique
    info_text = f"""
    <b>Boutique:</b> {boutique.nom}<br/>
    <b>Type:</b> {boutique.get_type_commerce_display()}<br/>
    <b>Adresse:</b> {boutique.adresse}, {boutique.ville}<br/>
    <b>Date de génération:</b> {datetime.now().strftime('%d/%m/%Y à %H:%M')}<br/>
    <b>Nombre d'articles:</b> {articles.count()}<br/>
    """
    info = Paragraph(info_text, styles['Normal'])
    story.append(info)
    story.append(Spacer(1, 30))
    
    # Créer une grille de QR codes (3 colonnes)
    qr_data = []
    current_row = []
    
    for i, article in enumerate(articles):
        try:
            # Vérifier que le fichier QR code existe
            if article.qr_code and os.path.exists(article.qr_code.path):
                # Créer une cellule avec QR code + informations
                cell_content = [
                    ReportLabImage(article.qr_code.path, width=4*cm, height=4*cm),
                    Paragraph(f"<b>{article.nom}</b>", styles['Normal']),
                    Paragraph(f"Code: {article.code}", styles['Normal']),
                    Paragraph(f"Prix: {article.prix_vente} CDF", styles['Normal'])
                ]
                current_row.append(cell_content)
            else:
                # Si pas de QR code, générer un nouveau
                if generer_qr_code_article(article):
                    cell_content = [
                        ReportLabImage(article.qr_code.path, width=4*cm, height=4*cm),
                        Paragraph(f"<b>{article.nom}</b>", styles['Normal']),
                        Paragraph(f"Code: {article.code}", styles['Normal']),
                        Paragraph(f"Prix: {article.prix_vente} CDF", styles['Normal'])
                    ]
                    current_row.append(cell_content)
                else:
                    # QR code manquant
                    cell_content = [
                        Paragraph("QR Code<br/>non disponible", styles['Normal']),
                        Paragraph(f"<b>{article.nom}</b>", styles['Normal']),
                        Paragraph(f"Code: {article.code}", styles['Normal']),
                        Paragraph(f"Prix: {article.prix_vente} CDF", styles['Normal'])
                    ]
                    current_row.append(cell_content)
            
            # Ajouter la ligne quand on a 3 colonnes ou à la fin
            if len(current_row) == 3 or i == len(articles) - 1:
                # Compléter la ligne si nécessaire
                while len(current_row) < 3:
                    empty_cell = [
                        Paragraph("", styles['Normal']),
                        Paragraph("", styles['Normal']),
                        Paragraph("", styles['Normal']),
                        Paragraph("", styles['Normal'])
                    ]
                    current_row.append(empty_cell)
                qr_data.append(current_row)
                current_row = []
                
        except Exception as e:
            print(f"Erreur avec l'article {article.id}: {str(e)}")
            continue
    
    # Créer le tableau avec les QR codes
    if qr_data:
        table = Table(qr_data, colWidths=[6*cm, 6*cm, 6*cm])
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(table)
    
    # Construire le PDF
    doc.build(story)
    
    # Préparer la réponse
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="QR_Codes_{boutique.nom}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    return response

@login_required
@commercant_required
@boutique_access_required
def modifier_terminal(request, boutique_id, terminal_id):
    """Modifier un terminal MAUI"""
    boutique = request.boutique
    terminal = get_object_or_404(Client, id=terminal_id, boutique=boutique)
    
    if request.method == 'POST':
        terminal.nom_terminal = request.POST.get('nom_terminal', terminal.nom_terminal)
        terminal.description = request.POST.get('description', '')
        terminal.notes = request.POST.get('notes', '')
        terminal.save()
        
        messages.success(request, f"Terminal '{terminal.nom_terminal}' modifié avec succès!")
        return redirect('inventory:commercant_terminaux_boutique', boutique_id=boutique.id)
    
    context = {
        'boutique': boutique,
        'terminal': terminal
    }
    
    return render(request, 'inventory/commercant/modifier_terminal.html', context)

@login_required
@commercant_required
@boutique_access_required
def toggle_terminal(request, boutique_id, terminal_id):
    """Activer/Désactiver un terminal MAUI"""
    boutique = request.boutique
    terminal = get_object_or_404(Client, id=terminal_id, boutique=boutique)
    
    terminal.est_actif = not terminal.est_actif
    terminal.save()
    
    statut = "activé" if terminal.est_actif else "désactivé"
    messages.success(request, f"Terminal '{terminal.nom_terminal}' {statut} avec succès!")
    
    return redirect('inventory:commercant_terminaux_boutique', boutique_id=boutique.id)

@login_required
@commercant_required
@boutique_access_required
def supprimer_terminal(request, boutique_id, terminal_id):
    """Supprimer un terminal MAUI"""
    boutique = request.boutique
    terminal = get_object_or_404(Client, id=terminal_id, boutique=boutique)
    
    if request.method == 'POST':
        nom_terminal = terminal.nom_terminal
        terminal.delete()
        messages.success(request, f"Terminal '{nom_terminal}' supprimé avec succès!")
        return redirect('inventory:commercant_terminaux_boutique', boutique_id=boutique.id)
    
    context = {
        'boutique': boutique,
        'terminal': terminal
    }
    
    return render(request, 'inventory/commercant/supprimer_terminal.html', context)

@login_required
@commercant_required
@boutique_access_required
def ventes_boutique(request, boutique_id):
    """Afficher les ventes d'une boutique spécifique"""
    boutique = request.boutique
    
    # ⭐ ISOLATION: Récupérer UNIQUEMENT les ventes de CETTE boutique
    ventes = Vente.objects.filter(
        boutique=boutique  # Filtrage direct par boutique
    ).select_related('client_maui', 'boutique').prefetch_related('lignes__article').order_by('-date_vente')
    terminaux = boutique.clients.all().order_by('nom_terminal')
    
    # Filtres optionnels
    periode = request.GET.get('periode') or 'CUSTOM'
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    terminal_id = request.GET.get('terminal_id')

    # Calculer les dates selon la période sélectionnée
    today = timezone.now().date()
    if periode == 'TODAY':
        date_debut = today.isoformat()
        date_fin = today.isoformat()
    elif periode == 'YESTERDAY':
        jour = today - timedelta(days=1)
        date_debut = jour.isoformat()
        date_fin = jour.isoformat()
    elif periode == 'THIS_MONTH':
        premier_jour = today.replace(day=1)
        date_debut = premier_jour.isoformat()
        date_fin = today.isoformat()

    # Appliquer les filtres de dates
    if date_debut:
        try:
            date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d')
            ventes = ventes.filter(date_vente__gte=date_debut_obj)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d') + timedelta(days=1)
            ventes = ventes.filter(date_vente__lt=date_fin_obj)
        except ValueError:
            pass

    # Filtre par terminal (client MAUI)
    if terminal_id:
        if terminal_id != 'TOUS':
            try:
                terminal_id_int = int(terminal_id)
                ventes = ventes.filter(client_maui_id=terminal_id_int)
                terminal_selected = terminal_id_int
            except (TypeError, ValueError):
                terminal_selected = 'TOUS'
        else:
            terminal_selected = 'TOUS'
    else:
        terminal_selected = 'TOUS'

    # Filtre par catégorie d'article
    categorie_id = request.GET.get('categorie_id', '').strip()
    categorie_selected = ''
    if categorie_id:
        try:
            categorie_id_int = int(categorie_id)
            ventes = ventes.filter(lignes__article__categorie_id=categorie_id_int).distinct()
            categorie_selected = categorie_id_int
        except (TypeError, ValueError):
            pass

    # Catégories disponibles pour le filtre
    categories = boutique.categories.all().order_by('nom')

    # Statistiques globales
    stats = ventes.aggregate(
        total_ventes=Count('id'),
        chiffre_affaires=Sum('montant_total')
    )

    # Statistiques par catégorie (nb ventes distinctes + montant des lignes)
    stats_par_categorie = LigneVente.objects.filter(
        vente__in=ventes,
        vente__est_annulee=False,
    ).values(
        cat_nom=F('article__categorie__nom'),
        cat_id=F('article__categorie_id'),
    ).annotate(
        nb_ventes=Count('vente', distinct=True),
        montant=Sum('sous_total'),
        nb_lignes=Count('id'),
    ).order_by('-montant')
    
    # Regrouper les ventes par période
    from collections import OrderedDict
    maintenant = datetime.now()
    aujourd_hui = maintenant.date()
    
    ventes_groupees = OrderedDict()
    periodes_ordre = [
        "Aujourd'hui",
        "Hier",
        "Cette semaine",
        "Ce mois",
        "Mois précédent"
    ]
    
    for vente in ventes:
        # Déterminer la période
        date_vente = vente.date_vente.date()
        
        if date_vente == aujourd_hui:
            periode = "Aujourd'hui"
        elif date_vente == aujourd_hui - timedelta(days=1):
            periode = "Hier"
        elif date_vente > aujourd_hui - timedelta(days=7):
            periode = "Cette semaine"
        elif vente.date_vente.year == aujourd_hui.year and vente.date_vente.month == aujourd_hui.month:
            periode = "Ce mois"
        else:
            mois_precedent = aujourd_hui.replace(day=1) - timedelta(days=1)
            if vente.date_vente.year == mois_precedent.year and vente.date_vente.month == mois_precedent.month:
                periode = "Mois précédent"
            else:
                mois_fr = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 
                          'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
                periode = f"{mois_fr[vente.date_vente.month - 1]} {vente.date_vente.year}"
        
        if periode not in ventes_groupees:
            ventes_groupees[periode] = []
        ventes_groupees[periode].append(vente)
    
    # Trier les périodes dans l'ordre souhaité
    ventes_groupees_triees = OrderedDict()
    for periode in periodes_ordre:
        if periode in ventes_groupees:
            ventes_groupees_triees[periode] = ventes_groupees[periode]
    # Ajouter les autres périodes (mois passés)
    for periode, ventes_list in ventes_groupees.items():
        if periode not in ventes_groupees_triees:
            ventes_groupees_triees[periode] = ventes_list
    
    context = {
        'boutique': boutique,
        'ventes': ventes,
        'ventes_groupees': ventes_groupees_triees,
        'total_ventes': stats['total_ventes'] or 0,
        'chiffre_affaires': stats['chiffre_affaires'] or 0,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'terminaux': terminaux,
        'terminal_selected': terminal_selected,
        'periode_selected': periode,
        'categories': categories,
        'categorie_selected': categorie_selected,
        'stats_par_categorie': stats_par_categorie,
    }
    
    return render(request, 'inventory/commercant/ventes_boutique.html', context)


# ===== VENTES REFUSÉES =====

@login_required
@commercant_required
@boutique_access_required
def ventes_refusees_boutique(request, boutique_id):
    """Affiche les ventes refusées de la boutique avec statistiques"""
    boutique = request.boutique
    aujourd_hui = timezone.now().date()
    
    # Récupérer les ventes refusées de la boutique
    ventes_refusees = VenteRejetee.objects.filter(
        boutique=boutique
    ).select_related('terminal').order_by('-date_tentative')
    
    # Filtrer par date si demandé
    date_filter = request.GET.get('date', '')
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            ventes_refusees = ventes_refusees.filter(date_tentative__date=date_obj)
        except ValueError:
            pass
    
    # Ventes refusées du jour
    ventes_refusees_jour = VenteRejetee.objects.filter(
        boutique=boutique,
        date_tentative__date=aujourd_hui
    )
    
    # Calculer la somme des articles refusés du jour (extraire du JSON)
    total_refusees_jour = Decimal('0')
    for vente in ventes_refusees_jour:
        try:
            donnees = vente.donnees_vente
            if isinstance(donnees, dict):
                montant = donnees.get('montant_total', 0)
                total_refusees_jour += Decimal(str(montant))
        except (TypeError, ValueError, KeyError):
            pass
    
    # Recette normale du jour (ventes payées)
    recette_jour = Vente.objects.filter(
        client_maui__boutique=boutique,
        date_vente__date=aujourd_hui,
        paye=True
    ).aggregate(total=Sum('montant_total'))['total'] or Decimal('0')
    
    # Total combiné (recette + ventes refusées)
    total_combine = recette_jour + total_refusees_jour
    
    # Statistiques par raison de rejet
    stats_par_raison = ventes_refusees.values('raison_rejet').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Mapping des raisons pour l'affichage
    raisons_display = dict(VenteRejetee.RAISONS_REJET)
    for stat in stats_par_raison:
        stat['raison_display'] = raisons_display.get(stat['raison_rejet'], stat['raison_rejet'])
    
    # Préparer les données détaillées des ventes refusées
    ventes_details = []
    for vente in ventes_refusees[:50]:  # Limiter à 50 pour la performance
        try:
            donnees = vente.donnees_vente
            montant = Decimal('0')
            articles_info = []
            
            if isinstance(donnees, dict):
                montant = Decimal(str(donnees.get('montant_total', 0)))
                lignes = donnees.get('lignes', donnees.get('articles', []))
                for ligne in lignes:
                    if isinstance(ligne, dict):
                        articles_info.append({
                            'nom': ligne.get('nom', ligne.get('article_nom', 'N/A')),
                            'quantite': ligne.get('quantite', 1),
                            'prix': ligne.get('prix_unitaire', ligne.get('prix', 0))
                        })
            
            ventes_details.append({
                'vente': vente,
                'montant': montant,
                'articles': articles_info,
                'raison_display': raisons_display.get(vente.raison_rejet, vente.raison_rejet)
            })
        except (TypeError, ValueError, KeyError):
            ventes_details.append({
                'vente': vente,
                'montant': Decimal('0'),
                'articles': [],
                'raison_display': raisons_display.get(vente.raison_rejet, vente.raison_rejet)
            })
    
    context = {
        'boutique': boutique,
        'ventes_refusees': ventes_refusees,
        'ventes_details': ventes_details,
        'total_refusees_jour': total_refusees_jour,
        'recette_jour': recette_jour,
        'total_combine': total_combine,
        'nb_refusees_jour': ventes_refusees_jour.count(),
        'nb_refusees_total': ventes_refusees.count(),
        'stats_par_raison': stats_par_raison,
        'date_filter': date_filter,
        'aujourd_hui': aujourd_hui.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'inventory/commercant/ventes_refusees_boutique.html', context)

@login_required
@commercant_required
def api_vente_details(request, vente_id):
    """API pour récupérer les détails d'une vente avec ses articles"""
    try:
        vente = get_object_or_404(Vente, id=vente_id)
        
        # Vérifier que le commerçant a accès à cette vente
        commercant = request.user.profil_commercant
        if vente.boutique.commercant != commercant:
            return JsonResponse({'error': 'Accès non autorisé'}, status=403)
        
        # Récupérer les lignes de vente avec les articles
        lignes = vente.lignes.select_related('article').all()
        
        # Construire les lignes avec gestion des articles supprimés et devise
        lignes_data = []
        for ligne in lignes:
            try:
                # Déterminer la devise et le prix à afficher
                devise_ligne = getattr(ligne, 'devise', vente.devise) or 'CDF'
                if devise_ligne == 'USD':
                    prix_affiche = ligne.prix_unitaire_usd or ligne.prix_unitaire
                    total_affiche = (ligne.prix_unitaire_usd or 0) * ligne.quantite
                    symbole = '$'
                else:
                    prix_affiche = ligne.prix_unitaire
                    total_affiche = ligne.total_ligne
                    symbole = 'FC'
                
                lignes_data.append({
                    'article_nom': ligne.article.nom if ligne.article else 'Article supprimé',
                    'article_code': getattr(ligne.article, 'code', '') or '-' if ligne.article else '-',
                    'quantite': ligne.quantite,
                    'prix_unitaire': f"{prix_affiche:,.2f}" if devise_ligne == 'USD' else f"{prix_affiche:,.0f}",
                    'total_ligne': f"{total_affiche:,.2f}" if devise_ligne == 'USD' else f"{total_affiche:,.0f}",
                    'devise': devise_ligne,
                    'symbole': symbole
                })
            except Exception:
                lignes_data.append({
                    'article_nom': 'Article inconnu',
                    'article_code': '-',
                    'quantite': ligne.quantite,
                    'prix_unitaire': f"{ligne.prix_unitaire:,.0f}",
                    'total_ligne': f"{ligne.total_ligne:,.0f}",
                    'devise': 'CDF',
                    'symbole': 'FC'
                })
        
        # Montant total selon devise
        if vente.devise == 'USD':
            montant_affiche = f"{vente.montant_total_usd or 0:,.2f}"
            symbole_total = '$'
        else:
            montant_affiche = f"{vente.montant_total:,.0f}"
            symbole_total = 'FC'
        
        # Construire la réponse JSON
        data = {
            'id': vente.id,
            'numero_facture': vente.numero_facture,
            'date_vente': vente.date_vente.strftime('%d/%m/%Y %H:%M'),
            'mode_paiement': vente.get_mode_paiement_display(),
            'terminal': vente.client_maui.nom_terminal if vente.client_maui else None,
            'devise': vente.devise,
            'symbole': symbole_total,
            'montant_total': montant_affiche,
            'lignes': lignes_data
        }
        
        return JsonResponse(data)
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# ===== GESTION DU DÉPÔT CENTRAL =====

@login_required
@commercant_required
def liste_depots(request):
    """Liste des dépôts du commerçant"""
    commercant = request.user.profil_commercant
    depots = commercant.boutiques.filter(est_depot=True)
    
    # Statistiques pour chaque dépôt
    for depot in depots:
        articles_depot = depot.articles.filter(est_actif=True)
        depot.nb_articles = articles_depot.count()
        # Valeur stock en CDF
        depot.valeur_stock_cdf = articles_depot.filter(devise='CDF').aggregate(
            total=Sum(F('quantite_stock') * F('prix_achat'))
        )['total'] or 0
        # Valeur stock en USD
        depot.valeur_stock_usd = articles_depot.filter(devise='USD').aggregate(
            total=Sum(F('quantite_stock') * F('prix_achat'))
        )['total'] or 0
        depot.valeur_stock = depot.valeur_stock_cdf
        depot.nb_transferts_mois = TransfertStock.objects.filter(
            depot_source=depot,
            date_transfert__gte=timezone.now().replace(day=1)
        ).count()
    
    context = {
        'commercant': commercant,
        'depots': depots,
    }
    
    return render(request, 'inventory/commercant/liste_depots.html', context)

@login_required
@commercant_required
def detail_depot(request, depot_id):
    """Détail d'un dépôt avec ses articles et statistiques"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Articles du dépôt
    articles = depot.articles.filter(est_actif=True).select_related('categorie').order_by('nom')
    
    # Statistiques
    total_articles = articles.count()
    
    # Valeur du stock en CDF (articles en CDF + conversion des articles en USD)
    articles_cdf = articles.filter(devise='CDF')
    articles_usd = articles.filter(devise='USD')
    
    valeur_stock_cdf = articles_cdf.aggregate(
        total=Sum(F('quantite_stock') * F('prix_achat'))
    )['total'] or 0
    
    # Valeur du stock en USD (articles en USD)
    valeur_stock_usd = articles_usd.aggregate(
        total=Sum(F('quantite_stock') * F('prix_achat'))
    )['total'] or 0
    
    # Valeur totale (pour compatibilité - en CDF)
    valeur_stock = valeur_stock_cdf
    
    articles_stock_bas = articles.filter(quantite_stock__lte=depot.alerte_stock_bas).count()
    
    # Transferts récents
    transferts_recents = TransfertStock.objects.filter(
        depot_source=depot
    ).select_related('article', 'boutique_destination').order_by('-date_transfert')[:20]
    
    # Boutiques de destination disponibles (non-dépôts)
    boutiques_destination = commercant.boutiques.filter(est_depot=False, est_active=True)
    
    context = {
        'depot': depot,
        'articles': articles,
        'total_articles': total_articles,
        'valeur_stock': valeur_stock,
        'valeur_stock_cdf': valeur_stock_cdf,
        'valeur_stock_usd': valeur_stock_usd,
        'articles_stock_bas': articles_stock_bas,
        'transferts_recents': transferts_recents,
        'boutiques_destination': boutiques_destination,
    }
    
    return render(request, 'inventory/commercant/detail_depot.html', context)

@login_required
@commercant_required
def creer_transfert_stock(request, depot_id):
    """Créer un transfert de stock du dépôt vers une boutique"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    if request.method == 'POST':
        try:
            article_id = request.POST.get('article_id')
            boutique_dest_id = request.POST.get('boutique_destination')
            quantite = int(request.POST.get('quantite', 0))
            commentaire = request.POST.get('commentaire', '')
            
            article = get_object_or_404(Article, id=article_id, boutique=depot)
            boutique_dest = get_object_or_404(Boutique, id=boutique_dest_id, commercant=commercant, est_depot=False)
            
            if quantite <= 0:
                messages.error(request, "La quantité doit être supérieure à 0")
                return redirect('inventory:detail_depot', depot_id=depot.id)
            
            if article.quantite_stock < quantite:
                messages.error(request, f"Stock insuffisant. Disponible: {article.quantite_stock}")
                return redirect('inventory:detail_depot', depot_id=depot.id)
            
            # Créer le transfert
            transfert = TransfertStock.objects.create(
                article=article,
                depot_source=depot,
                boutique_destination=boutique_dest,
                quantite=quantite,
                effectue_par=request.user.username,
                commentaire=commentaire,
                statut='EN_ATTENTE'
            )
            
            messages.success(request, f"Transfert créé: {quantite} x {article.nom} vers {boutique_dest.nom}")
            return redirect('inventory:detail_transfert', transfert_id=transfert.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création du transfert: {str(e)}")
            return redirect('inventory:detail_depot', depot_id=depot.id)
    
    return redirect('inventory:detail_depot', depot_id=depot.id)

@login_required
@commercant_required
def detail_transfert(request, transfert_id):
    """Détail d'un transfert de stock"""
    commercant = request.user.profil_commercant
    transfert = get_object_or_404(TransfertStock, id=transfert_id)
    
    # Vérifier l'accès
    if transfert.depot_source.commercant != commercant:
        messages.error(request, "Accès non autorisé")
        return redirect('inventory:commercant_dashboard')
    
    context = {
        'transfert': transfert,
    }
    
    return render(request, 'inventory/commercant/detail_transfert.html', context)

@login_required
@commercant_required
def valider_transfert(request, transfert_id):
    """Valider un transfert de stock"""
    commercant = request.user.profil_commercant
    transfert = get_object_or_404(TransfertStock, id=transfert_id)
    
    # Vérifier l'accès
    if transfert.depot_source.commercant != commercant:
        messages.error(request, "Accès non autorisé")
        return redirect('inventory:commercant_dashboard')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                transfert.valider_transfert(request.user.username)
            
            messages.success(request, f"Transfert validé avec succès! {transfert.quantite} x {transfert.article.nom} transféré vers {transfert.boutique_destination.nom}")
            return redirect('inventory:detail_depot', depot_id=transfert.depot_source.id)
            
        except ValidationError as e:
            messages.error(request, str(e))
        except Exception as e:
            messages.error(request, f"Erreur lors de la validation: {str(e)}")
    
    return redirect('inventory:detail_transfert', transfert_id=transfert.id)

@login_required
@commercant_required
def annuler_transfert(request, transfert_id):
    """Annuler un transfert de stock en attente"""
    commercant = request.user.profil_commercant
    transfert = get_object_or_404(TransfertStock, id=transfert_id)
    
    # Vérifier l'accès
    if transfert.depot_source.commercant != commercant:
        messages.error(request, "Accès non autorisé")
        return redirect('inventory:commercant_dashboard')
    
    if request.method == 'POST':
        if transfert.statut == 'EN_ATTENTE':
            transfert.statut = 'ANNULE'
            transfert.save()
            messages.success(request, "Transfert annulé")
        else:
            messages.error(request, "Ce transfert ne peut plus être annulé")
    
    return redirect('inventory:detail_depot', depot_id=transfert.depot_source.id)

@login_required
@commercant_required
def approvisionner_depot(request, depot_id):
    """Approvisionner le dépôt avec de nouveaux articles ou ajouter du stock"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    if request.method == 'POST':
        type_appro = request.POST.get('type_appro', 'nouveau')
        
        try:
            if type_appro == 'nouveau':
                # Créer un nouvel article
                code = request.POST.get('code', '').strip()
                nom = request.POST.get('nom', '').strip()
                devise = request.POST.get('devise', 'CDF')
                prix_achat = Decimal(request.POST.get('prix_achat', '0'))
                prix_vente = Decimal(request.POST.get('prix_vente', '0'))
                quantite = int(request.POST.get('quantite_initiale', '1'))
                description = request.POST.get('description', '')
                
                # Prix en USD (optionnels)
                prix_achat_usd_str = request.POST.get('prix_achat_usd', '').strip()
                prix_vente_usd_str = request.POST.get('prix_vente_usd', '').strip()
                prix_achat_usd = Decimal(prix_achat_usd_str) if prix_achat_usd_str else None
                prix_vente_usd = Decimal(prix_vente_usd_str) if prix_vente_usd_str else None
                
                if not nom:
                    messages.error(request, "Le nom de l'article est obligatoire")
                    return redirect('inventory:detail_depot', depot_id=depot.id)
                
                # Valider qu'au moins un prix de vente est fourni (CDF ou USD)
                if prix_vente == 0 and not prix_vente_usd:
                    messages.error(request, "Au moins un prix de vente est requis (CDF ou USD)")
                    return redirect('inventory:detail_depot', depot_id=depot.id)
                
                # Générer un code unique si non renseigné
                if not code:
                    import uuid
                    code = f"ART-{uuid.uuid4().hex[:8].upper()}"
                
                # Vérifier si l'article existe déjà dans ce dépôt
                if Article.objects.filter(code=code, boutique=depot).exists():
                    messages.error(request, f"Un article avec le code '{code}' existe déjà dans ce dépôt")
                    return redirect('inventory:detail_depot', depot_id=depot.id)
                
                # Créer l'article
                article = Article.objects.create(
                    code=code,
                    nom=nom,
                    description=description,
                    devise=devise,
                    prix_achat=prix_achat,
                    prix_vente=prix_vente,
                    prix_achat_usd=prix_achat_usd,
                    prix_vente_usd=prix_vente_usd,
                    quantite_stock=quantite,
                    boutique=depot,
                    est_actif=True
                )
                
                # Enregistrer le mouvement de stock
                MouvementStock.objects.create(
                    article=article,
                    type_mouvement='ENTREE',
                    quantite=quantite,
                    stock_avant=0,
                    stock_apres=quantite,
                    commentaire=f"Approvisionnement initial du dépôt",
                    reference_document=f"APPRO-{depot.id}-{article.id}",
                    utilisateur=request.user.username
                )
                
                messages.success(request, f"Article '{nom}' créé avec {quantite} unités en stock")
                
            else:
                # Ajouter du stock à un article existant
                article_id = request.POST.get('article_existant')
                quantite_ajout = int(request.POST.get('quantite_ajout', '1'))
                
                if not article_id:
                    messages.error(request, "Veuillez sélectionner un article")
                    return redirect('inventory:detail_depot', depot_id=depot.id)
                
                article = get_object_or_404(Article, id=article_id, boutique=depot)
                stock_avant = article.quantite_stock
                article.quantite_stock += quantite_ajout
                article.save()
                
                # Enregistrer le mouvement de stock
                MouvementStock.objects.create(
                    article=article,
                    type_mouvement='ENTREE',
                    quantite=quantite_ajout,
                    stock_avant=stock_avant,
                    stock_apres=article.quantite_stock,
                    commentaire=f"Approvisionnement du dépôt",
                    reference_document=f"APPRO-{depot.id}-{article.id}",
                    utilisateur=request.user.username
                )
                
                messages.success(request, f"{quantite_ajout} unités ajoutées à '{article.nom}' (nouveau stock: {article.quantite_stock})")
                
        except Exception as e:
            messages.error(request, f"Erreur lors de l'approvisionnement: {str(e)}")
    
    return redirect('inventory:detail_depot', depot_id=depot.id)

@login_required
@commercant_required
def detail_article_depot(request, depot_id, article_id):
    """Voir les détails d'un article du dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    article = get_object_or_404(Article, id=article_id, boutique=depot)
    
    # Historique des mouvements de stock
    mouvements = MouvementStock.objects.filter(article=article).order_by('-date_mouvement')[:50]
    
    context = {
        'depot': depot,
        'article': article,
        'mouvements': mouvements,
    }
    
    return render(request, 'inventory/commercant/detail_article_depot.html', context)

@login_required
@commercant_required
def modifier_article_depot(request, depot_id, article_id):
    """Modifier un article du dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    article = get_object_or_404(Article, id=article_id, boutique=depot)
    
    if request.method == 'POST':
        try:
            stock_avant = article.quantite_stock
            
            article.nom = request.POST.get('nom', article.nom).strip()
            article.code = request.POST.get('code', article.code).strip()
            article.description = request.POST.get('description', '')
            article.devise = request.POST.get('devise', 'CDF')
            article.prix_achat = Decimal(request.POST.get('prix_achat', '0'))
            article.prix_vente = Decimal(request.POST.get('prix_vente', '0'))
            article.quantite_stock = int(request.POST.get('quantite_stock', article.quantite_stock))
            
            # Catégorie
            categorie_id = request.POST.get('categorie', '').strip()
            if categorie_id:
                article.categorie = Categorie.objects.filter(id=categorie_id, boutique=depot).first()
            else:
                article.categorie = None
            
            # Date d'expiration
            date_exp_str = request.POST.get('date_expiration', '').strip()
            if date_exp_str:
                from datetime import datetime
                article.date_expiration = datetime.strptime(date_exp_str, '%Y-%m-%d').date()
            else:
                article.date_expiration = None
            
            # Statut actif
            article.est_actif = request.POST.get('est_actif') == 'on'
            
            # Prix USD optionnels
            prix_achat_usd_str = request.POST.get('prix_achat_usd', '').strip()
            prix_vente_usd_str = request.POST.get('prix_vente_usd', '').strip()
            article.prix_achat_usd = Decimal(prix_achat_usd_str) if prix_achat_usd_str else None
            article.prix_vente_usd = Decimal(prix_vente_usd_str) if prix_vente_usd_str else None
            
            article.save()
            
            # Enregistrer le mouvement de stock si la quantité a changé
            if article.quantite_stock != stock_avant:
                diff = article.quantite_stock - stock_avant
                MouvementStock.objects.create(
                    article=article,
                    type_mouvement='ENTREE' if diff > 0 else 'SORTIE',
                    quantite=diff,
                    stock_avant=stock_avant,
                    stock_apres=article.quantite_stock,
                    commentaire=f"Modification manuelle du stock",
                    reference_document=f"MODIF-{depot.id}-{article.id}",
                    utilisateur=request.user.username
                )
            
            messages.success(request, f"Article '{article.nom}' modifié avec succès")
            return redirect('inventory:detail_depot', depot_id=depot.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
    
    # Récupérer les catégories pour le select (du dépôt + globales + catégorie actuelle de l'article)
    from django.db.models import Q
    categories_query = Q(boutique=depot) | Q(boutique__isnull=True)
    if article.categorie_id:
        categories_query |= Q(id=article.categorie_id)
    categories = Categorie.objects.filter(categories_query).distinct().order_by('nom')
    
    # Récupérer le taux de change (priorité: dépôt > commerçant > défaut)
    taux_dollar = depot.taux_dollar or commercant.taux_dollar or 2800
    
    context = {
        'depot': depot,
        'article': article,
        'categories': categories,
        'taux_dollar': taux_dollar,
    }
    
    return render(request, 'inventory/commercant/modifier_article_depot.html', context)

@login_required
@commercant_required
def supprimer_article_depot(request, depot_id, article_id):
    """Supprimer un article du dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    article = get_object_or_404(Article, id=article_id, boutique=depot)
    
    nom_article = article.nom
    article.est_actif = False
    article.save()
    
    messages.success(request, f"Article '{nom_article}' supprimé du dépôt")
    return redirect('inventory:detail_depot', depot_id=depot.id)

@login_required
@commercant_required
def importer_articles_vers_depot(request, depot_id):
    """Importer des articles existants depuis les points de vente vers le dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Récupérer tous les articles des boutiques du commerçant (sauf ceux du dépôt)
    articles_boutiques = Article.objects.filter(
        boutique__commercant=commercant,
        boutique__est_depot=False,
        est_actif=True
    ).select_related('boutique', 'categorie').order_by('nom')
    
    # Articles déjà présents dans le dépôt (basé sur le code)
    codes_depot = set(Article.objects.filter(
        boutique=depot,
        est_actif=True
    ).values_list('code', flat=True))
    
    # Grouper par nom d'article pour éviter les doublons visuels
    articles_uniques = {}
    for article in articles_boutiques:
        if article.code not in codes_depot:
            if article.nom not in articles_uniques:
                articles_uniques[article.nom] = {
                    'article': article,
                    'boutiques': [article.boutique.nom],
                    'deja_dans_depot': False
                }
            else:
                articles_uniques[article.nom]['boutiques'].append(article.boutique.nom)
    
    if request.method == 'POST':
        articles_selectionnes = request.POST.getlist('articles_selectionnes')
        
        if not articles_selectionnes:
            messages.error(request, "Veuillez sélectionner au moins un article à importer")
            return redirect('inventory:importer_articles_vers_depot', depot_id=depot.id)
        
        articles_importes = 0
        erreurs = []
        
        try:
            with transaction.atomic():
                for article_id in articles_selectionnes:
                    try:
                        article_source = Article.objects.get(id=article_id)
                        quantite_key = f'quantite_{article_id}'
                        quantite_initiale = int(request.POST.get(quantite_key, 0))
                        
                        if quantite_initiale < 0:
                            erreurs.append(f"{article_source.nom}: la quantité ne peut pas être négative")
                            continue
                        
                        # Vérifier si l'article existe déjà dans le dépôt
                        if Article.objects.filter(code=article_source.code, boutique=depot).exists():
                            erreurs.append(f"{article_source.nom} (code: {article_source.code}): existe déjà dans le dépôt")
                            continue
                        
                        # Créer l'article dans le dépôt
                        article_depot = Article.objects.create(
                            code=article_source.code,
                            nom=article_source.nom,
                            description=article_source.description,
                            devise=article_source.devise,
                            prix_achat=article_source.prix_achat,
                            prix_vente=article_source.prix_vente,
                            prix_achat_usd=article_source.prix_achat_usd,
                            prix_vente_usd=article_source.prix_vente_usd,
                            categorie=article_source.categorie,
                            boutique=depot,
                            quantite_stock=quantite_initiale,
                            est_actif=True
                        )
                        
                        # Créer un mouvement de stock si quantité > 0
                        if quantite_initiale > 0:
                            MouvementStock.objects.create(
                                article=article_depot,
                                type_mouvement='ENTREE',
                                quantite=quantite_initiale,
                                stock_avant=0,
                                stock_apres=quantite_initiale,
                                commentaire=f"Import depuis point de vente: {article_source.boutique.nom}",
                                reference_document=f"IMPORT-{depot.id}-{article_depot.id}",
                                utilisateur=request.user.username
                            )
                        
                        articles_importes += 1
                        
                    except Article.DoesNotExist:
                        erreurs.append(f"Article ID {article_id} introuvable")
                    except Exception as e:
                        erreurs.append(f"Erreur lors de l'import de l'article {article_id}: {str(e)}")
                
                if articles_importes == 0 and erreurs:
                    raise Exception("Aucun article importé")
        
        except Exception as e:
            for erreur in erreurs:
                messages.error(request, erreur)
            return redirect('inventory:importer_articles_vers_depot', depot_id=depot.id)
        
        # Messages de résultat
        if articles_importes > 0:
            messages.success(request, f"{articles_importes} article(s) importé(s) avec succès dans le dépôt")
        
        for erreur in erreurs:
            messages.warning(request, erreur)
        
        return redirect('inventory:detail_depot', depot_id=depot.id)
    
    context = {
        'depot': depot,
        'articles_uniques': articles_uniques,
        'commercant': commercant,
    }
    
    return render(request, 'inventory/commercant/importer_articles_depot.html', context)

@login_required
@commercant_required
def importer_excel_depot(request, depot_id):
    """Importer des articles depuis un fichier Excel vers le dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    if request.method == 'POST':
        fichier_excel = request.FILES.get('fichier_excel')
        
        if not fichier_excel:
            messages.error(request, "Veuillez sélectionner un fichier Excel")
            return redirect('inventory:importer_excel_depot', depot_id=depot.id)
        
        # Vérifier l'extension du fichier
        if not fichier_excel.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Le fichier doit être au format Excel (.xlsx ou .xls)")
            return redirect('inventory:importer_excel_depot', depot_id=depot.id)
        
        try:
            import openpyxl
            from io import BytesIO
            
            # Lire le fichier Excel
            wb = openpyxl.load_workbook(BytesIO(fichier_excel.read()))
            ws = wb.active
            
            # Récupérer les en-têtes (première ligne)
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
            
            # Mapper les colonnes attendues
            colonnes = {
                'code': None,
                'nom': None,
                'prix_achat': None,
                'prix_vente': None,
                'stock': None,
                'devise': None,
            }
            
            # Variantes acceptées pour chaque colonne
            variantes = {
                'code': ['code', 'code article', 'code_article', 'ref', 'reference', 'référence'],
                'nom': ['nom', 'nom article', 'nom_article', 'designation', 'désignation', 'libelle', 'libellé', 'article'],
                'prix_achat': ['prix achat', 'prix_achat', 'prix d\'achat', 'achat', 'pa', 'cout', 'coût'],
                'prix_vente': ['prix vente', 'prix_vente', 'prix de vente', 'vente', 'pv', 'prix'],
                'stock': ['stock', 'quantite', 'quantité', 'qte', 'qty', 'quantite_stock'],
                'devise': ['devise', 'monnaie', 'currency'],
            }
            
            # Trouver les indices des colonnes
            for i, header in enumerate(headers):
                for col_name, variants in variantes.items():
                    if header in variants:
                        colonnes[col_name] = i
                        break
            
            # Vérifier les colonnes obligatoires
            if colonnes['nom'] is None:
                messages.error(request, "Colonne 'Nom' non trouvée dans le fichier Excel. Colonnes attendues: code, nom, prix_achat, prix_vente, stock")
                return redirect('inventory:importer_excel_depot', depot_id=depot.id)
            
            articles_importes = 0
            articles_mis_a_jour = 0
            erreurs = []
            
            # Parcourir les lignes (à partir de la 2ème)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraire les valeurs
                    code = str(row[colonnes['code']]).strip() if colonnes['code'] is not None and row[colonnes['code']] else None
                    nom = str(row[colonnes['nom']]).strip() if colonnes['nom'] is not None and row[colonnes['nom']] else None
                    
                    if not nom:
                        continue  # Ignorer les lignes sans nom
                    
                    # Prix d'achat
                    prix_achat = 0
                    if colonnes['prix_achat'] is not None and row[colonnes['prix_achat']]:
                        try:
                            prix_achat = Decimal(str(row[colonnes['prix_achat']]).replace(',', '.').replace(' ', ''))
                        except:
                            prix_achat = 0
                    
                    # Prix de vente
                    prix_vente = 0
                    if colonnes['prix_vente'] is not None and row[colonnes['prix_vente']]:
                        try:
                            prix_vente = Decimal(str(row[colonnes['prix_vente']]).replace(',', '.').replace(' ', ''))
                        except:
                            prix_vente = 0
                    
                    # Stock
                    stock = 0
                    if colonnes['stock'] is not None and row[colonnes['stock']]:
                        try:
                            stock = int(float(str(row[colonnes['stock']]).replace(',', '.').replace(' ', '')))
                        except:
                            stock = 0
                    
                    # Devise (par défaut CDF)
                    devise = 'CDF'
                    if colonnes['devise'] is not None and row[colonnes['devise']]:
                        devise_val = str(row[colonnes['devise']]).upper().strip()
                        if devise_val in ['USD', '$', 'DOLLAR', 'DOLLARS']:
                            devise = 'USD'
                    
                    # Générer un code si non fourni
                    if not code:
                        import random
                        import string
                        code = 'ART-' + ''.join(random.choices(string.ascii_uppercase + string.digits, k=8))
                    
                    # Vérifier si l'article existe déjà dans le dépôt (par code)
                    article_existant = Article.objects.filter(code=code, boutique=depot).first()
                    
                    if article_existant:
                        # Mettre à jour le stock existant
                        stock_avant = article_existant.quantite_stock
                        article_existant.quantite_stock += stock
                        article_existant.prix_achat = prix_achat if prix_achat > 0 else article_existant.prix_achat
                        article_existant.prix_vente = prix_vente if prix_vente > 0 else article_existant.prix_vente
                        article_existant.save()
                        
                        if stock > 0:
                            MouvementStock.objects.create(
                                article=article_existant,
                                type_mouvement='ENTREE',
                                quantite=stock,
                                stock_avant=stock_avant,
                                stock_apres=article_existant.quantite_stock,
                                commentaire="Import Excel - Mise à jour stock",
                                reference_document=f"IMPORT-EXCEL-{depot.id}",
                                utilisateur=request.user.username
                            )
                        
                        articles_mis_a_jour += 1
                    else:
                        # Créer un nouvel article
                        article = Article.objects.create(
                            code=code,
                            nom=nom,
                            devise=devise,
                            prix_achat=prix_achat,
                            prix_vente=prix_vente,
                            boutique=depot,
                            quantite_stock=stock,
                            est_actif=True
                        )
                        
                        if stock > 0:
                            MouvementStock.objects.create(
                                article=article,
                                type_mouvement='ENTREE',
                                quantite=stock,
                                stock_avant=0,
                                stock_apres=stock,
                                commentaire="Import Excel - Nouvel article",
                                reference_document=f"IMPORT-EXCEL-{depot.id}",
                                utilisateur=request.user.username
                            )
                        
                        articles_importes += 1
                        
                except Exception as e:
                    erreurs.append(f"Ligne {row_num}: {str(e)}")
            
            # Messages de résultat
            if articles_importes > 0:
                messages.success(request, f"{articles_importes} nouvel(aux) article(s) importé(s)")
            if articles_mis_a_jour > 0:
                messages.success(request, f"{articles_mis_a_jour} article(s) mis à jour (stock ajouté)")
            if not articles_importes and not articles_mis_a_jour:
                messages.warning(request, "Aucun article n'a été importé. Vérifiez le format du fichier.")
            
            for erreur in erreurs[:10]:  # Limiter à 10 erreurs
                messages.warning(request, erreur)
            
            if len(erreurs) > 10:
                messages.warning(request, f"... et {len(erreurs) - 10} autres erreurs")
            
            return redirect('inventory:detail_depot', depot_id=depot.id)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la lecture du fichier Excel: {str(e)}")
            return redirect('inventory:importer_excel_depot', depot_id=depot.id)
    
    context = {
        'depot': depot,
        'commercant': commercant,
    }
    
    return render(request, 'inventory/commercant/importer_excel_depot.html', context)

@login_required
@commercant_required
def historique_transferts(request, depot_id):
    """Historique des transferts d'un dépôt"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Filtres
    statut = request.GET.get('statut', '')
    boutique_dest = request.GET.get('boutique', '')
    
    transferts = TransfertStock.objects.filter(depot_source=depot).select_related(
        'article', 'boutique_destination'
    ).order_by('-date_transfert')
    
    if statut:
        transferts = transferts.filter(statut=statut)
    if boutique_dest:
        transferts = transferts.filter(boutique_destination_id=boutique_dest)
    
    # Statistiques
    stats = {
        'total': transferts.count(),
        'en_attente': transferts.filter(statut='EN_ATTENTE').count(),
        'valides': transferts.filter(statut='VALIDE').count(),
        'annules': transferts.filter(statut='ANNULE').count(),
    }
    
    boutiques = commercant.boutiques.filter(est_depot=False, est_active=True)
    
    context = {
        'depot': depot,
        'transferts': transferts[:100],
        'stats': stats,
        'boutiques': boutiques,
        'statut_filter': statut,
        'boutique_filter': boutique_dest,
    }
    
    return render(request, 'inventory/commercant/historique_transferts.html', context)


# ===== TRANSFERT MULTIPLE =====

@login_required
@commercant_required
def transfert_multiple(request, depot_id):
    """Page de transfert multiple d'articles du dépôt vers une boutique"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Articles disponibles dans le dépôt (avec stock > 0)
    articles = depot.articles.filter(
        est_actif=True,
        quantite_stock__gt=0
    ).select_related('categorie').order_by('nom')
    
    # Boutiques de destination
    boutiques_destination = commercant.boutiques.filter(est_depot=False, est_active=True)
    
    if request.method == 'POST':
        boutique_dest_id = request.POST.get('boutique_destination')
        commentaire_global = request.POST.get('commentaire', '')
        
        if not boutique_dest_id:
            messages.error(request, "Veuillez sélectionner une boutique de destination")
            return redirect('inventory:transfert_multiple', depot_id=depot.id)
        
        boutique_dest = get_object_or_404(Boutique, id=boutique_dest_id, commercant=commercant, est_depot=False)
        
        # Récupérer les articles sélectionnés avec leurs quantités
        articles_selectionnes = request.POST.getlist('articles_selectionnes')
        
        if not articles_selectionnes:
            messages.error(request, "Veuillez sélectionner au moins un article à transférer")
            return redirect('inventory:transfert_multiple', depot_id=depot.id)
        
        transferts_crees = []
        erreurs = []
        
        # Générer une référence de lot unique
        import uuid as _uuid
        reference_lot = f"LOT-{timezone.now().strftime('%Y%m%d%H%M%S')}-{_uuid.uuid4().hex[:6].upper()}"
        
        try:
            with transaction.atomic():
                for article_id in articles_selectionnes:
                    quantite_key = f'quantite_{article_id}'
                    quantite = int(request.POST.get(quantite_key, 0))
                    
                    if quantite <= 0:
                        continue
                    
                    try:
                        article = Article.objects.get(id=article_id, boutique=depot)
                        
                        if article.quantite_stock < quantite:
                            erreurs.append(f"{article.nom}: stock insuffisant (dispo: {article.quantite_stock}, demandé: {quantite})")
                            continue
                        
                        # Créer le transfert et le valider immédiatement
                        transfert = TransfertStock.objects.create(
                            article=article,
                            depot_source=depot,
                            boutique_destination=boutique_dest,
                            quantite=quantite,
                            effectue_par=request.user.username,
                            commentaire=commentaire_global,
                            reference_lot=reference_lot,
                            statut='EN_ATTENTE'
                        )
                        # Validation directe - mise à jour des stocks
                        transfert.valider_transfert(request.user.username)
                        transferts_crees.append(transfert)
                        
                    except Article.DoesNotExist:
                        erreurs.append(f"Article ID {article_id} introuvable")
                
                if not transferts_crees and erreurs:
                    raise Exception("Aucun transfert créé")
        
        except Exception as e:
            for erreur in erreurs:
                messages.error(request, erreur)
            return redirect('inventory:transfert_multiple', depot_id=depot.id)
        
        for erreur in erreurs:
            messages.warning(request, erreur)
        
        # Rediriger vers le bon de transfert imprimable
        if transferts_crees:
            return redirect('inventory:bon_transfert', depot_id=depot.id, reference_lot=reference_lot)
        
        return redirect('inventory:detail_depot', depot_id=depot.id)
    
    context = {
        'depot': depot,
        'articles': articles,
        'boutiques_destination': boutiques_destination,
    }
    
    return render(request, 'inventory/commercant/transfert_multiple.html', context)


@login_required
@commercant_required
def bon_transfert(request, depot_id, reference_lot):
    """Bon de transfert imprimable - récapitulatif après validation"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    transferts = TransfertStock.objects.filter(
        depot_source=depot,
        reference_lot=reference_lot
    ).select_related('article', 'article__categorie', 'boutique_destination').order_by('article__nom')
    
    if not transferts.exists():
        messages.error(request, "Bon de transfert introuvable.")
        return redirect('inventory:detail_depot', depot_id=depot.id)
    
    # Calculs récapitulatifs — toujours en Francs Congolais (FC)
    taux = depot.taux_dollar or Decimal('1')
    transferts_list = list(transferts)
    premier = transferts_list[0] if transferts_list else None
    total_articles = len(transferts_list)
    total_quantite = 0
    total_cout_vente = Decimal('0')
    for t in transferts_list:
        pv = t.article.prix_vente or Decimal('0')
        # Si l'article est en USD, convertir en FC
        if getattr(t.article, 'devise', 'CDF') == 'USD':
            pv = pv * taux
        t.prix_vente_fc = pv
        t.cout_vente_ligne = t.quantite * pv
        total_quantite += t.quantite
        total_cout_vente += t.cout_vente_ligne
    
    context = {
        'depot': depot,
        'transferts': transferts_list,
        'reference_lot': reference_lot,
        'boutique_destination': premier.boutique_destination,
        'date_transfert': premier.date_transfert,
        'effectue_par': premier.effectue_par,
        'valide_par': premier.valide_par,
        'total_articles': total_articles,
        'total_quantite': total_quantite,
        'total_cout_vente': total_cout_vente,
        'commercant': commercant,
    }
    
    return render(request, 'inventory/commercant/bon_transfert.html', context)


@login_required
@commercant_required
def valider_transferts_multiples(request, depot_id):
    """Valider plusieurs transferts en attente en une seule fois"""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    if request.method == 'POST':
        transferts_ids = request.POST.getlist('transferts_selectionnes')
        
        if not transferts_ids:
            messages.error(request, "Veuillez sélectionner au moins un transfert à valider")
            return redirect('inventory:historique_transferts', depot_id=depot.id)
        
        valides = 0
        erreurs = []
        
        for transfert_id in transferts_ids:
            try:
                transfert = TransfertStock.objects.get(id=transfert_id, depot_source=depot, statut='EN_ATTENTE')
                with transaction.atomic():
                    transfert.valider_transfert(request.user.username)
                valides += 1
            except TransfertStock.DoesNotExist:
                erreurs.append(f"Transfert {transfert_id} introuvable ou déjà traité")
            except Exception as e:
                erreurs.append(f"Erreur transfert {transfert_id}: {str(e)}")
        
        if valides:
            messages.success(request, f"{valides} transfert(s) validé(s) avec succès")
        for erreur in erreurs:
            messages.warning(request, erreur)
    
    return redirect('inventory:historique_transferts', depot_id=depot.id)


# ===== HISTORIQUE DES MOUVEMENTS DE STOCK =====

@login_required
@commercant_required
@boutique_access_required
def historique_mouvements_stock(request, boutique_id):
    """Historique des mouvements de stock d'une boutique"""
    boutique = request.boutique
    
    # Filtres
    type_mouvement = request.GET.get('type', '')
    article_id = request.GET.get('article', '')
    search_query = request.GET.get('search', '').strip()
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    
    # Récupérer les mouvements de stock pour les articles de cette boutique
    mouvements = MouvementStock.objects.filter(
        article__boutique=boutique
    ).select_related('article', 'article__categorie').order_by('-date_mouvement')
    
    # Appliquer les filtres
    if type_mouvement:
        mouvements = mouvements.filter(type_mouvement=type_mouvement)
    
    if article_id:
        mouvements = mouvements.filter(article_id=article_id)
    
    # Recherche par nom ou code d'article
    if search_query:
        from django.db.models import Q
        mouvements = mouvements.filter(
            Q(article__nom__icontains=search_query) |
            Q(article__code__icontains=search_query) |
            Q(reference_document__icontains=search_query) |
            Q(commentaire__icontains=search_query)
        )
    
    if date_debut:
        try:
            date_debut_parsed = datetime.strptime(date_debut, '%Y-%m-%d')
            mouvements = mouvements.filter(date_mouvement__date__gte=date_debut_parsed)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_parsed = datetime.strptime(date_fin, '%Y-%m-%d')
            mouvements = mouvements.filter(date_mouvement__date__lte=date_fin_parsed)
        except ValueError:
            pass
    
    # Statistiques
    stats = {
        'total_mouvements': mouvements.count(),
        'entrees': mouvements.filter(type_mouvement='ENTREE').count(),
        'sorties': mouvements.filter(type_mouvement='SORTIE').count(),
        'ventes': mouvements.filter(type_mouvement='VENTE').count(),
        'ajustements': mouvements.filter(type_mouvement='AJUSTEMENT').count(),
        'retours': mouvements.filter(type_mouvement='RETOUR').count(),
    }
    
    # Liste des articles pour le filtre
    articles = boutique.articles.filter(est_actif=True).order_by('nom')
    
    # Types de mouvement pour le filtre
    types_mouvement = MouvementStock.TYPES
    
    # Réponse AJAX : retourner JSON uniquement les données du tableau
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        mouvements_data = []
        for m in mouvements[:200]:
            mouvements_data.append({
                'date': m.date_mouvement.strftime('%d/%m/%Y'),
                'heure': m.date_mouvement.strftime('%H:%M'),
                'type': m.type_mouvement,
                'type_display': m.get_type_mouvement_display(),
                'article_nom': m.article.nom,
                'article_code': m.article.code or '',
                'quantite': m.quantite,
                'stock_avant': m.stock_avant,
                'stock_apres': m.stock_apres,
                'reference': m.reference_document or '',
                'utilisateur': str(m.utilisateur) if m.utilisateur else '',
                'commentaire': (m.commentaire or '')[:60],
            })
        return JsonResponse({
            'mouvements': mouvements_data,
            'stats': stats,
        })

    context = {
        'boutique': boutique,
        'mouvements': mouvements[:200],  # Limiter à 200 pour performance
        'stats': stats,
        'articles': articles,
        'types_mouvement': types_mouvement,
        'type_filter': type_mouvement,
        'article_filter': int(article_id) if article_id else None,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'search_query': search_query,
    }
    
    return render(request, 'inventory/commercant/historique_mouvements_stock.html', context)


# ===== APPROVISIONNEMENT PAR FACTURE =====

@login_required
@commercant_required
def approvisionner_facture(request, depot_id):
    """Créer une facture d'approvisionnement avec plusieurs articles."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Récupérer les fournisseurs et catégories
    fournisseurs = Fournisseur.objects.filter(commercant=commercant, est_actif=True)
    categories = Categorie.objects.filter(boutique=depot)
    articles_existants = Article.objects.filter(boutique=depot, est_actif=True).order_by('nom')
    
    # Récupérer les dernières données d'approvisionnement par article (pour pré-remplissage)
    derniers_appros = {}
    last_lignes = LigneApprovisionnement.objects.filter(
        article__in=articles_existants
    ).select_related('article', 'facture').order_by('article_id', '-date_creation')
    seen_ids = set()
    for ligne in last_lignes:
        if ligne.article_id not in seen_ids:
            seen_ids.add(ligne.article_id)
            # Devise de la facture: CDF = FC, USD = USD
            devise_facture = ligne.facture.devise if ligne.facture else 'CDF'
            derniers_appros[ligne.article_id] = {
                'type_quantite': ligne.type_quantite,
                'nombre_cartons': ligne.nombre_cartons,
                'pieces_par_carton': ligne.pieces_par_carton,
                'prix_achat_carton': float(ligne.prix_achat_carton),
                'prix_achat_unitaire': float(ligne.prix_achat_unitaire),
                'devise_saisie': devise_facture,  # Devise utilisée lors du dernier appro
            }
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Données de la facture
                numero_facture = request.POST.get('numero_facture', '').strip()
                fournisseur_id = request.POST.get('fournisseur_id', '').strip()
                fournisseur_nom = request.POST.get('fournisseur_nom', '').strip()
                date_facture_str = request.POST.get('date_facture', '')
                devise = request.POST.get('devise', 'CDF')
                notes = request.POST.get('notes', '')
                
                if not numero_facture:
                    messages.error(request, "Le numéro de facture est obligatoire")
                    return redirect('inventory:approvisionner_facture', depot_id=depot.id)
                
                # Gérer le fournisseur
                fournisseur = None
                if fournisseur_id:
                    fournisseur = Fournisseur.objects.filter(id=fournisseur_id, commercant=commercant).first()
                elif fournisseur_nom:
                    # Créer un nouveau fournisseur
                    fournisseur, created = Fournisseur.objects.get_or_create(
                        nom=fournisseur_nom,
                        commercant=commercant,
                        defaults={'est_actif': True}
                    )
                
                # Date de facture
                try:
                    date_facture = datetime.strptime(date_facture_str, '%Y-%m-%d').date() if date_facture_str else timezone.now().date()
                except ValueError:
                    date_facture = timezone.now().date()
                
                # Vérifier unicité du numéro de facture
                if FactureApprovisionnement.objects.filter(numero_facture=numero_facture, depot=depot).exists():
                    msg = f"Le numéro de facture '{numero_facture}' existe déjà."
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': msg}, status=400)
                    messages.error(request, msg)
                    return redirect('inventory:approvisionner_facture', depot_id=depot.id)
                
                # Créer la facture
                facture = FactureApprovisionnement.objects.create(
                    numero_facture=numero_facture,
                    fournisseur=fournisseur,
                    fournisseur_nom=fournisseur_nom if not fournisseur else '',
                    depot=depot,
                    date_facture=date_facture,
                    devise=devise,
                    notes=notes,
                    created_by=request.user.username
                )
                
                # Traiter les lignes d'articles
                articles_json_str = request.POST.get('articles_json', '[]')
                articles_data = json.loads(articles_json_str) if articles_json_str else []
                
                if not articles_data:
                    messages.error(request, "Veuillez ajouter au moins un article à la facture")
                    return redirect('inventory:approvisionner_facture', depot_id=depot.id)
                
                for art_data in articles_data:
                    # Récupérer ou créer l'article
                    article_id = art_data.get('article_id')
                    if article_id:
                        article = Article.objects.get(id=article_id, boutique=depot)
                    else:
                        # Créer un nouvel article
                        code = art_data.get('code', '').strip()
                        if not code:
                            code = f"ART-{uuid.uuid4().hex[:8].upper()}"
                        
                        categorie_id = art_data.get('categorie_id')
                        categorie_nom = art_data.get('categorie_nom', '').strip()
                        if categorie_id:
                            categorie = Categorie.objects.filter(id=categorie_id).first()
                        elif categorie_nom:
                            categorie, _ = Categorie.objects.get_or_create(
                                nom__iexact=categorie_nom,
                                boutique=depot,
                                defaults={'nom': categorie_nom}
                            )
                        else:
                            categorie = None
                        
                        # Si facture en USD, l'article sera stocké en CDF (après conversion)
                        devise_article = 'CDF' if devise == 'USD' else devise
                        
                        article = Article.objects.create(
                            code=code,
                            nom=art_data.get('nom', 'Article sans nom'),
                            description=art_data.get('description', ''),
                            devise=devise_article,
                            prix_achat=Decimal(str(art_data.get('prix_achat_unitaire', 0))),
                            prix_vente=Decimal(str(art_data.get('prix_vente', 0))),
                            quantite_stock=0,
                            boutique=depot,
                            categorie=categorie,
                            est_actif=True
                        )
                    
                    # Données de la ligne
                    type_quantite = art_data.get('type_quantite', 'UNITE')
                    nombre_cartons = int(art_data.get('nombre_cartons', 0))
                    pieces_par_carton = int(art_data.get('pieces_par_carton', 1))
                    pieces_supplementaires = int(art_data.get('pieces_supplementaires', 0))
                    quantite_unites = int(art_data.get('quantite_unites', 0))
                    prix_achat_carton = Decimal(str(art_data.get('prix_achat_carton', 0)))
                    prix_achat_unitaire = Decimal(str(art_data.get('prix_achat_unitaire', 0)))
                    prix_piece_sup = Decimal(str(art_data.get('prix_piece_sup', 0)))
                    prix_vente = Decimal(str(art_data.get('prix_vente', 0)))
                    
                    # Calculs pour cartons (avec pièces supplémentaires)
                    if type_quantite == 'CARTON':
                        if prix_achat_carton > 0 and pieces_par_carton > 0:
                            prix_achat_unitaire = prix_achat_carton / pieces_par_carton
                        # Si prix pièce sup non renseigné, utiliser le prix unitaire
                        if prix_piece_sup == 0 and pieces_supplementaires > 0:
                            prix_piece_sup = prix_achat_unitaire
                        # Total unités = (cartons × pièces/carton) + pièces supplémentaires
                        quantite_unites = (nombre_cartons * pieces_par_carton) + pieces_supplementaires
                    
                    # NB: La conversion USD→CDF est déjà effectuée côté JavaScript (ajouterArticle)
                    # Les prix reçus ici sont toujours en Francs Congolais
                    
                    # Calcul du prix total (cartons + pièces supplémentaires)
                    if type_quantite == 'CARTON':
                        prix_achat_total = (nombre_cartons * prix_achat_carton) + (pieces_supplementaires * prix_piece_sup)
                    else:
                        prix_achat_total = quantite_unites * prix_achat_unitaire
                    
                    # Créer la ligne d'approvisionnement
                    LigneApprovisionnement.objects.create(
                        facture=facture,
                        article=article,
                        categorie=article.categorie,
                        type_quantite=type_quantite,
                        nombre_cartons=nombre_cartons,
                        pieces_par_carton=pieces_par_carton,
                        quantite_unites=quantite_unites,
                        prix_achat_carton=prix_achat_carton,
                        prix_achat_unitaire=prix_achat_unitaire,
                        prix_achat_total=prix_achat_total,
                        prix_vente_unitaire=prix_vente
                    )
                    
                    # Mettre à jour le stock de l'article
                    stock_avant = article.quantite_stock
                    article.quantite_stock += quantite_unites
                    article.prix_achat = prix_achat_unitaire
                    if prix_vente > 0:
                        article.prix_vente = prix_vente
                    article.save()
                    
                    # Enregistrer le mouvement de stock
                    MouvementStock.objects.create(
                        article=article,
                        type_mouvement='ENTREE',
                        quantite=quantite_unites,
                        stock_avant=stock_avant,
                        stock_apres=article.quantite_stock,
                        commentaire=f"Facture {numero_facture} - {fournisseur.nom if fournisseur else fournisseur_nom}",
                        reference_document=f"FACT-{facture.id}",
                        utilisateur=request.user.username
                    )
                
                # Recalculer le montant total
                facture.calculer_montant_total()
                
                # Si requête AJAX, retourner JSON sans rediriger
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f"Facture {numero_facture} enregistrée avec {len(articles_data)} article(s)",
                        'facture_id': facture.id,
                    })
                
                messages.success(request, f"Facture {numero_facture} enregistrée avec {len(articles_data)} article(s)")
                return redirect('inventory:detail_depot', depot_id=depot.id)
                
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
            messages.error(request, f"Erreur: {str(e)}")
    
    context = {
        'depot': depot,
        'fournisseurs': fournisseurs,
        'categories': categories,
        'articles_existants': articles_existants,
        'derniers_appros_json': json.dumps(derniers_appros),
        'today': timezone.now().date().isoformat(),
    }
    
    return render(request, 'inventory/commercant/approvisionner_facture.html', context)


@login_required
@commercant_required
@boutique_access_required
def approvisionner_facture_boutique(request, boutique_id):
    """Créer une facture d'approvisionnement pour un point de vente (local à la boutique)."""
    boutique = request.boutique
    commercant = request.user.profil_commercant
    
    # Récupérer les fournisseurs et catégories
    fournisseurs = Fournisseur.objects.filter(commercant=commercant, est_actif=True)
    categories = Categorie.objects.filter(boutique=boutique)
    articles_existants = Article.objects.filter(boutique=boutique, est_actif=True).order_by('nom')
    
    # Récupérer les dernières données d'approvisionnement par article (pour pré-remplissage)
    derniers_appros = {}
    last_lignes = LigneApprovisionnement.objects.filter(
        article__in=articles_existants
    ).select_related('article', 'facture').order_by('article_id', '-date_creation')
    seen_ids = set()
    for ligne in last_lignes:
        if ligne.article_id not in seen_ids:
            seen_ids.add(ligne.article_id)
            # Devise de la facture: CDF = FC, USD = USD
            devise_facture = ligne.facture.devise if ligne.facture else 'CDF'
            derniers_appros[ligne.article_id] = {
                'type_quantite': ligne.type_quantite,
                'nombre_cartons': ligne.nombre_cartons,
                'pieces_par_carton': ligne.pieces_par_carton,
                'prix_achat_carton': float(ligne.prix_achat_carton),
                'prix_achat_unitaire': float(ligne.prix_achat_unitaire),
                'devise_saisie': devise_facture,
            }
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Données de la facture
                numero_facture = request.POST.get('numero_facture', '').strip()
                fournisseur_id = request.POST.get('fournisseur_id', '').strip()
                fournisseur_nom = request.POST.get('fournisseur_nom', '').strip()
                date_facture_str = request.POST.get('date_facture', '')
                devise = request.POST.get('devise', 'CDF')
                notes = request.POST.get('notes', '')
                
                if not numero_facture:
                    messages.error(request, "Le numéro de facture est obligatoire")
                    return redirect('inventory:approvisionner_facture_boutique', boutique_id=boutique.id)
                
                # Gérer le fournisseur
                fournisseur = None
                if fournisseur_id:
                    fournisseur = Fournisseur.objects.filter(id=fournisseur_id, commercant=commercant).first()
                elif fournisseur_nom:
                    # Créer un nouveau fournisseur
                    fournisseur, created = Fournisseur.objects.get_or_create(
                        nom=fournisseur_nom,
                        commercant=commercant,
                        defaults={'est_actif': True}
                    )
                
                # Date de facture
                try:
                    date_facture = datetime.strptime(date_facture_str, '%Y-%m-%d').date() if date_facture_str else timezone.now().date()
                except ValueError:
                    date_facture = timezone.now().date()
                
                # Vérifier unicité du numéro de facture pour cette boutique
                if FactureApprovisionnement.objects.filter(numero_facture=numero_facture, depot=boutique).exists():
                    msg = f"Le numéro de facture '{numero_facture}' existe déjà."
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': msg}, status=400)
                    messages.error(request, msg)
                    return redirect('inventory:approvisionner_facture_boutique', boutique_id=boutique.id)
                
                # Créer la facture (utiliser le champ depot même pour une boutique)
                facture = FactureApprovisionnement.objects.create(
                    numero_facture=numero_facture,
                    fournisseur=fournisseur,
                    fournisseur_nom=fournisseur_nom if not fournisseur else '',
                    depot=boutique,  # Boutique stockée dans le champ depot
                    date_facture=date_facture,
                    devise=devise,
                    notes=notes,
                    created_by=request.user.username
                )
                
                # Traiter les lignes d'articles
                articles_json_str = request.POST.get('articles_json', '[]')
                articles_data = json.loads(articles_json_str) if articles_json_str else []
                
                if not articles_data:
                    messages.error(request, "Veuillez ajouter au moins un article à la facture")
                    return redirect('inventory:approvisionner_facture_boutique', boutique_id=boutique.id)
                
                for art_data in articles_data:
                    # Récupérer ou créer l'article
                    article_id = art_data.get('article_id')
                    if article_id:
                        article = Article.objects.get(id=article_id, boutique=boutique)
                    else:
                        # Créer un nouvel article
                        code = art_data.get('code', '').strip()
                        if not code:
                            code = f"ART-{uuid.uuid4().hex[:8].upper()}"
                        
                        categorie_id = art_data.get('categorie_id')
                        categorie_nom = art_data.get('categorie_nom', '').strip()
                        if categorie_id:
                            categorie = Categorie.objects.filter(id=categorie_id).first()
                        elif categorie_nom:
                            categorie, _ = Categorie.objects.get_or_create(
                                nom__iexact=categorie_nom,
                                boutique=boutique,
                                defaults={'nom': categorie_nom}
                            )
                        else:
                            categorie = None
                        
                        # Si facture en USD, l'article sera stocké en CDF (après conversion)
                        devise_article = 'CDF' if devise == 'USD' else devise
                        
                        article = Article.objects.create(
                            code=code,
                            nom=art_data.get('nom', 'Article sans nom'),
                            description=art_data.get('description', ''),
                            devise=devise_article,
                            prix_achat=Decimal(str(art_data.get('prix_achat_unitaire', 0))),
                            prix_vente=Decimal(str(art_data.get('prix_vente', 0))),
                            quantite_stock=0,
                            boutique=boutique,
                            categorie=categorie,
                            est_actif=True
                        )
                    
                    # Données de la ligne
                    type_quantite = art_data.get('type_quantite', 'UNITE')
                    nombre_cartons = int(art_data.get('nombre_cartons', 0))
                    pieces_par_carton = int(art_data.get('pieces_par_carton', 1))
                    pieces_supplementaires = int(art_data.get('pieces_supplementaires', 0))
                    quantite_unites = int(art_data.get('quantite_unites', 0))
                    prix_achat_carton = Decimal(str(art_data.get('prix_achat_carton', 0)))
                    prix_achat_unitaire = Decimal(str(art_data.get('prix_achat_unitaire', 0)))
                    prix_piece_sup = Decimal(str(art_data.get('prix_piece_sup', 0)))
                    prix_vente = Decimal(str(art_data.get('prix_vente', 0)))
                    
                    # Calculs pour cartons (avec pièces supplémentaires)
                    if type_quantite == 'CARTON':
                        if prix_achat_carton > 0 and pieces_par_carton > 0:
                            prix_achat_unitaire = prix_achat_carton / pieces_par_carton
                        if prix_piece_sup == 0 and pieces_supplementaires > 0:
                            prix_piece_sup = prix_achat_unitaire
                        quantite_unites = (nombre_cartons * pieces_par_carton) + pieces_supplementaires
                    
                    # Calcul du prix total
                    if type_quantite == 'CARTON':
                        prix_achat_total = (nombre_cartons * prix_achat_carton) + (pieces_supplementaires * prix_piece_sup)
                    else:
                        prix_achat_total = quantite_unites * prix_achat_unitaire
                    
                    # Créer la ligne d'approvisionnement
                    LigneApprovisionnement.objects.create(
                        facture=facture,
                        article=article,
                        categorie=article.categorie,
                        type_quantite=type_quantite,
                        nombre_cartons=nombre_cartons,
                        pieces_par_carton=pieces_par_carton,
                        quantite_unites=quantite_unites,
                        prix_achat_carton=prix_achat_carton,
                        prix_achat_unitaire=prix_achat_unitaire,
                        prix_achat_total=prix_achat_total,
                        prix_vente_unitaire=prix_vente
                    )
                    
                    # Mettre à jour le stock de l'article
                    stock_avant = article.quantite_stock
                    article.quantite_stock += quantite_unites
                    article.prix_achat = prix_achat_unitaire
                    if prix_vente > 0:
                        article.prix_vente = prix_vente
                    article.save()
                    
                    # Enregistrer le mouvement de stock
                    MouvementStock.objects.create(
                        article=article,
                        type_mouvement='ENTREE',
                        quantite=quantite_unites,
                        stock_avant=stock_avant,
                        stock_apres=article.quantite_stock,
                        commentaire=f"Facture {numero_facture} - {fournisseur.nom if fournisseur else fournisseur_nom}",
                        reference_document=f"FACT-{facture.id}",
                        utilisateur=request.user.username
                    )
                
                # Recalculer le montant total
                facture.calculer_montant_total()
                
                # Si requête AJAX, retourner JSON
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f"Facture {numero_facture} enregistrée avec {len(articles_data)} article(s)",
                        'facture_id': facture.id,
                    })
                
                messages.success(request, f"Facture {numero_facture} enregistrée avec {len(articles_data)} article(s)")
                return redirect('inventory:detail_boutique', boutique_id=boutique.id)
                
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': str(e)}, status=400)
            messages.error(request, f"Erreur: {str(e)}")
    
    context = {
        'boutique': boutique,
        'depot': boutique,  # Alias pour compatibilité avec le template
        'fournisseurs': fournisseurs,
        'categories': categories,
        'articles_existants': articles_existants,
        'derniers_appros_json': json.dumps(derniers_appros),
        'today': timezone.now().date().isoformat(),
    }
    
    return render(request, 'inventory/commercant/approvisionner_facture.html', context)


@login_required
@commercant_required
def liste_factures_depot(request, depot_id):
    """Liste des factures d'approvisionnement d'un dépôt avec recherche."""
    from django.core.paginator import Paginator
    from django.db.models import Prefetch
    from django.shortcuts import redirect

    commercant = request.user.profil_commercant
    
    # Si le dépôt n'existe pas, rediriger vers la page centralisée des factures
    try:
        depot = Boutique.objects.get(id=depot_id, commercant=commercant, est_depot=True)
    except Boutique.DoesNotExist:
        messages.warning(request, f"Le dépôt demandé n'existe pas. Voici toutes vos factures.")
        return redirect('inventory:liste_toutes_factures_commercant')

    # Afficher TOUTES les factures du commerçant (dépôt + boutiques)
    factures = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant
    ).select_related(
        'fournisseur', 'depot'
    ).prefetch_related(
        Prefetch('lignes', queryset=LigneApprovisionnement.objects.select_related('article'))
    )

    # --- Recherche ---
    q = request.GET.get('q', '').strip()
    type_filter = request.GET.get('type', '').strip()  # 'depot' ou 'boutique'
    boutique_filter = request.GET.get('boutique', '').strip()
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    tri = request.GET.get('tri', '-date_facture')

    total_count = factures.count()

    if q:
        factures = factures.filter(
            Q(numero_facture__icontains=q) |
            Q(fournisseur__nom__icontains=q) |
            Q(fournisseur_nom__icontains=q) |
            Q(notes__icontains=q) |
            Q(depot__nom__icontains=q) |
            Q(lignes__article__nom__icontains=q)
        ).distinct()

    # Filtre par type (dépôt vs boutique)
    if type_filter == 'depot':
        factures = factures.filter(depot__est_depot=True)
    elif type_filter == 'boutique':
        factures = factures.filter(depot__est_depot=False)
    
    # Filtre par boutique/dépôt spécifique
    boutique_selectionnee = None
    if boutique_filter:
        try:
            boutique_id = int(boutique_filter)
            boutique_selectionnee = Boutique.objects.filter(
                id=boutique_id, commercant=commercant
            ).first()
            factures = factures.filter(depot_id=boutique_id)
        except ValueError:
            pass

    if date_debut:
        try:
            factures = factures.filter(date_facture__gte=datetime.strptime(date_debut, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_fin:
        try:
            factures = factures.filter(date_facture__lte=datetime.strptime(date_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Tri
    allowed_sorts = {
        '-date_facture': '-date_facture',
        'date_facture': 'date_facture',
        '-montant_total': '-montant_total',
        'montant_total': 'montant_total',
        '-numero_facture': '-numero_facture',
        'numero_facture': 'numero_facture',
    }
    factures = factures.order_by(allowed_sorts.get(tri, '-date_facture'))

    filtered_count = factures.count()

    # Pagination
    paginator = Paginator(factures, 25)
    page_num = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_num)

    # Stats rapides (toutes les factures du commerçant)
    stats = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant
    ).aggregate(
        total_montant=Sum('montant_total'),  # Toutes les factures sont en CDF
        nb_factures=Count('id'),
        nb_factures_depot=Count('id', filter=Q(depot__est_depot=True)),
        nb_factures_boutique=Count('id', filter=Q(depot__est_depot=False)),
    )

    # Fournisseurs distincts pour filtre rapide
    fournisseurs = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant
    ).exclude(
        fournisseur__isnull=True
    ).values_list('fournisseur__nom', flat=True).distinct().order_by('fournisseur__nom')
    
    # Liste des boutiques/dépôts pour filtre
    boutiques_depots = Boutique.objects.filter(commercant=commercant).order_by('-est_depot', 'nom')

    context = {
        'depot': depot,
        'factures': page_obj,
        'page_obj': page_obj,
        'q': q,
        'type_filter': type_filter,
        'boutique_filter': boutique_filter,
        'boutique_selectionnee': boutique_selectionnee,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'tri': tri,
        'total_count': total_count,
        'filtered_count': filtered_count,
        'stats': stats,
        'fournisseurs': list(fournisseurs),
        'boutiques_depots': boutiques_depots,
    }

    return render(request, 'inventory/commercant/liste_factures_depot.html', context)


@login_required
@commercant_required
def liste_toutes_factures_commercant(request):
    """Liste centralisée de TOUTES les factures du commerçant (dépôts + boutiques) avec distinction."""
    from django.core.paginator import Paginator
    from django.db.models import Prefetch
    
    commercant = request.user.profil_commercant
    
    # Récupérer TOUTES les factures du commerçant (dépôts ET boutiques)
    factures = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant
    ).select_related(
        'fournisseur', 'depot'
    ).prefetch_related(
        Prefetch('lignes', queryset=LigneApprovisionnement.objects.select_related('article'))
    )
    
    # --- Filtres ---
    q = request.GET.get('q', '').strip()
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    type_filter = request.GET.get('type', '').strip()  # 'depot' ou 'boutique'
    boutique_filter = request.GET.get('boutique', '').strip()
    tri = request.GET.get('tri', '-date_facture')
    
    total_count = factures.count()
    
    # Filtre par recherche
    if q:
        factures = factures.filter(
            Q(numero_facture__icontains=q) |
            Q(fournisseur__nom__icontains=q) |
            Q(fournisseur_nom__icontains=q) |
            Q(notes__icontains=q) |
            Q(depot__nom__icontains=q) |
            Q(lignes__article__nom__icontains=q)
        ).distinct()
    
    # Filtre par type (dépôt vs boutique)
    if type_filter == 'depot':
        factures = factures.filter(depot__est_depot=True)
    elif type_filter == 'boutique':
        factures = factures.filter(depot__est_depot=False)
    
    # Filtre par boutique/dépôt spécifique
    if boutique_filter:
        try:
            factures = factures.filter(depot_id=int(boutique_filter))
        except ValueError:
            pass
    
    # Filtre par date
    if date_debut:
        try:
            factures = factures.filter(date_facture__gte=datetime.strptime(date_debut, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_fin:
        try:
            factures = factures.filter(date_facture__lte=datetime.strptime(date_fin, '%Y-%m-%d').date())
        except ValueError:
            pass
    
    # Tri
    allowed_sorts = {
        '-date_facture': '-date_facture',
        'date_facture': 'date_facture',
        '-montant_total': '-montant_total',
        'montant_total': 'montant_total',
        '-numero_facture': '-numero_facture',
        'numero_facture': 'numero_facture',
    }
    factures = factures.order_by(allowed_sorts.get(tri, '-date_facture'))
    
    filtered_count = factures.count()
    
    # Pagination
    paginator = Paginator(factures, 25)
    page_num = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_num)
    
    # Stats globales (toutes les factures sont en CDF)
    stats = FactureApprovisionnement.objects.filter(depot__commercant=commercant).aggregate(
        total_montant=Sum('montant_total'),  # Toutes les factures sont en CDF
        nb_factures=Count('id'),
        nb_factures_depot=Count('id', filter=Q(depot__est_depot=True)),
        nb_factures_boutique=Count('id', filter=Q(depot__est_depot=False)),
    )
    
    # Liste des boutiques/dépôts pour filtre
    boutiques_depots = Boutique.objects.filter(commercant=commercant).order_by('-est_depot', 'nom')
    
    # ⭐ Boutique/dépôt sélectionné (pour boutons contextuels)
    boutique_selectionnee = None
    if boutique_filter:
        try:
            boutique_selectionnee = Boutique.objects.filter(
                id=int(boutique_filter), commercant=commercant
            ).first()
        except ValueError:
            pass
    
    # Fournisseurs distincts
    fournisseurs = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant
    ).exclude(
        fournisseur__isnull=True
    ).values_list('fournisseur__nom', flat=True).distinct().order_by('fournisseur__nom')
    
    context = {
        'factures': page_obj,
        'page_obj': page_obj,
        'q': q,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'type_filter': type_filter,
        'boutique_filter': boutique_filter,
        'boutique_selectionnee': boutique_selectionnee,
        'tri': tri,
        'total_count': total_count,
        'filtered_count': filtered_count,
        'stats': stats,
        'fournisseurs': list(fournisseurs),
        'boutiques_depots': boutiques_depots,
    }
    
    return render(request, 'inventory/commercant/liste_toutes_factures.html', context)


@login_required
@commercant_required
def detail_facture_depot(request, depot_id, facture_id):
    """Détail d'une facture d'approvisionnement."""
    commercant = request.user.profil_commercant
    # Récupérer la facture et vérifier qu'elle appartient au commerçant
    facture = get_object_or_404(
        FactureApprovisionnement, 
        id=facture_id, 
        depot__commercant=commercant
    )
    # Utiliser le vrai dépôt de la facture, pas celui dans l'URL
    depot = facture.depot
    
    lignes = facture.lignes.select_related('article', 'categorie')
    
    # Calculer le bénéfice pour chaque ligne
    lignes_avec_benefice = []
    total_benefice = Decimal('0')
    total_achat = Decimal('0')
    total_vente = Decimal('0')
    
    for ligne in lignes:
        prix_achat = ligne.prix_achat_unitaire or Decimal('0')
        prix_vente = ligne.prix_vente_unitaire or Decimal('0')
        qte = ligne.quantite_unites or 0
        
        benefice_unitaire = prix_vente - prix_achat
        benefice_total = benefice_unitaire * qte
        achat_total = prix_achat * qte
        vente_total = prix_vente * qte
        
        # Pourcentage de bénéfice (par rapport au prix d'achat)
        if prix_achat > 0:
            pourcentage = (benefice_unitaire / prix_achat) * 100
        else:
            pourcentage = Decimal('0')
        
        ligne.benefice_unitaire = benefice_unitaire
        ligne.benefice_total = benefice_total
        ligne.benefice_pourcentage = pourcentage
        ligne.vente_total = vente_total
        
        lignes_avec_benefice.append(ligne)
        total_benefice += benefice_total
        total_achat += achat_total
        total_vente += vente_total
    
    # Pourcentage global
    if total_achat > 0:
        pourcentage_global = (total_benefice / total_achat) * 100
    else:
        pourcentage_global = Decimal('0')
    
    context = {
        'depot': depot,
        'facture': facture,
        'lignes': lignes_avec_benefice,
        'total_benefice': total_benefice,
        'total_achat': total_achat,
        'total_vente': total_vente,
        'pourcentage_global': pourcentage_global,
    }
    
    return render(request, 'inventory/commercant/detail_facture_depot.html', context)


@login_required
@commercant_required
def detail_facture_flexible(request, facture_id):
    """
    Détail d'une facture d'approvisionnement (route flexible sans depot_id).
    Meilleure pratique: détecte automatiquement le dépôt/boutique de la facture.
    """
    commercant = request.user.profil_commercant
    
    # Récupérer la facture et vérifier qu'elle appartient au commerçant
    facture = get_object_or_404(
        FactureApprovisionnement.objects.select_related('depot', 'fournisseur'),
        id=facture_id,
        depot__commercant=commercant
    )
    
    depot = facture.depot
    lignes = facture.lignes.select_related('article', 'categorie')
    
    # Calculer le bénéfice pour chaque ligne
    lignes_avec_benefice = []
    total_benefice = Decimal('0')
    total_achat = Decimal('0')
    total_vente = Decimal('0')
    
    for ligne in lignes:
        prix_achat = ligne.prix_achat_unitaire or Decimal('0')
        prix_vente = ligne.prix_vente_unitaire or Decimal('0')
        qte = ligne.quantite_unites or 0
        
        benefice_unitaire = prix_vente - prix_achat
        benefice_total = benefice_unitaire * qte
        achat_total = prix_achat * qte
        vente_total = prix_vente * qte
        
        # Pourcentage de bénéfice (par rapport au prix d'achat)
        if prix_achat > 0:
            pourcentage = (benefice_unitaire / prix_achat) * 100
        else:
            pourcentage = Decimal('0')
        
        ligne.benefice_unitaire = benefice_unitaire
        ligne.benefice_total = benefice_total
        ligne.benefice_pourcentage = pourcentage
        ligne.vente_total = vente_total
        
        lignes_avec_benefice.append(ligne)
        total_benefice += benefice_total
        total_achat += achat_total
        total_vente += vente_total
    
    # Pourcentage global
    if total_achat > 0:
        pourcentage_global = (total_benefice / total_achat) * 100
    else:
        pourcentage_global = Decimal('0')
    
    context = {
        'depot': depot,
        'facture': facture,
        'lignes': lignes_avec_benefice,
        'total_benefice': total_benefice,
        'total_achat': total_achat,
        'total_vente': total_vente,
        'pourcentage_global': pourcentage_global,
    }
    
    return render(request, 'inventory/commercant/detail_facture_depot.html', context)


@login_required
@commercant_required
def supprimer_facture_depot(request, depot_id, facture_id):
    """
    Supprimer une facture d'approvisionnement.
    
    ATTENTION: Cette action supprime uniquement la facture et ses lignes.
    Le stock n'est PAS modifié (les articles restent en stock).
    Pour ajuster le stock, utilisez l'inventaire ou les mouvements de stock.
    """
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Vérifier si la facture existe (évite erreur 404 sur double soumission)
    facture = FactureApprovisionnement.objects.filter(id=facture_id, depot=depot).first()
    
    if not facture:
        # La facture n'existe pas ou a déjà été supprimée
        messages.info(request, 'Cette facture a déjà été supprimée ou n\'existe pas.')
        return redirect('inventory:liste_factures_depot', depot_id=depot_id)
    
    if request.method == 'POST':
        numero = facture.numero_facture
        
        # Supprimer la facture (les lignes sont supprimées en cascade)
        facture.delete()
        
        messages.success(request, f'La facture "{numero}" a été supprimée avec succès.')
        return redirect('inventory:liste_factures_depot', depot_id=depot_id)
    
    # GET: afficher confirmation
    return render(request, 'inventory/commercant/confirmer_suppression_facture.html', {
        'depot': depot,
        'facture': facture,
    })


@login_required
@commercant_required
def modifier_ligne_facture(request, depot_id, facture_id, ligne_id):
    """
    Modifier une ligne de facture d'approvisionnement.
    
    Note: Cette fonction modifie UNIQUEMENT la ligne de facture, pas l'article.
    Une facture est un historique et doit refléter les prix au moment de l'achat.
    """
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    facture = get_object_or_404(FactureApprovisionnement, id=facture_id, depot=depot)
    ligne = get_object_or_404(LigneApprovisionnement, id=ligne_id, facture=facture)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        data = json.loads(request.body)
        
        # Récupérer et valider les nouvelles valeurs
        quantite_unites = int(data.get('quantite_unites', ligne.quantite_unites) or 0)
        prix_achat_unitaire = Decimal(str(data.get('prix_achat_unitaire', ligne.prix_achat_unitaire) or 0))
        prix_vente_unitaire = Decimal(str(data.get('prix_vente_unitaire', ligne.prix_vente_unitaire) or 0))
        
        # Validation des données
        if quantite_unites < 1:
            return JsonResponse({'success': False, 'error': 'La quantité doit être supérieure à 0'}, status=400)
        
        if prix_achat_unitaire < 0:
            return JsonResponse({'success': False, 'error': 'Le prix d\'achat ne peut pas être négatif'}, status=400)
        
        if prix_vente_unitaire < 0:
            return JsonResponse({'success': False, 'error': 'Le prix de vente ne peut pas être négatif'}, status=400)
        
        # Transaction atomique pour garantir la cohérence des données
        from django.db import transaction
        with transaction.atomic():
            # Mettre à jour la ligne de facture uniquement
            ligne.quantite_unites = quantite_unites
            ligne.prix_achat_unitaire = prix_achat_unitaire
            ligne.prix_vente_unitaire = prix_vente_unitaire
            ligne.prix_achat_total = quantite_unites * prix_achat_unitaire
            ligne.save()
            
            # Recalculer le montant total de la facture
            total = sum(l.prix_achat_total for l in facture.lignes.all())
            facture.montant_total = total
            facture.save()
        
        # Calculer le bénéfice pour l'affichage
        benefice_total = (prix_vente_unitaire - prix_achat_unitaire) * quantite_unites
        if prix_achat_unitaire > 0:
            pourcentage = ((prix_vente_unitaire - prix_achat_unitaire) / prix_achat_unitaire) * 100
        else:
            pourcentage = Decimal('0')
        
        return JsonResponse({
            'success': True,
            'ligne': {
                'id': ligne.id,
                'quantite_unites': ligne.quantite_unites,
                'prix_achat_unitaire': float(ligne.prix_achat_unitaire),
                'prix_vente_unitaire': float(ligne.prix_vente_unitaire),
                'prix_achat_total': float(ligne.prix_achat_total),
                'benefice_total': float(benefice_total),
                'benefice_pourcentage': float(pourcentage),
            },
            'facture_total': float(facture.montant_total)
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'}, status=400)
    except (ValueError, InvalidOperation) as e:
        return JsonResponse({'success': False, 'error': f'Valeur invalide: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'}, status=500)


@login_required
@commercant_required  
def api_fournisseurs(request, depot_id):
    """API pour récupérer les fournisseurs."""
    commercant = request.user.profil_commercant
    fournisseurs = Fournisseur.objects.filter(commercant=commercant, est_actif=True).values('id', 'nom', 'contact')
    return JsonResponse({'fournisseurs': list(fournisseurs)})


@login_required
@commercant_required
def creer_fournisseur(request, depot_id):
    """Créer un nouveau fournisseur."""
    if request.method == 'POST':
        commercant = request.user.profil_commercant
        nom = request.POST.get('nom', '').strip()
        contact = request.POST.get('contact', '').strip()
        adresse = request.POST.get('adresse', '').strip()
        
        if not nom:
            return JsonResponse({'success': False, 'error': 'Le nom est obligatoire'})
        
        fournisseur, created = Fournisseur.objects.get_or_create(
            nom=nom,
            commercant=commercant,
            defaults={'contact': contact, 'adresse': adresse}
        )
        
        if not created:
            return JsonResponse({'success': False, 'error': 'Ce fournisseur existe déjà'})
        
        return JsonResponse({
            'success': True,
            'fournisseur': {'id': fournisseur.id, 'nom': fournisseur.nom}
        })
    
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})


@login_required
@commercant_required
def modifier_taux_dollar_depot(request, depot_id):
    """Modifier le taux de change USD vers CDF pour le dépôt."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    if request.method == 'POST':
        try:
            taux = request.POST.get('taux_dollar', '').strip()
            if not taux:
                messages.error(request, "Le taux est obligatoire")
                return redirect('inventory:detail_depot', depot_id=depot.id)
            
            depot.taux_dollar = Decimal(taux)
            depot.save()
            messages.success(request, f"Taux de change mis à jour: 1 USD = {depot.taux_dollar:,.0f} CDF")
        except (ValueError, InvalidOperation):
            messages.error(request, "Taux invalide")
    
    return redirect('inventory:detail_depot', depot_id=depot.id)


# ========== INVENTAIRE ==========

@login_required
@commercant_required
def liste_inventaires(request, depot_id):
    """Liste des inventaires d'un dépôt."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    inventaires = Inventaire.objects.filter(boutique=depot).order_by('-date_creation')
    
    context = {
        'depot': depot,
        'inventaires': inventaires,
    }
    return render(request, 'inventory/commercant/liste_inventaires.html', context)


@login_required
@commercant_required
def nouvel_inventaire(request, depot_id):
    """Créer un nouvel inventaire."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    
    # Vérifier s'il n'y a pas déjà un inventaire en cours
    inventaire_en_cours = Inventaire.objects.filter(boutique=depot, statut='EN_COURS').first()
    if inventaire_en_cours:
        messages.warning(request, f"Un inventaire est déjà en cours ({inventaire_en_cours.reference})")
        return redirect('inventory:detail_inventaire', depot_id=depot.id, inventaire_id=inventaire_en_cours.id)
    
    if request.method == 'POST':
        date_inventaire = request.POST.get('date_inventaire', timezone.now().date())
        notes = request.POST.get('notes', '')
        
        # Créer l'inventaire
        inventaire = Inventaire.objects.create(
            boutique=depot,
            date_inventaire=date_inventaire,
            notes=notes,
            cree_par=request.user,
            statut='EN_COURS'
        )
        
        # Créer les lignes pour tous les articles du dépôt
        articles = Article.objects.filter(boutique=depot, est_actif=True)
        lignes_creees = 0
        for article in articles:
            LigneInventaire.objects.create(
                inventaire=inventaire,
                article=article,
                stock_theorique=article.quantite_stock,
                prix_unitaire=article.prix_achat or 0
            )
            lignes_creees += 1
        
        inventaire.nb_articles = lignes_creees
        inventaire.save()
        
        messages.success(request, f"Inventaire {inventaire.reference} créé avec {lignes_creees} articles")
        return redirect('inventory:saisir_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
    
    context = {
        'depot': depot,
        'nb_articles': Article.objects.filter(boutique=depot, est_actif=True).count(),
        'today': timezone.now().date(),
    }
    return render(request, 'inventory/commercant/nouvel_inventaire.html', context)


@login_required
@commercant_required
def detail_inventaire(request, depot_id, inventaire_id):
    """Détail d'un inventaire."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=depot)
    
    lignes = inventaire.lignes.select_related('article', 'article__categorie').all()
    
    # Statistiques
    lignes_saisies = lignes.filter(stock_physique__isnull=False).count()
    lignes_avec_ecart = lignes.filter(stock_physique__isnull=False).exclude(ecart=0).count()
    ecarts_positifs = lignes.filter(ecart__gt=0)
    ecarts_negatifs = lignes.filter(ecart__lt=0)
    
    context = {
        'depot': depot,
        'inventaire': inventaire,
        'lignes': lignes,
        'lignes_saisies': lignes_saisies,
        'lignes_avec_ecart': lignes_avec_ecart,
        'ecarts_positifs': ecarts_positifs,
        'ecarts_negatifs': ecarts_negatifs,
        'total_ecart_positif': sum(l.valeur_ecart for l in ecarts_positifs),
        'total_ecart_negatif': abs(sum(l.valeur_ecart for l in ecarts_negatifs)),
    }
    return render(request, 'inventory/commercant/detail_inventaire.html', context)


@login_required
@commercant_required
def saisir_inventaire(request, depot_id, inventaire_id):
    """Saisir les quantités physiques de l'inventaire."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=depot, statut='EN_COURS')
    
    if request.method == 'POST':
        lignes_mises_a_jour = 0
        for key, value in request.POST.items():
            if key.startswith('stock_physique_'):
                ligne_id = key.replace('stock_physique_', '')
                try:
                    ligne = LigneInventaire.objects.get(id=ligne_id, inventaire=inventaire)
                    if value.strip():
                        ligne.stock_physique = int(value)
                        commentaire_key = f'commentaire_{ligne_id}'
                        if commentaire_key in request.POST:
                            ligne.commentaire = request.POST[commentaire_key]
                        ligne.save()
                        lignes_mises_a_jour += 1
                except (LigneInventaire.DoesNotExist, ValueError):
                    pass
        
        messages.success(request, f"{lignes_mises_a_jour} lignes mises à jour")
        
        # Recalculer les statistiques
        inventaire.calculer_statistiques()
        
        if 'continuer' in request.POST:
            return redirect('inventory:saisir_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
        return redirect('inventory:detail_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
    
    # Filtres
    filtre = request.GET.get('filtre', 'non_saisis')
    categorie_id = request.GET.get('categorie')
    recherche = request.GET.get('q', '')
    
    lignes = inventaire.lignes.select_related('article', 'article__categorie')
    
    if filtre == 'non_saisis':
        lignes = lignes.filter(stock_physique__isnull=True)
    elif filtre == 'saisis':
        lignes = lignes.filter(stock_physique__isnull=False)
    elif filtre == 'ecarts':
        lignes = lignes.filter(stock_physique__isnull=False).exclude(ecart=0)
    
    if categorie_id:
        lignes = lignes.filter(article__categorie_id=categorie_id)
    
    if recherche:
        lignes = lignes.filter(
            Q(article__nom__icontains=recherche) |
            Q(article__code__icontains=recherche)
        )
    
    categories = Categorie.objects.filter(boutique=depot)
    
    context = {
        'depot': depot,
        'inventaire': inventaire,
        'lignes': lignes[:50],
        'total_lignes': lignes.count(),
        'categories': categories,
        'filtre': filtre,
        'categorie_id': categorie_id,
        'recherche': recherche,
        'nb_non_saisis': inventaire.lignes.filter(stock_physique__isnull=True).count(),
        'nb_saisis': inventaire.lignes.filter(stock_physique__isnull=False).count(),
    }
    return render(request, 'inventory/commercant/saisir_inventaire.html', context)


@login_required
@commercant_required
def terminer_inventaire(request, depot_id, inventaire_id):
    """Terminer un inventaire (clôturer la saisie)."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=depot, statut='EN_COURS')
    
    if request.method == 'POST':
        nb_saisis = inventaire.lignes.filter(stock_physique__isnull=False).count()
        if nb_saisis == 0:
            messages.error(request, "Vous devez saisir au moins un article avant de terminer")
            return redirect('inventory:saisir_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
        
        inventaire.statut = 'TERMINE'
        inventaire.date_cloture = timezone.now()
        inventaire.calculer_statistiques()
        inventaire.save()
        
        messages.success(request, f"Inventaire {inventaire.reference} terminé")
        return redirect('inventory:detail_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
    
    return redirect('inventory:detail_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)


@login_required
@commercant_required
def regulariser_inventaire(request, depot_id, inventaire_id):
    """Régulariser le stock selon l'inventaire."""
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant, est_depot=True)
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=depot, statut='TERMINE')
    
    if request.method == 'POST':
        with transaction.atomic():
            lignes_regularisees = 0
            
            for ligne in inventaire.lignes.filter(stock_physique__isnull=False, est_regularise=False):
                if ligne.ecart != 0:
                    article = ligne.article
                    stock_avant = article.quantite_stock
                    
                    article.quantite_stock = ligne.stock_physique
                    article.save()
                    
                    type_mvt = 'ENTREE' if ligne.ecart > 0 else 'SORTIE'
                    MouvementStock.objects.create(
                        article=article,
                        type_mouvement=type_mvt,
                        quantite=abs(ligne.ecart),
                        stock_avant=stock_avant,
                        stock_apres=article.quantite_stock,
                        commentaire=f"Régularisation inventaire {inventaire.reference}",
                        reference_document=f"INV-{inventaire.id}"
                    )
                    
                    ligne.est_regularise = True
                    ligne.save()
                    lignes_regularisees += 1
            
            inventaire.statut = 'REGULARISE'
            inventaire.date_regularisation = timezone.now()
            inventaire.save()
            
            messages.success(request, f"Inventaire régularisé: {lignes_regularisees} articles ajustés")
        
        return redirect('inventory:detail_inventaire', depot_id=depot.id, inventaire_id=inventaire.id)
    
    lignes_ecarts = inventaire.lignes.filter(
        stock_physique__isnull=False,
        est_regularise=False
    ).exclude(ecart=0).select_related('article')
    
    context = {
        'depot': depot,
        'inventaire': inventaire,
        'lignes_ecarts': lignes_ecarts,
        'nb_ecarts': lignes_ecarts.count(),
        'total_positif': sum(l.valeur_ecart for l in lignes_ecarts if l.ecart > 0),
        'total_negatif': abs(sum(l.valeur_ecart for l in lignes_ecarts if l.ecart < 0)),
    }
    return render(request, 'inventory/commercant/regulariser_inventaire.html', context)


# ========== INVENTAIRE BOUTIQUES ==========

@login_required
@commercant_required
@boutique_access_required
def liste_inventaires_boutique(request, boutique_id):
    """Liste des inventaires d'une boutique."""
    boutique = request.boutique
    commercant = request.user.profil_commercant
    inventaires = Inventaire.objects.filter(boutique=boutique).order_by('-date_creation')
    
    # Compter les collaborateurs actifs pour afficher un message si nécessaire
    from .models import Collaborateur
    nb_collaborateurs = Collaborateur.objects.filter(commercant=commercant, est_actif=True).count()
    
    context = {
        'boutique': boutique,
        'inventaires': inventaires,
        'nb_collaborateurs': nb_collaborateurs,
    }
    return render(request, 'inventory/commercant/liste_inventaires_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def nouvel_inventaire_boutique(request, boutique_id):
    """Créer un nouvel inventaire pour une boutique."""
    boutique = request.boutique
    
    inventaire_en_cours = Inventaire.objects.filter(boutique=boutique, statut='EN_COURS').first()
    if inventaire_en_cours:
        messages.warning(request, f"Un inventaire est déjà en cours ({inventaire_en_cours.reference})")
        return redirect('inventory:detail_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire_en_cours.id)
    
    if request.method == 'POST':
        date_inventaire = request.POST.get('date_inventaire', timezone.now().date())
        notes = request.POST.get('notes', '')
        
        inventaire = Inventaire.objects.create(
            boutique=boutique,
            date_inventaire=date_inventaire,
            notes=notes,
            cree_par=request.user,
            statut='EN_COURS'
        )
        
        articles = Article.objects.filter(boutique=boutique, est_actif=True)
        lignes_creees = 0
        for article in articles:
            LigneInventaire.objects.create(
                inventaire=inventaire,
                article=article,
                stock_theorique=article.quantite_stock,
                prix_unitaire=article.prix_achat or 0
            )
            lignes_creees += 1
        
        inventaire.nb_articles = lignes_creees
        inventaire.save()
        
        messages.success(request, f"Inventaire {inventaire.reference} créé avec {lignes_creees} articles")
        return redirect('inventory:saisir_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)
    
    context = {
        'boutique': boutique,
        'nb_articles': Article.objects.filter(boutique=boutique, est_actif=True).count(),
        'today': timezone.now().date(),
    }
    return render(request, 'inventory/commercant/nouvel_inventaire_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def detail_inventaire_boutique(request, boutique_id, inventaire_id):
    """Détail d'un inventaire de boutique."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique)
    
    lignes = inventaire.lignes.select_related('article', 'article__categorie').all()
    lignes_saisies = lignes.filter(stock_physique__isnull=False).count()
    lignes_avec_ecart = lignes.filter(stock_physique__isnull=False).exclude(ecart=0).count()
    ecarts_positifs = lignes.filter(ecart__gt=0)
    ecarts_negatifs = lignes.filter(ecart__lt=0)
    
    context = {
        'boutique': boutique,
        'inventaire': inventaire,
        'lignes': lignes,
        'lignes_saisies': lignes_saisies,
        'lignes_avec_ecart': lignes_avec_ecart,
        'ecarts_positifs': ecarts_positifs,
        'ecarts_negatifs': ecarts_negatifs,
        'total_ecart_positif': sum(l.valeur_ecart for l in ecarts_positifs),
        'total_ecart_negatif': abs(sum(l.valeur_ecart for l in ecarts_negatifs)),
    }
    return render(request, 'inventory/commercant/detail_inventaire_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def saisir_inventaire_boutique(request, boutique_id, inventaire_id):
    """Saisir les quantités physiques de l'inventaire d'une boutique."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique, statut='EN_COURS')
    
    if request.method == 'POST':
        lignes_mises_a_jour = 0
        prix_modifies = 0
        
        # Traiter les modifications de prix de vente
        for key, value in request.POST.items():
            if key.startswith('prix_vente_'):
                ligne_id = key.replace('prix_vente_', '')
                try:
                    ligne = LigneInventaire.objects.get(id=ligne_id, inventaire=inventaire)
                    nouveau_prix = Decimal(value.strip())
                    if nouveau_prix != ligne.article.prix_vente and nouveau_prix > 0:
                        ligne.article.prix_vente = nouveau_prix
                        ligne.article.save(update_fields=['prix_vente'])
                        prix_modifies += 1
                except (LigneInventaire.DoesNotExist, ValueError, InvalidOperation):
                    pass
        
        # Nom d'affichage du collaborateur connecté
        if hasattr(request.user, 'profil_collaborateur'):
            user_nom_batch = request.user.profil_collaborateur.nom_complet
        else:
            user_nom_batch = request.user.get_full_name() or request.user.username

        # Traiter les stocks et autres données
        for key, value in request.POST.items():
            if key.startswith('stock_physique_'):
                ligne_id = key.replace('stock_physique_', '')
                try:
                    ligne = LigneInventaire.objects.get(id=ligne_id, inventaire=inventaire)
                    if value.strip():
                        ligne.stock_physique = int(value)
                        commentaire_key = f'commentaire_{ligne_id}'
                        if commentaire_key in request.POST:
                            ligne.commentaire = request.POST[commentaire_key]

                        # Traçabilité : enregistrer qui a saisi et quand
                        ligne.saisi_par = request.user
                        ligne.assigne_a = user_nom_batch
                        ligne.date_modification = timezone.now()

                        ligne.save()
                        lignes_mises_a_jour += 1
                except (LigneInventaire.DoesNotExist, ValueError):
                    pass
        
        msg_parts = []
        if lignes_mises_a_jour > 0:
            msg_parts.append(f"{lignes_mises_a_jour} lignes mises à jour")
        if prix_modifies > 0:
            msg_parts.append(f"{prix_modifies} prix modifiés")
        
        if msg_parts:
            messages.success(request, " | ".join(msg_parts))
        
        inventaire.calculer_statistiques()
        
        if 'continuer' in request.POST:
            return redirect('inventory:saisir_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)
        return redirect('inventory:detail_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)
    
    filtre = request.GET.get('filtre', 'non_saisis')
    categorie_id = request.GET.get('categorie')
    recherche = request.GET.get('q', '')
    assigne_a_filtre = request.GET.get('assigne_a', '')
    
    lignes = inventaire.lignes.select_related('article', 'article__categorie')
    
    if filtre == 'non_saisis':
        lignes = lignes.filter(stock_physique__isnull=True)
    elif filtre == 'saisis':
        lignes = lignes.filter(stock_physique__isnull=False)
    elif filtre == 'ecarts':
        lignes = lignes.filter(stock_physique__isnull=False).exclude(ecart=0)
    
    if categorie_id:
        lignes = lignes.filter(article__categorie_id=categorie_id)
    
    if recherche:
        lignes = lignes.filter(
            Q(article__nom__icontains=recherche) |
            Q(article__code__icontains=recherche)
        )
    
    # Filtre par collaborateur (assigne_a auto-rempli au moment de la saisie)
    if assigne_a_filtre:
        if assigne_a_filtre == 'NON_SAISI':
            lignes = lignes.filter(Q(assigne_a='') | Q(assigne_a__isnull=True))
        else:
            lignes = lignes.filter(assigne_a=assigne_a_filtre)

    categories = Categorie.objects.filter(boutique=boutique)

    # Liste des collaborateurs uniques ayant effectué des saisies
    employes_list = (
        inventaire.lignes
        .filter(stock_physique__isnull=False)
        .exclude(assigne_a='').exclude(assigne_a__isnull=True)
        .values_list('assigne_a', flat=True)
        .distinct().order_by('assigne_a')
    )

    # Nom d'affichage de l'utilisateur connecté
    if hasattr(request.user, 'profil_collaborateur'):
        current_user_nom = request.user.profil_collaborateur.nom_complet
    else:
        current_user_nom = request.user.get_full_name() or request.user.username
    
    # Données complètes pour l'autocomplete de la saisie rapide (toutes les lignes, sans filtre)
    all_articles_json = json.dumps([
        {
            'ligne_id': l.id,
            'article_id': l.article.id,
            'nom': l.article.nom,
            'code': l.article.code,
            'stock_theorique': l.stock_theorique,
            'stock_physique': l.stock_physique,
            'assigne_a': l.assigne_a or '',
            'quantite_stock': l.article.quantite_stock,
            'prix_vente': float(l.article.prix_vente),
            'devise': l.article.devise,
            'categorie_id': l.article.categorie_id,
            'categorie_nom': l.article.categorie.nom if l.article.categorie else '',
        }
        for l in inventaire.lignes.select_related('article', 'article__categorie').order_by('article__nom')
    ], ensure_ascii=False)

    categories_json = json.dumps([
        {'id': c.id, 'nom': c.nom}
        for c in Categorie.objects.filter(boutique=boutique).order_by('nom')
    ], ensure_ascii=False)

    is_collaborateur = hasattr(request.user, 'profil_collaborateur')

    context = {
        'boutique': boutique,
        'inventaire': inventaire,
        'lignes': lignes[:50],
        'total_lignes': lignes.count(),
        'categories': categories,
        'filtre': filtre,
        'categorie_id': categorie_id,
        'recherche': recherche,
        'assigne_a_filtre': assigne_a_filtre,
        'employes_list': employes_list,
        'current_user_nom': current_user_nom,
        'is_collaborateur': is_collaborateur,
        'nb_non_saisis': inventaire.lignes.filter(stock_physique__isnull=True).count(),
        'nb_saisis': inventaire.lignes.filter(stock_physique__isnull=False).count(),
        'all_articles_json': all_articles_json,
        'categories_json': categories_json,
        'ajax_url': f"/commercant/boutiques/{boutique.id}/inventaires/{inventaire.id}/saisir-ligne/",
    }
    return render(request, 'inventory/commercant/saisir_inventaire_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
@require_POST
def saisir_ligne_inventaire_ajax(request, boutique_id, inventaire_id):
    """AJAX : sauvegarder une seule ligne d'inventaire (saisie rapide)."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique, statut='EN_COURS')

    try:
        data = json.loads(request.body)
        ligne_id = int(data.get('ligne_id'))
        stock_physique = int(data.get('stock_physique'))
    except (ValueError, TypeError, KeyError):
        return JsonResponse({'success': False, 'error': 'Données invalides'}, status=400)

    try:
        ligne = LigneInventaire.objects.select_related('article', 'article__categorie').get(
            id=ligne_id, inventaire=inventaire
        )
    except LigneInventaire.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Ligne introuvable'}, status=404)

    # Nom d'affichage du collaborateur connecté
    if hasattr(request.user, 'profil_collaborateur'):
        user_nom_ajax = request.user.profil_collaborateur.nom_complet
    else:
        user_nom_ajax = request.user.get_full_name() or request.user.username

    ligne.stock_physique = stock_physique
    ligne.saisi_par = request.user
    ligne.assigne_a = user_nom_ajax
    ligne.date_modification = timezone.now()
    ligne.save()

    article_fields_updated = []

    # Mise à jour prix de vente
    nouveau_prix = data.get('prix_vente')
    if nouveau_prix is not None:
        try:
            nouveau_prix = Decimal(str(nouveau_prix))
            if nouveau_prix > 0 and nouveau_prix != ligne.article.prix_vente:
                ligne.article.prix_vente = nouveau_prix
                article_fields_updated.append('prix_vente')
        except Exception:
            pass

    # Mise à jour catégorie
    nouvelle_cat_id = data.get('categorie_id')
    if nouvelle_cat_id is not None:
        try:
            nouvelle_cat = Categorie.objects.get(id=int(nouvelle_cat_id), boutique=boutique)
            if ligne.article.categorie_id != nouvelle_cat.id:
                ligne.article.categorie = nouvelle_cat
                article_fields_updated.append('categorie')
        except Categorie.DoesNotExist:
            pass

    if article_fields_updated:
        ligne.article.save(update_fields=article_fields_updated)

    inventaire.calculer_statistiques()

    ecart = stock_physique - (ligne.stock_theorique or 0)
    nb_saisis = inventaire.lignes.filter(stock_physique__isnull=False).count()
    nb_total = inventaire.lignes.count()

    return JsonResponse({
        'success': True,
        'ligne_id': ligne.id,
        'article_nom': ligne.article.nom,
        'article_id': ligne.article.id,
        'stock_theorique': ligne.stock_theorique,
        'stock_physique': stock_physique,
        'prix_vente': float(ligne.article.prix_vente),
        'categorie_nom': ligne.article.categorie.nom if ligne.article.categorie else '',
        'ecart': ecart,
        'nb_saisis': nb_saisis,
        'nb_total': nb_total,
        'saisi_par_nom': user_nom_ajax,
    })


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 1 : VERROUILLAGE D'ARTICLES EN COURS DE SAISIE (cache Django)
# ─────────────────────────────────────────────────────────────────────────────

def _user_nom(user):
    if hasattr(user, 'profil_collaborateur'):
        return user.profil_collaborateur.nom_complet
    return user.get_full_name() or user.username


@login_required
@commercant_required
@boutique_access_required
@require_POST
def lock_ligne_inventaire_ajax(request, boutique_id, inventaire_id):
    """Verrouille une ligne pour la session de l'utilisateur courant."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique)
    try:
        data = json.loads(request.body)
        ligne_id = int(data.get('ligne_id'))
    except (ValueError, TypeError, KeyError):
        return JsonResponse({'error': 'Données invalides'}, status=400)

    cache_key = f'inv_lock_{inventaire_id}_{ligne_id}'
    existing = cache.get(cache_key)

    if existing and existing['user_id'] != request.user.id:
        return JsonResponse({'locked_by_other': True, 'nom': existing['nom']})

    cache.set(cache_key, {'nom': _user_nom(request.user), 'user_id': request.user.id}, 300)
    return JsonResponse({'locked_by_other': False})


@login_required
@commercant_required
@boutique_access_required
@require_POST
def unlock_ligne_inventaire_ajax(request, boutique_id, inventaire_id):
    """Libère le verrou de la ligne pour l'utilisateur courant."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique)
    try:
        data = json.loads(request.body)
        ligne_id = int(data.get('ligne_id'))
    except (ValueError, TypeError, KeyError):
        return JsonResponse({'error': 'Données invalides'}, status=400)

    cache_key = f'inv_lock_{inventaire_id}_{ligne_id}'
    existing = cache.get(cache_key)
    if existing and existing['user_id'] == request.user.id:
        cache.delete(cache_key)
    return JsonResponse({'ok': True})


@login_required
@commercant_required
@boutique_access_required
def get_locks_inventaire_ajax(request, boutique_id, inventaire_id):
    """Retourne tous les verrous actifs pour cet inventaire."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique)
    ligne_ids = list(inventaire.lignes.values_list('id', flat=True))
    locks = {}
    for lid in ligne_ids:
        lock = cache.get(f'inv_lock_{inventaire_id}_{lid}')
        if lock and lock['user_id'] != request.user.id:
            locks[str(lid)] = lock['nom']
    return JsonResponse({'locks': locks})


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 5 : TABLEAU DE BORD RESPONSABLE EN TEMPS RÉEL
# ─────────────────────────────────────────────────────────────────────────────

@login_required
@commercant_required
@boutique_access_required
def tableau_bord_inventaire(request, boutique_id, inventaire_id):
    """Tableau de bord temps réel de l'avancement par collaborateur."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique)

    nb_total = inventaire.lignes.count()
    nb_saisis = inventaire.lignes.filter(stock_physique__isnull=False).count()

    # Regrouper les saisies par collaborateur
    collab_stats = (
        inventaire.lignes
        .filter(stock_physique__isnull=False)
        .exclude(assigne_a='').exclude(assigne_a__isnull=True)
        .values('assigne_a')
        .annotate(
            nb_saisis=Count('id'),
            derniere_saisie=Max('date_modification'),
        )
        .order_by('-nb_saisis')
    )

    # Valeur totale saisie par collaborateur (en CDF)
    collab_list = []
    for cs in collab_stats:
        lignes_collab = inventaire.lignes.filter(
            assigne_a=cs['assigne_a'],
            stock_physique__isnull=False
        ).select_related('article')
        valeur_cdf = sum(
            (l.stock_physique or 0) * float(l.article.prix_vente)
            for l in lignes_collab
            if l.article.devise != 'USD'
        )
        valeur_usd = sum(
            (l.stock_physique or 0) * float(l.article.prix_vente)
            for l in lignes_collab
            if l.article.devise == 'USD'
        )
        nb_ecarts = lignes_collab.filter(
            stock_physique__isnull=False
        ).exclude(stock_physique=F('stock_theorique')).count()

        collab_list.append({
            'nom': cs['assigne_a'],
            'nb_saisis': cs['nb_saisis'],
            'derniere_saisie': cs['derniere_saisie'],
            'valeur_cdf': valeur_cdf,
            'valeur_usd': valeur_usd,
            'nb_ecarts': nb_ecarts,
        })

    # Articles non encore saisis
    non_saisis = inventaire.lignes.filter(stock_physique__isnull=True).select_related('article')

    context = {
        'boutique': boutique,
        'inventaire': inventaire,
        'nb_total': nb_total,
        'nb_saisis': nb_saisis,
        'nb_non_saisis': nb_total - nb_saisis,
        'pct': round(nb_saisis / nb_total * 100) if nb_total else 0,
        'collab_list': collab_list,
        'non_saisis': non_saisis[:20],
        'nb_non_saisis_total': nb_total - nb_saisis,
    }
    return render(request, 'inventory/commercant/tableau_bord_inventaire.html', context)


@login_required
@commercant_required
@boutique_access_required
def terminer_inventaire_boutique(request, boutique_id, inventaire_id):
    """Terminer un inventaire de boutique."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique, statut='EN_COURS')
    
    if request.method == 'POST':
        nb_saisis = inventaire.lignes.filter(stock_physique__isnull=False).count()
        if nb_saisis == 0:
            messages.error(request, "Vous devez saisir au moins un article avant de terminer")
            return redirect('inventory:saisir_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)
        
        inventaire.statut = 'TERMINE'
        inventaire.date_cloture = timezone.now()
        inventaire.calculer_statistiques()
        inventaire.save()
        
        messages.success(request, f"Inventaire {inventaire.reference} terminé")
    
    return redirect('inventory:detail_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)


@login_required
@commercant_required
@boutique_access_required
def regulariser_inventaire_boutique(request, boutique_id, inventaire_id):
    """Régulariser le stock d'une boutique selon l'inventaire."""
    boutique = request.boutique
    inventaire = get_object_or_404(Inventaire, id=inventaire_id, boutique=boutique, statut='TERMINE')
    
    if request.method == 'POST':
        with transaction.atomic():
            lignes_regularisees = 0
            
            for ligne in inventaire.lignes.filter(stock_physique__isnull=False, est_regularise=False):
                if ligne.ecart != 0:
                    article = ligne.article
                    stock_avant = article.quantite_stock
                    
                    article.quantite_stock = ligne.stock_physique
                    article.save()
                    
                    type_mvt = 'ENTREE' if ligne.ecart > 0 else 'SORTIE'
                    MouvementStock.objects.create(
                        article=article,
                        type_mouvement=type_mvt,
                        quantite=abs(ligne.ecart),
                        stock_avant=stock_avant,
                        stock_apres=article.quantite_stock,
                        commentaire=f"Régularisation inventaire {inventaire.reference}",
                        reference_document=f"INV-{inventaire.id}"
                    )
                    
                    ligne.est_regularise = True
                    ligne.save()
                    lignes_regularisees += 1
            
            inventaire.statut = 'REGULARISE'
            inventaire.date_regularisation = timezone.now()
            inventaire.save()
            
            messages.success(request, f"Inventaire régularisé: {lignes_regularisees} articles ajustés")
        
        return redirect('inventory:detail_inventaire_boutique', boutique_id=boutique.id, inventaire_id=inventaire.id)
    
    lignes_ecarts = inventaire.lignes.filter(
        stock_physique__isnull=False,
        est_regularise=False
    ).exclude(ecart=0).select_related('article')
    
    context = {
        'boutique': boutique,
        'inventaire': inventaire,
        'lignes_ecarts': lignes_ecarts,
        'nb_ecarts': lignes_ecarts.count(),
        'total_positif': sum(l.valeur_ecart for l in lignes_ecarts if l.ecart > 0),
        'total_negatif': abs(sum(l.valeur_ecart for l in lignes_ecarts if l.ecart < 0)),
    }
    return render(request, 'inventory/commercant/regulariser_inventaire_boutique.html', context)


@login_required
@commercant_required
@boutique_access_required
def transfert_entre_boutiques(request, boutique_id):
    """Transférer des articles d'un point de vente vers un autre"""
    commercant = request.user.profil_commercant
    boutique_source = request.boutique
    
    # Récupérer les autres boutiques du commerçant (pas la source)
    autres_boutiques = Boutique.objects.filter(
        commercant=commercant,
        est_depot=False
    ).exclude(id=boutique_source.id).order_by('nom')
    
    # Articles avec stock disponible
    articles = Article.objects.filter(
        boutique=boutique_source,
        est_actif=True,
        quantite_stock__gt=0
    ).order_by('nom')
    
    if request.method == 'POST':
        boutique_dest_id = request.POST.get('boutique_destination')
        commentaire = request.POST.get('commentaire', '').strip()
        articles_selectionnes = request.POST.getlist('articles_selectionnes')
        
        if not boutique_dest_id:
            messages.error(request, "Veuillez sélectionner un point de vente de destination")
            return redirect('inventory:transfert_entre_boutiques', boutique_id=boutique_source.id)
        
        boutique_dest = get_object_or_404(Boutique, id=boutique_dest_id, commercant=commercant)
        
        if not articles_selectionnes:
            messages.error(request, "Veuillez sélectionner au moins un article à transférer")
            return redirect('inventory:transfert_entre_boutiques', boutique_id=boutique_source.id)
        
        transferts_effectues = []
        erreurs = []
        
        for article_id in articles_selectionnes:
            try:
                quantite = int(request.POST.get(f'quantite_{article_id}', 1))
                article_source = Article.objects.get(id=article_id, boutique=boutique_source)
                
                if quantite <= 0:
                    continue
                
                if quantite > article_source.quantite_stock:
                    erreurs.append(f"{article_source.nom}: quantité demandée ({quantite}) > stock disponible ({article_source.quantite_stock})")
                    continue
                
                # Trouver ou créer l'article dans la boutique de destination
                article_dest, created = Article.objects.get_or_create(
                    boutique=boutique_dest,
                    code=article_source.code,
                    defaults={
                        'nom': article_source.nom,
                        'description': article_source.description,
                        'devise': article_source.devise,
                        'prix_vente': article_source.prix_vente,
                        'prix_achat': article_source.prix_achat,
                        'categorie': None,  # Catégorie peut être différente
                        'quantite_stock': 0,
                        'est_actif': True,
                    }
                )
                
                # Décrémenter le stock source
                stock_source_avant = article_source.quantite_stock
                article_source.quantite_stock -= quantite
                article_source.save()
                
                # Incrémenter le stock destination
                stock_dest_avant = article_dest.quantite_stock
                article_dest.quantite_stock += quantite
                article_dest.save()
                
                # Enregistrer les mouvements de stock
                MouvementStock.objects.create(
                    article=article_source,
                    type_mouvement='SORTIE',
                    quantite=-quantite,
                    stock_avant=stock_source_avant,
                    stock_apres=article_source.quantite_stock,
                    commentaire=f"Transfert vers {boutique_dest.nom}",
                    reference_document=f"TRANS-BTQ-{boutique_source.id}-{boutique_dest.id}",
                    utilisateur=request.user.username
                )
                
                MouvementStock.objects.create(
                    article=article_dest,
                    type_mouvement='ENTREE',
                    quantite=quantite,
                    stock_avant=stock_dest_avant,
                    stock_apres=article_dest.quantite_stock,
                    commentaire=f"Transfert depuis {boutique_source.nom}",
                    reference_document=f"TRANS-BTQ-{boutique_source.id}-{boutique_dest.id}",
                    utilisateur=request.user.username
                )
                
                transferts_effectues.append({
                    'article': article_source.nom,
                    'quantite': quantite,
                    'nouveau': created
                })
                
            except Article.DoesNotExist:
                erreurs.append(f"Article ID {article_id} introuvable")
            except Exception as e:
                erreurs.append(f"Erreur article {article_id}: {str(e)}")
        
        if transferts_effectues:
            nb = len(transferts_effectues)
            messages.success(request, f"✅ {nb} article(s) transféré(s) vers {boutique_dest.nom}")
        
        for erreur in erreurs:
            messages.warning(request, erreur)
        
        return redirect('inventory:commercant_articles_boutique', boutique_id=boutique_source.id)
    
    context = {
        'boutique': boutique_source,
        'articles': articles,
        'autres_boutiques': autres_boutiques,
    }
    
    return render(request, 'inventory/commercant/transfert_entre_boutiques.html', context)


# ===== ALERTES STOCK (RÉGULARISATION) =====

@login_required
def alertes_stock_boutique(request, boutique_id):
    """
    Affiche les alertes de stock pour une boutique.
    Permet au commerçant de voir et régulariser les écarts de stock.
    """
    boutique = get_object_or_404(Boutique, id=boutique_id)
    
    # Vérifier que l'utilisateur a accès à cette boutique (ou est superuser)
    if not request.user.is_superuser:
        try:
            commercant = request.user.profil_commercant
            if boutique not in commercant.boutiques.all():
                return HttpResponseForbidden("Accès non autorisé à cette boutique")
        except:
            return HttpResponseForbidden("Accès non autorisé")
    
    
    # Filtres
    statut_filtre = request.GET.get('statut', 'EN_ATTENTE')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Récupérer les alertes
    alertes = AlerteStock.objects.filter(boutique=boutique).select_related(
        'article', 'variante', 'vente', 'terminal', 'regularise_par'
    ).order_by('-date_creation')
    
    # Appliquer les filtres
    if statut_filtre and statut_filtre != 'TOUS':
        alertes = alertes.filter(statut=statut_filtre)
    
    if date_debut:
        try:
            date_debut_dt = datetime.strptime(date_debut, '%Y-%m-%d')
            alertes = alertes.filter(date_creation__date__gte=date_debut_dt)
        except ValueError:
            pass
    
    if date_fin:
        try:
            date_fin_dt = datetime.strptime(date_fin, '%Y-%m-%d')
            alertes = alertes.filter(date_creation__date__lte=date_fin_dt)
        except ValueError:
            pass
    
    # Statistiques
    stats = {
        'total': AlerteStock.objects.filter(boutique=boutique).count(),
        'en_attente': AlerteStock.objects.filter(boutique=boutique, statut='EN_ATTENTE').count(),
        'regularisees': AlerteStock.objects.filter(boutique=boutique, statut='REGULARISE').count(),
        'ignorees': AlerteStock.objects.filter(boutique=boutique, statut='IGNORE').count(),
    }
    
    # Traitement des actions POST
    if request.method == 'POST':
        action = request.POST.get('action')
        alerte_id = request.POST.get('alerte_id')
        notes = request.POST.get('notes', '')
        
        if alerte_id:
            try:
                alerte = AlerteStock.objects.get(id=alerte_id, boutique=boutique)
                
                if action == 'regulariser':
                    alerte.regulariser(request.user, notes)
                    messages.success(request, f"✅ Alerte régularisée pour {alerte.nom_article_complet}")
                elif action == 'ignorer':
                    alerte.ignorer(request.user, notes)
                    messages.info(request, f"ℹ️ Alerte ignorée pour {alerte.nom_article_complet}")
                elif action == 'ajuster_stock':
                    # Ajuster le stock à 0 si négatif
                    if alerte.variante:
                        if alerte.variante.quantite_stock < 0:
                            alerte.variante.quantite_stock = 0
                            alerte.variante.save()
                    else:
                        if alerte.article.quantite_stock < 0:
                            alerte.article.quantite_stock = 0
                            alerte.article.save()
                    alerte.regulariser(request.user, f"Stock ajusté à 0. {notes}")
                    messages.success(request, f"✅ Stock ajusté à 0 pour {alerte.nom_article_complet}")
                    
            except AlerteStock.DoesNotExist:
                messages.error(request, "Alerte non trouvée")
        
        return redirect('inventory:commercant_alertes_stock', boutique_id=boutique_id)
    
    context = {
        'boutique': boutique,
        'alertes': alertes[:100],  # Limiter à 100 alertes
        'stats': stats,
        'statut_filtre': statut_filtre,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    
    return render(request, 'inventory/commercant/alertes_stock.html', context)


# ===== ANALYSE IA DES MOUVEMENTS =====

@login_required
def analyse_ia_mouvements(request, boutique_id):
    """
    🧠 Dashboard d'analyse intelligente des mouvements de stock.
    Calcule le score de santé, les insights, les anomalies et les recommandations.
    """
    from django.db.models import ExpressionWrapper, DecimalField

    boutique = get_object_or_404(Boutique, id=boutique_id)
    if not request.user.is_superuser:
        try:
            commercant = request.user.profil_commercant
            if boutique not in commercant.boutiques.all():
                return HttpResponseForbidden("Accès non autorisé à cette boutique")
        except Exception:
            return HttpResponseForbidden("Accès non autorisé")

    # --- POST : régularisation rapide ---
    if request.method == 'POST':
        article_id = request.POST.get('article_id')
        quantite_ajout = request.POST.get('quantite_ajout', '0')
        notes = request.POST.get('notes', 'Régularisation manuelle')
        try:
            article = Article.objects.get(id=article_id, boutique=boutique)
            qte = int(quantite_ajout)
            if qte > 0:
                stock_avant = article.quantite_stock
                article.quantite_stock += qte
                article.save(update_fields=['quantite_stock'])
                MouvementStock.objects.create(
                    article=article,
                    type_mouvement='ENTREE',
                    quantite=qte,
                    stock_avant=stock_avant,
                    stock_apres=article.quantite_stock,
                    reference_document=f"REGUL-WEB-{timezone.now().strftime('%Y%m%d%H%M%S')}",
                    utilisateur=request.user.username,
                    commentaire=f"Régularisation web: {notes}"
                )
                AlerteStock.objects.filter(
                    article=article, boutique=boutique, statut='EN_ATTENTE'
                ).update(
                    statut='REGULARISE',
                    date_regularisation=timezone.now(),
                    regularise_par=request.user,
                    notes_regularisation=notes
                )
                messages.success(request, f"✅ Stock de « {article.nom} » ajusté : +{qte} unités")
        except (Article.DoesNotExist, ValueError):
            messages.error(request, "Erreur lors de la régularisation")
        return redirect('inventory:analyse_ia_mouvements', boutique_id=boutique_id)

    # --- Plage de dates ---
    maintenant = timezone.now()
    debut_mois = maintenant.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    debut_str = request.GET.get('debut', '')
    fin_str   = request.GET.get('fin',   '')

    try:
        debut = timezone.make_aware(datetime.strptime(debut_str, '%Y-%m-%d')) if debut_str else debut_mois
    except ValueError:
        debut = debut_mois
    try:
        fin = timezone.make_aware(datetime.strptime(fin_str, '%Y-%m-%d').replace(
            hour=23, minute=59, second=59)) if fin_str else maintenant
    except ValueError:
        fin = maintenant

    nb_jours = max(1, (fin - debut).days + 1)

    # --- QuerySets ---
    articles_qs  = Article.objects.filter(boutique=boutique, est_actif=True)
    mouvements_qs = MouvementStock.objects.filter(
        article__boutique=boutique, date_mouvement__gte=debut, date_mouvement__lte=fin)
    ventes_qs = Vente.objects.filter(
        boutique=boutique, date_vente__gte=debut, date_vente__lte=fin, est_annulee=False)
    alertes_qs = AlerteStock.objects.filter(boutique=boutique, statut='EN_ATTENTE')

    # --- 1. Résumé financier ---
    def ca_devise(devise):
        return float(
            LigneVente.objects.filter(
                vente__boutique=boutique,
                vente__date_vente__gte=debut, vente__date_vente__lte=fin,
                vente__est_annulee=False, devise=devise
            ).aggregate(
                ca=Sum(ExpressionWrapper(F('quantite') * F('prix_unitaire'),
                                         output_field=DecimalField()))
            )['ca'] or 0
        )

    ca_cdf = ca_devise('CDF')
    ca_usd = ca_devise('USD')
    ca_negocie = float(
        LigneVente.objects.filter(
            vente__boutique=boutique,
            vente__date_vente__gte=debut, vente__date_vente__lte=fin,
            vente__est_annulee=False, est_negocie=True
        ).aggregate(
            ca=Sum(ExpressionWrapper(F('quantite') * F('prix_unitaire'),
                                     output_field=DecimalField()))
        )['ca'] or 0
    )

    agg_stock = articles_qs.aggregate(
        val_vente=Sum(ExpressionWrapper(F('quantite_stock') * F('prix_vente'),
                                        output_field=DecimalField())),
        val_cout=Sum(ExpressionWrapper(F('quantite_stock') * F('prix_achat'),
                                       output_field=DecimalField()))
    )
    stock_val_vente = float(agg_stock['val_vente'] or 0)
    stock_val_cout  = float(agg_stock['val_cout'] or 0)

    # --- 2. Mouvements par type ---
    from django.db.models import Count
    mouv_raw = mouvements_qs.values('type_mouvement').annotate(
        nb=Count('id'), total_qte=Sum('quantite'))
    mouv = {m['type_mouvement']: m for m in mouv_raw}

    def gm(typ, abs_val=False):
        d = mouv.get(typ, {})
        q = d.get('total_qte', 0) or 0
        return {'nb': d.get('nb', 0) or 0, 'total_qte': abs(q) if abs_val else q}

    mouvements_resume = {
        'entrees':     gm('ENTREE'),
        'sorties':     gm('SORTIE', True),
        'ventes':      gm('VENTE', True),
        'retours':     gm('RETOUR'),
        'ajustements': gm('AJUSTEMENT'),
    }

    # --- 3. Alertes & stocks négatifs ---
    articles_negatifs = list(
        articles_qs.filter(quantite_stock__lt=0)
        .select_related('categorie').order_by('quantite_stock')[:30]
    )

    alertes_list = list(alertes_qs.select_related('article', 'vente').order_by('-date_creation')[:100])
    alertes_par_art = {}
    for al in alertes_list:
        aid = al.article_id
        if aid not in alertes_par_art:
            alertes_par_art[aid] = {
                'article': al.article,
                'nb_alertes': 0,
                'qte_ecart': 0,
                'ventes': []
            }
        alertes_par_art[aid]['nb_alertes'] += 1
        alertes_par_art[aid]['qte_ecart'] += abs(al.ecart)
        if al.numero_facture not in alertes_par_art[aid]['ventes']:
            alertes_par_art[aid]['ventes'].append(al.numero_facture)

    articles_a_regulariser = sorted(
        alertes_par_art.values(), key=lambda x: x['nb_alertes'], reverse=True)[:20]

    pertes_estimees = sum(
        abs(a.quantite_stock) * float(a.prix_achat) for a in articles_negatifs)
    nb_alertes_total = alertes_qs.count()

    # --- 4. Top articles vendus ---
    top_articles = list(
        LigneVente.objects.filter(
            vente__boutique=boutique,
            vente__date_vente__gte=debut, vente__date_vente__lte=fin,
            vente__est_annulee=False
        )
        .values('article__id', 'article__nom', 'article__code', 'article__devise')
        .annotate(
            total_qte=Sum('quantite'),
            ca=Sum(ExpressionWrapper(F('quantite') * F('prix_unitaire'),
                                     output_field=DecimalField())),
            nb_ventes=Count('vente', distinct=True)
        ).order_by('-ca')[:10]
    )

    # --- 5. Analyse IA ---
    score = 100
    insights = []
    recommandations = []
    anomalies = []

    total_articles = articles_qs.count()
    nb_neg = len(articles_negatifs)

    if total_articles > 0:
        score -= min(40, round(nb_neg / total_articles * 100 * 2))
    score -= min(20, round(nb_alertes_total * 0.5))

    rotation = 0
    if stock_val_vente > 0 and ca_cdf > 0:
        rotation = (ca_cdf / stock_val_vente) * (365 / nb_jours)
        if rotation < 2:
            score -= 10
            insights.append({'type': 'ROTATION_FAIBLE', 'niveau': 'warning',
                'icon': 'fa-sync-alt', 'color': 'warning',
                'message': f"Rotation stock faible : {rotation:.1f}×/an — marchandise immobilisée."})
        elif rotation >= 12:
            insights.append({'type': 'ROTATION_ELEVEE', 'niveau': 'success',
                'icon': 'fa-check-circle', 'color': 'success',
                'message': f"Excellente rotation : {rotation:.1f}×/an — stock bien géré."})

    score = max(0, min(100, score))

    if nb_neg > 0:
        deficit = sum(abs(a.quantite_stock) for a in articles_negatifs)
        insights.append({'type': 'STOCK_NEGATIF', 'niveau': 'danger',
            'icon': 'fa-exclamation-circle', 'color': 'danger',
            'message': f"{nb_neg} article(s) en stock négatif — déficit : {deficit} unités — perte estimée : {round(pertes_estimees):,} CDF"})
        recommandations.append({'priorite': 'HAUTE', 'color': 'danger',
            'icon': 'fa-first-aid',
            'action': 'Réapprovisionner ou ajuster manuellement les articles en déficit.'})

    if ca_negocie > 0 and ca_cdf > 0:
        pct = ca_negocie / ca_cdf * 100
        lvl = 'warning' if pct > 25 else 'info'
        insights.append({'type': 'PRIX_NEGOCIE', 'niveau': lvl,
            'icon': 'fa-tag', 'color': lvl,
            'message': f"{pct:.1f}% du CA CDF vendu à prix négocié — écart avec catalogue."})
        if pct > 30:
            recommandations.append({'priorite': 'MOYENNE', 'color': 'warning',
                'icon': 'fa-shield-alt',
                'action': 'Plus de 30% du CA à prix réduit. Vérifier les autorisations de remise.'})

    if nb_alertes_total > 0:
        insights.append({'type': 'ALERTES_EN_ATTENTE', 'niveau': 'warning',
            'icon': 'fa-bell', 'color': 'warning',
            'message': f"{nb_alertes_total} alerte(s) de stock non régularisée(s)."})
        recommandations.append({'priorite': 'HAUTE', 'color': 'danger',
            'icon': 'fa-tools',
            'action': f"Régulariser les {nb_alertes_total} alerte(s) de sur-vente en attente."})

    nb_ajust = mouv.get('AJUSTEMENT', {}).get('nb', 0) or 0
    if nb_ajust > 10:
        anomalies.append({'color': 'warning', 'icon': 'fa-random',
            'message': f"{nb_ajust} ajustements manuels sur la période — vérifier la cohérence."})

    nb_ventes_periode = ventes_qs.count()
    nb_mouv_vente = mouv.get('VENTE', {}).get('nb', 0) or 0
    if nb_ventes_periode > 0 and nb_mouv_vente == 0:
        anomalies.append({'color': 'danger', 'icon': 'fa-unlink',
            'message': f"{nb_ventes_periode} vente(s) enregistrée(s) mais aucun mouvement VENTE trouvé."})

    niveau_sante = ('EXCELLENT' if score >= 85 else 'BON' if score >= 70
                    else 'MOYEN' if score >= 50 else 'CRITIQUE')
    score_color   = ('success' if score >= 85 else 'info' if score >= 70
                     else 'warning' if score >= 50 else 'danger')

    context = {
        'boutique': boutique,
        'debut': debut.strftime('%Y-%m-%d'),
        'fin': fin.strftime('%Y-%m-%d'),
        'nb_jours': nb_jours,
        'ca_cdf': ca_cdf,
        'ca_usd': ca_usd,
        'ca_negocie': ca_negocie,
        'stock_val_vente': stock_val_vente,
        'stock_val_cout': stock_val_cout,
        'marge_potentielle': max(0, stock_val_vente - stock_val_cout),
        'nb_ventes': nb_ventes_periode,
        'mouvements_resume': mouvements_resume,
        'articles_negatifs': articles_negatifs,
        'articles_a_regulariser': articles_a_regulariser,
        'pertes_estimees': pertes_estimees,
        'nb_alertes_total': nb_alertes_total,
        'top_articles': top_articles,
        'score': score,
        'score_color': score_color,
        'niveau_sante': niveau_sante,
        'rotation': round(rotation, 1),
        'insights': insights,
        'recommandations': recommandations,
        'anomalies': anomalies,
    }
    return render(request, 'inventory/commercant/analyse_ia_mouvements.html', context)


@login_required
@commercant_required
def suivi_articles_recents(request, depot_id):
    """Suivi des factures récemment enregistrées avec leurs détails et coûts."""
    from django.db.models import Sum, Count, Q, Prefetch
    from datetime import datetime, timedelta, date as date_type
    
    commercant = request.user.profil_commercant
    depot = get_object_or_404(Boutique, id=depot_id, commercant=commercant)
    
    # Filtres de date
    date_debut = request.GET.get('date_debut', '').strip()
    date_fin = request.GET.get('date_fin', '').strip()
    
    # Par défaut: aujourd'hui uniquement
    aujourd_hui = datetime.now().date()
    if not date_debut:
        date_debut = aujourd_hui.strftime('%Y-%m-%d')
    if not date_fin:
        date_fin = aujourd_hui.strftime('%Y-%m-%d')
    
    # Convertir en objets date
    try:
        debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
        fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
    except ValueError:
        debut = aujourd_hui
        fin = aujourd_hui
        date_debut = debut.strftime('%Y-%m-%d')
        date_fin = fin.strftime('%Y-%m-%d')
    
    # Récupérer toutes les factures dans la période
    factures = FactureApprovisionnement.objects.filter(
        depot__commercant=commercant,
        date_facture__gte=debut,
        date_facture__lte=fin
    ).select_related(
        'fournisseur', 'depot'
    ).prefetch_related(
        Prefetch('lignes', queryset=LigneApprovisionnement.objects.select_related('article'))
    ).order_by('-date_facture', '-id')
    
    # Préparer les données des factures avec détails
    factures_list = []
    for facture in factures:
        lignes_detail = []
        for ligne in facture.lignes.all():
            article_nom = ligne.article.nom if ligne.article else (ligne.nom_custom or "Article sans nom")
            lignes_detail.append({
                'article_nom': article_nom,
                'quantite': ligne.quantite_unites,
                'prix_achat_unitaire': ligne.prix_achat_unitaire,
                'prix_vente_unitaire': ligne.prix_vente_unitaire,
                'cout_achat': ligne.prix_achat_total,
                'valeur_vente': ligne.quantite_unites * ligne.prix_vente_unitaire,
                'benefice': (ligne.prix_vente_unitaire - ligne.prix_achat_unitaire) * ligne.quantite_unites,
            })
        
        factures_list.append({
            'facture': facture,
            'lignes': lignes_detail,
            'nb_articles': len(lignes_detail),
            'est_aujourd_hui': facture.date_facture == aujourd_hui,
        })
    
    # Statistiques globales
    stats = factures.aggregate(
        nb_factures=Count('id'),
        cout_total=Sum('montant_total'),
    )
    
    # Compter les factures d'aujourd'hui
    stats['nb_factures_aujourd_hui'] = factures.filter(date_facture=aujourd_hui).count()
    
    # Calculer valeur de vente et bénéfice potentiel
    valeur_vente_totale = Decimal('0')
    benefice_potentiel_total = Decimal('0')
    nb_articles_total = 0
    
    for f in factures_list:
        for ligne in f['lignes']:
            valeur_vente_totale += ligne['valeur_vente']
            benefice_potentiel_total += ligne['benefice']
            nb_articles_total += 1
    
    stats['valeur_vente_totale'] = valeur_vente_totale
    stats['benefice_potentiel_total'] = benefice_potentiel_total
    stats['nb_articles_total'] = nb_articles_total
    
    # Calculer le pourcentage de bénéfice global
    if stats['cout_total'] and stats['cout_total'] > 0:
        stats['pourcentage_benefice'] = (stats['benefice_potentiel_total'] / stats['cout_total']) * 100
    else:
        stats['pourcentage_benefice'] = Decimal('0')
    
    context = {
        'depot': depot,
        'factures': factures_list,
        'stats': stats,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'aujourd_hui': aujourd_hui,
    }
    
    return render(request, 'inventory/commercant/suivi_articles_recents.html', context)
